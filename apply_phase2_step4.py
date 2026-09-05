import os
import shutil
from datetime import datetime

print("🚀 Applying Phase 2 - Step 4: Native Longest Path Calculation...")

# 1. Create Backup Folder
timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
backup_dir = f"_backup_phase2_step4_{timestamp}"
os.makedirs(backup_dir, exist_ok=True)
print(f"📦 Created backup folder: {backup_dir}")

files_to_backup = [
    "app.py",
    "data_engine.py",
    "templates/gantt.html",
    "static/gantt.js",
]

for file_path in files_to_backup:
    if os.path.exists(file_path):
        dest = os.path.join(backup_dir, os.path.basename(file_path.replace("/", os.sep)))
        shutil.copy2(file_path, dest)
        print(f"   Backed up {file_path}")


# ==============================================================================
# FILE 1: longest_path_engine.py (NEW - Native CPM Longest Path Calculator)
# ==============================================================================

LONGEST_PATH_ENGINE_CODE = '''"""
LONGEST PATH ENGINE
====================
Computes the true "Longest Path" (Driving Critical Path) using backward-pass
graph traversal — the industry-standard method for forensic delay analysis.

Unlike TF ≤ 0 heuristic (which can be distorted by constraints), Longest Path
traces the actual chain of driving predecessors from project completion 
back to project start.

Algorithm:
1. Identify project finish milestone(s) — activity with latest early_finish + no successors
2. From each finish node, walk backward through predecessors
3. For each activity, select the predecessor with the LATEST early_finish
   (that is the "driving" predecessor)
4. Continue until reaching a start milestone or no more predecessors
5. Mark all activities in the traced chain as "on longest path"
"""

from datetime import datetime
from collections import defaultdict
import logging

logger = logging.getLogger(__name__)


class LongestPathEngine:
    """Native CPM Longest Path Calculator."""

    def __init__(self, engine):
        self.engine = engine
        self.longest_path_ids = set()
        self.longest_path_chain = []
        self.driving_edges = []  # List of (pred_id, succ_id, lag_days) tuples
        self.results = {}

    def calculate(self):
        """Run longest path calculation."""
        logger.info("🎯 Calculating Longest Path (driving critical path)...")
        
        try:
            # Step 1: Find project finish milestone(s)
            finish_nodes = self._find_finish_nodes()
            if not finish_nodes:
                logger.warning("No project finish node found for Longest Path calculation.")
                self.results = {'error': 'No project finish node identified.'}
                return self.results
            
            logger.info(f"  Found {len(finish_nodes)} candidate finish node(s)")
            
            # Step 2: Trace backward from each finish node
            for finish_node in finish_nodes:
                self._trace_backward(finish_node)
            
            # Step 3: Build ordered chain from finish backward to start
            self.longest_path_chain = self._build_ordered_chain(finish_nodes)
            
            # Step 4: Compile results
            self.results = self._compile_results()
            
            logger.info(f"  ✅ Longest Path identified: {len(self.longest_path_ids)} activities")
        except Exception as e:
            logger.exception("Longest Path calculation error: %s", e)
            self.results = {'error': str(e)}
        
        return self.results

    def _find_finish_nodes(self):
        """
        Identify project finish nodes:
        - Activities with no successors (or all successors are external/completed)
        - Prioritize incomplete finish milestones
        - Fall back to latest early_finish activity
        """
        candidates = []
        
        for act in self.engine.activities:
            if act.get('task_type') in ('TT_WBS', 'TT_LOE'):
                continue
            
            task_id = str(act.get('task_id', ''))
            succs = self.engine.successors.get(task_id, [])
            
            # No successors = potential finish node
            if not succs:
                ef = act.get('early_end_date_parsed') or act.get('target_end_date_parsed')
                if ef:
                    candidates.append({
                        'id': task_id,
                        'act': act,
                        'ef': ef,
                        'is_milestone': act.get('task_type') in ('TT_Mile', 'TT_FinMile'),
                        'is_complete': act.get('status_code') == 'TK_Complete',
                    })
        
        if not candidates:
            return []
        
        # Prefer incomplete finish milestones
        incomplete_milestones = [c for c in candidates if c['is_milestone'] and not c['is_complete']]
        if incomplete_milestones:
            # Return the one with latest EF
            best = max(incomplete_milestones, key=lambda c: c['ef'])
            return [best['id']]
        
        # Otherwise, take the single activity with latest EF
        best = max(candidates, key=lambda c: c['ef'])
        return [best['id']]

    def _trace_backward(self, start_node_id):
        """
        Backward pass from start_node_id.
        For each activity, find its DRIVING predecessor (the one with latest EF).
        Add that predecessor to longest_path_ids, then recurse.
        """
        visited = set()
        stack = [start_node_id]
        
        while stack:
            current_id = stack.pop()
            if current_id in visited:
                continue
            visited.add(current_id)
            self.longest_path_ids.add(current_id)
            
            # Get predecessors of current activity
            preds = self.engine.predecessors.get(current_id, [])
            if not preds:
                continue
            
            # Find the "driving" predecessor (latest EF among preds)
            driving_pred = None
            latest_ef = None
            
            for pred in preds:
                pred_id = str(pred.get('task_id', ''))
                pred_act = self.engine.activity_by_id.get(pred_id)
                if not pred_act:
                    continue
                
                pred_ef = pred_act.get('early_end_date_parsed') or pred_act.get('target_end_date_parsed')
                if not pred_ef:
                    continue
                
                if latest_ef is None or pred_ef > latest_ef:
                    latest_ef = pred_ef
                    driving_pred = pred
            
            if driving_pred:
                driving_pred_id = str(driving_pred.get('task_id', ''))
                self.driving_edges.append({
                    'pred_id': driving_pred_id,
                    'succ_id': current_id,
                    'lag_days': driving_pred.get('lag_days', 0),
                    'type': driving_pred.get('type', ''),
                })
                stack.append(driving_pred_id)

    def _build_ordered_chain(self, finish_nodes):
        """
        Return the longest path as an ordered list from start to finish.
        Uses topological ordering based on early_start dates.
        """
        chain = []
        for act_id in self.longest_path_ids:
            act = self.engine.activity_by_id.get(act_id)
            if not act:
                continue
            es = act.get('early_start_date_parsed') or act.get('target_start_date_parsed')
            chain.append({
                'id': act_id,
                'code': act.get('task_code', ''),
                'name': act.get('task_name', ''),
                'wbs': act.get('wbs_name', ''),
                'early_start': es.strftime('%Y-%m-%d') if es else '',
                'early_finish': act.get('early_end_date_parsed').strftime('%Y-%m-%d') if act.get('early_end_date_parsed') else '',
                'duration_days': round(float(act.get('original_duration_days', 0) or 0), 1),
                'total_float_days': round(float(act.get('total_float_days', 0) or 0), 1),
                'status': act.get('status_text', ''),
                'is_milestone': act.get('task_type') in ('TT_Mile', 'TT_FinMile'),
                'is_completed': act.get('status_code') == 'TK_Complete',
                '_sort_date': es or datetime.min,
            })
        
        # Sort chronologically
        chain.sort(key=lambda x: x['_sort_date'])
        for item in chain:
            del item['_sort_date']
        
        return chain

    def _compile_results(self):
        """Build final results dictionary."""
        # Compute total path duration (finish - start of first activity)
        if self.longest_path_chain:
            first_act = self.engine.activity_by_id.get(self.longest_path_chain[0]['id'])
            last_act = self.engine.activity_by_id.get(self.longest_path_chain[-1]['id'])
            path_start = first_act.get('early_start_date_parsed') if first_act else None
            path_end = last_act.get('early_end_date_parsed') if last_act else None
            total_days = (path_end - path_start).days if (path_start and path_end) else 0
        else:
            total_days = 0
        
        # Comparison to TF ≤ 0 critical
        tf_critical_ids = {
            str(a.get('task_id', '')) 
            for a in self.engine.activities 
            if a.get('is_critical') and a.get('task_type') not in ('TT_LOE', 'TT_WBS')
        }
        
        overlap = self.longest_path_ids & tf_critical_ids
        only_in_lp = self.longest_path_ids - tf_critical_ids
        only_in_tf = tf_critical_ids - self.longest_path_ids
        
        # Stats
        real_activity_count = sum(
            1 for a in self.engine.activities 
            if a.get('task_type') not in ('TT_WBS', 'TT_LOE') 
            and a.get('status_code') != 'TK_Complete'
        ) or 1
        
        lp_pct = round(len(self.longest_path_ids) / real_activity_count * 100, 2)
        tf_pct = round(len(tf_critical_ids) / real_activity_count * 100, 2)
        
        return {
            'longest_path_count': len(self.longest_path_ids),
            'longest_path_ids': list(self.longest_path_ids),
            'longest_path_chain': self.longest_path_chain,
            'driving_edges': self.driving_edges,
            'total_path_duration_days': total_days,
            'longest_path_percentage': lp_pct,
            'tf_critical_percentage': tf_pct,
            'overlap_count': len(overlap),
            'only_in_longest_path': len(only_in_lp),
            'only_in_tf_critical': len(only_in_tf),
            'agreement_pct': round(len(overlap) / max(len(self.longest_path_ids), 1) * 100, 2),
        }
'''

with open("longest_path_engine.py", "w", encoding="utf-8") as f:
    f.write(LONGEST_PATH_ENGINE_CODE)
print("  ✅ Created longest_path_engine.py")


# ==============================================================================
# FILE 2: app.py (Add /api/longest-path endpoint)
# ==============================================================================

APP_CODE = '''"""
P6 SCHEDULE ANALYZER - MAIN WEB APPLICATION (app.py)
=====================================================
Now with Longest Path Engine (Phase 2, Step 4)
"""

from flask import Flask, render_template, request, jsonify, send_file, session
from flask_cors import CORS
from werkzeug.utils import secure_filename
import os
import uuid
import time
import logging
from datetime import datetime

try:
    from config import (
        get_config, MAX_UPLOAD_SIZE_MB, SECRET_KEY, SESSION_LIFETIME_HOURS,
    )
except ImportError:
    MAX_UPLOAD_SIZE_MB = 100
    SECRET_KEY = 'dev-only-CHANGE-ME'
    SESSION_LIFETIME_HOURS = 24
    def get_config():
        return {
            'company_name': 'MK Constructions', 'app_title': 'P6 Schedule Analyzer',
            'app_subtitle': 'DCMA 14-Point Check & Analytics',
            'use_logo_image': False, 'theme': {'primary': '#1e40af', 'accent': '#3b82f6'},
            'features': {'gantt': True, 'comparison': True, 'evm': True, 'export': True, 'health': True}
        }

from parser import XERParser
from data_engine import ScheduleEngine
from reports import ReportGenerator

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Optional engines
try:
    from comparison_engine import ScheduleComparator
    logger.info("✅ ScheduleComparator imported")
except Exception as e:
    ScheduleComparator = None
    logger.warning("❌ ScheduleComparator import failed: %s", e)

try:
    from evm_engine import EVMEngine
    logger.info("✅ EVMEngine imported")
except Exception as e:
    EVMEngine = None
    logger.warning("❌ EVMEngine import failed: %s", e)

try:
    from advanced_health_engine import AdvancedHealthEngine
    logger.info("✅ AdvancedHealthEngine imported")
except Exception as e:
    AdvancedHealthEngine = None
    logger.warning("❌ AdvancedHealthEngine import failed: %s", e)

try:
    from pdf_report_generator import PDFReportGenerator
    logger.info("✅ PDFReportGenerator imported")
except Exception as e:
    PDFReportGenerator = None
    logger.warning("❌ PDFReportGenerator import failed: %s", e)

try:
    from ai_narrative_engine import AINarrativeEngine
    logger.info("✅ AINarrativeEngine imported")
except Exception as e:
    AINarrativeEngine = None
    logger.warning("❌ AINarrativeEngine import failed: %s", e)

try:
    from resource_engine import ResourceEngine
    logger.info("✅ ResourceEngine imported")
except Exception as e:
    ResourceEngine = None
    logger.warning("❌ ResourceEngine import failed: %s", e)

try:
    from longest_path_engine import LongestPathEngine
    logger.info("✅ LongestPathEngine imported")
except Exception as e:
    LongestPathEngine = None
    logger.warning("❌ LongestPathEngine import failed: %s", e)


app = Flask(__name__)
app.secret_key = SECRET_KEY
CORS(app)

UPLOAD_FOLDER = 'uploads'
OUTPUT_FOLDER = 'output'
ALLOWED_EXTENSIONS = {'xer'}

app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['OUTPUT_FOLDER'] = OUTPUT_FOLDER
app.config['MAX_CONTENT_LENGTH'] = MAX_UPLOAD_SIZE_MB * 1024 * 1024

os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(OUTPUT_FOLDER, exist_ok=True)


def cleanup_old_files(folder, max_age_hours=SESSION_LIFETIME_HOURS):
    if not os.path.exists(folder):
        return
    cutoff = time.time() - (max_age_hours * 3600)
    for fname in os.listdir(folder):
        fpath = os.path.join(folder, fname)
        try:
            if os.path.isfile(fpath) and os.path.getmtime(fpath) < cutoff:
                os.remove(fpath)
                logger.info("🧹 Cleaned up old file: %s", fname)
        except Exception as e:
            logger.warning("Could not delete %s: %s", fname, e)

cleanup_old_files(UPLOAD_FOLDER)
cleanup_old_files(OUTPUT_FOLDER)


SESSION_STORAGE = {}

def get_session_data():
    sid = session.get('sid')
    if not sid or sid not in SESSION_STORAGE:
        sid = uuid.uuid4().hex
        session['sid'] = sid
        SESSION_STORAGE[sid] = {
            'analysis': {'engine': None, 'dashboard_data': None, 'file_name': None, 'analyzed_at': None},
            'comparison': {'comparator': None, 'results': None, 'baseline_file': None, 'current_file': None},
            'health_cache': {},
            'longest_path_cache': None,
        }
    return SESSION_STORAGE[sid]


def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


def analyze_xer_file(file_path_or_stream, original_filename, session_data):
    logger.info("🔍 Analyzing XER: %s", original_filename)
    parser = XERParser()
    tables = parser.parse(file_path_or_stream)
    if tables is None or not tables:
        return {'error': 'Failed to parse XER file or file is empty.'}
    engine = ScheduleEngine()
    engine.load_data(tables)
    engine.analyze()
    
    # Auto-compute Longest Path and inject into engine
    if LongestPathEngine is not None:
        try:
            lp_engine = LongestPathEngine(engine)
            lp_results = lp_engine.calculate()
            engine.longest_path_ids = lp_engine.longest_path_ids
            engine.longest_path_results = lp_results
        except Exception as e:
            logger.warning("Longest Path auto-calc failed: %s", e)
            engine.longest_path_ids = set()
            engine.longest_path_results = {}
    else:
        engine.longest_path_ids = set()
        engine.longest_path_results = {}
    
    dashboard_data = engine.get_dashboard_data()
    analysis = session_data['analysis']
    analysis['engine'] = engine
    analysis['dashboard_data'] = dashboard_data
    analysis['file_name'] = original_filename
    analysis['analyzed_at'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    session_data['health_cache'] = {}
    session_data['longest_path_cache'] = None
    return dashboard_data


@app.context_processor
def inject_config():
    return {'config': get_config()}


@app.errorhandler(413)
def request_entity_too_large(error):
    return jsonify({'error': f'File size exceeds maximum limit of {MAX_UPLOAD_SIZE_MB} MB.'}), 413


# Base routes (unchanged)
@app.route('/')
def home():
    return render_template('index.html')

@app.route('/api/upload', methods=['POST'])
def upload_file():
    if 'file' not in request.files:
        return jsonify({'error': 'No file provided'}), 400
    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'No file selected'}), 400
    if not allowed_file(file.filename):
        return jsonify({'error': 'File must be a .xer file'}), 400
    original_name = secure_filename(file.filename)
    unique_name = f"{uuid.uuid4().hex}_{original_name}"
    file_path = os.path.join(app.config['UPLOAD_FOLDER'], unique_name)
    file.save(file_path)
    sess_data = get_session_data()
    try:
        dashboard_data = analyze_xer_file(file_path, original_name, sess_data)
        if 'error' in dashboard_data:
            return jsonify(dashboard_data), 400
        analysis = sess_data['analysis']
        return jsonify({
            'success': True, 'file_name': analysis['file_name'],
            'analyzed_at': analysis['analyzed_at'], 'data': dashboard_data
        })
    except Exception as e:
        logger.exception("Upload analysis error")
        return jsonify({'error': str(e)}), 500


@app.route('/api/dashboard')
def get_dashboard():
    sess_data = get_session_data()
    analysis = sess_data['analysis']
    if analysis['dashboard_data'] is None:
        return jsonify({'has_data': False})
    return jsonify({
        'has_data': True, 'file_name': analysis['file_name'],
        'analyzed_at': analysis['analyzed_at'], 'data': analysis['dashboard_data']
    })


@app.route('/api/load-sample')
def load_sample():
    sample_path = os.path.join('input', 'sample.xer')
    if not os.path.exists(sample_path):
        return jsonify({'error': 'sample.xer not found in input/ folder'}), 404
    sess_data = get_session_data()
    try:
        dashboard_data = analyze_xer_file(sample_path, 'sample.xer', sess_data)
        analysis = sess_data['analysis']
        return jsonify({
            'success': True, 'file_name': 'sample.xer',
            'analyzed_at': analysis['analyzed_at'], 'data': dashboard_data
        })
    except Exception as e:
        logger.exception("Sample load error")
        return jsonify({'error': str(e)}), 500


@app.route('/api/export-excel')
def export_excel():
    sess_data = get_session_data()
    engine = sess_data['analysis']['engine']
    if engine is None:
        return jsonify({'error': 'No schedule loaded. Please upload an XER file first.'}), 400
    out_name = f"schedule_report_{uuid.uuid4().hex[:8]}.xlsx"
    output_path = os.path.join(app.config['OUTPUT_FOLDER'], out_name)
    reporter = ReportGenerator(engine)
    res_path = reporter.generate_full_report(output_path)
    if not res_path or not os.path.exists(res_path):
        return jsonify({'error': 'Failed to generate Excel report.'}), 500
    return send_file(res_path, as_attachment=True,
        download_name=f"schedule_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@app.route('/gantt')
def gantt_view():
    return render_template('gantt.html')


@app.route('/api/gantt-data')
def get_gantt_data():
    sess_data = get_session_data()
    analysis = sess_data['analysis']
    if analysis['engine'] is None:
        return jsonify({'error': 'No data loaded. Please upload an XER file first.'}), 400
    max_acts = request.args.get('max', 2000, type=int)
    gantt_data = analysis['engine'].get_gantt_data(max_activities=max_acts)
    return jsonify({'success': True, 'file_name': analysis['file_name'], 'data': gantt_data})


# ═══════════════════════════════════════════
# NEW ROUTE: LONGEST PATH
# ═══════════════════════════════════════════

@app.route('/api/longest-path')
def get_longest_path():
    sess_data = get_session_data()
    analysis = sess_data['analysis']
    if analysis['engine'] is None:
        return jsonify({'error': 'No data loaded. Upload an XER file first.'}), 400
    if LongestPathEngine is None:
        return jsonify({'error': 'longest_path_engine.py is missing!'}), 500
    
    # Use cached result if available
    if sess_data.get('longest_path_cache'):
        return jsonify({
            'success': True, 'file_name': analysis['file_name'],
            'data': sess_data['longest_path_cache'], 'cached': True
        })
    
    try:
        lp_engine = LongestPathEngine(analysis['engine'])
        results = lp_engine.calculate()
        sess_data['longest_path_cache'] = results
        # Sync back to engine for Gantt
        analysis['engine'].longest_path_ids = lp_engine.longest_path_ids
        analysis['engine'].longest_path_results = results
        return jsonify({
            'success': True, 'file_name': analysis['file_name'], 'data': results
        })
    except Exception as e:
        logger.exception("Longest Path calculation error")
        return jsonify({'error': str(e)}), 500


@app.route('/comparison')
def comparison_view():
    return render_template('comparison.html')

@app.route('/api/compare', methods=['POST'])
def compare_schedules():
    if ScheduleComparator is None:
        return jsonify({'error': 'comparison_engine.py is missing!'}), 500
    if 'baseline' not in request.files or 'current' not in request.files:
        return jsonify({'error': 'Both baseline and current files required'}), 400
    baseline_file = request.files['baseline']
    current_file = request.files['current']
    if not (allowed_file(baseline_file.filename) and allowed_file(current_file.filename)):
        return jsonify({'error': 'Both files must be .xer files'}), 400
    baseline_name = secure_filename(baseline_file.filename)
    current_name = secure_filename(current_file.filename)
    baseline_path = os.path.join(app.config['UPLOAD_FOLDER'], f"bl_{uuid.uuid4().hex[:8]}_{baseline_name}")
    current_path = os.path.join(app.config['UPLOAD_FOLDER'], f"cur_{uuid.uuid4().hex[:8]}_{current_name}")
    baseline_file.save(baseline_path)
    current_file.save(current_path)
    sess_data = get_session_data()
    try:
        comparator = ScheduleComparator()
        comparator.load_baseline(baseline_path)
        comparator.load_current(current_path)
        results = comparator.compare()
        comp = sess_data['comparison']
        comp['comparator'] = comparator
        comp['results'] = results
        comp['baseline_file'] = baseline_name
        comp['current_file'] = current_name
        return jsonify({'success': True, 'baseline_file': baseline_name,
            'current_file': current_name, 'results': results})
    except Exception as e:
        logger.exception("Comparison error")
        return jsonify({'error': str(e)}), 500


@app.route('/api/comparison-data')
def get_comparison_data():
    sess_data = get_session_data()
    comp = sess_data['comparison']
    if comp['results'] is None:
        return jsonify({'has_data': False})
    return jsonify({
        'has_data': True, 'baseline_file': comp['baseline_file'],
        'current_file': comp['current_file'], 'results': comp['results']
    })


@app.route('/evm')
def evm_view():
    return render_template('evm.html')

@app.route('/api/evm-data')
def get_evm_data():
    sess_data = get_session_data()
    analysis = sess_data['analysis']
    if analysis['engine'] is None:
        return jsonify({'error': 'No data loaded. Upload an XER file first.'}), 400
    if EVMEngine is None:
        return jsonify({'error': 'evm_engine.py is missing!'}), 500
    try:
        evm = EVMEngine(analysis['engine'])
        results = evm.calculate()
        return jsonify({'success': True, 'file_name': analysis['file_name'], 'data': results})
    except Exception as e:
        logger.exception("EVM calculation error")
        return jsonify({'error': str(e)}), 500


@app.route('/api/resource-data')
def get_resource_data():
    sess_data = get_session_data()
    analysis = sess_data['analysis']
    if analysis['engine'] is None:
        return jsonify({'error': 'No data loaded. Upload an XER file first.'}), 400
    if ResourceEngine is None:
        return jsonify({'error': 'resource_engine.py is missing!'}), 500
    try:
        engine = ResourceEngine(analysis['engine'])
        results = engine.calculate()
        return jsonify({'success': True, 'file_name': analysis['file_name'], 'data': results})
    except Exception as e:
        logger.exception("Resource analytics error")
        return jsonify({'error': str(e)}), 500


@app.route('/health')
def health_view():
    return render_template('health.html')

@app.route('/api/health-data')
def get_health_data():
    sess_data = get_session_data()
    analysis = sess_data['analysis']
    if analysis['engine'] is None:
        return jsonify({'error': 'No data loaded. Upload a file first.'}), 400
    if AdvancedHealthEngine is None:
        return jsonify({'error': 'advanced_health_engine.py is missing!'}), 500
    selected_standard = request.args.get('standard', 'all')
    try:
        health = AdvancedHealthEngine(analysis['engine'])
        results = health.run_all_checks(selected_standard=selected_standard)
        return jsonify({'success': True, 'file_name': analysis['file_name'], 'data': results})
    except Exception as e:
        logger.exception("Health analysis error")
        return jsonify({'error': str(e)}), 500


@app.route('/api/executive-pdf')
def download_executive_pdf():
    sess_data = get_session_data()
    analysis = sess_data['analysis']
    if analysis['engine'] is None:
        return jsonify({'error': 'No data loaded'}), 400
    if PDFReportGenerator is None or AdvancedHealthEngine is None:
        return jsonify({'error': 'PDF or Health engine module missing!'}), 500
    selected_standard = request.args.get('standard', 'all')
    severity_filter = request.args.get('severity', 'all')
    try:
        health = AdvancedHealthEngine(analysis['engine'])
        results = health.run_all_checks(selected_standard=selected_standard)
        generator = PDFReportGenerator(results, analysis['file_name'], severity_filter=severity_filter)
        pdf_buffer = generator.generate_executive_report()
        return send_file(pdf_buffer, as_attachment=True,
            download_name=f"executive_report_{selected_standard}_{severity_filter}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf",
            mimetype='application/pdf')
    except Exception as e:
        logger.exception("PDF generation error")
        return jsonify({'error': str(e)}), 500


@app.route('/api/actions-pdf')
def download_actions_pdf():
    sess_data = get_session_data()
    analysis = sess_data['analysis']
    if analysis['engine'] is None:
        return jsonify({'error': 'No data loaded'}), 400
    if PDFReportGenerator is None or AdvancedHealthEngine is None:
        return jsonify({'error': 'PDF or Health engine module missing!'}), 500
    selected_standard = request.args.get('standard', 'all')
    severity_filter = request.args.get('severity', 'all')
    try:
        health = AdvancedHealthEngine(analysis['engine'])
        results = health.run_all_checks(selected_standard=selected_standard)
        generator = PDFReportGenerator(results, analysis['file_name'], severity_filter=severity_filter)
        pdf_buffer = generator.generate_actions_report()
        return send_file(pdf_buffer, as_attachment=True,
            download_name=f"action_list_{selected_standard}_{severity_filter}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf",
            mimetype='application/pdf')
    except Exception as e:
        logger.exception("PDF actions error")
        return jsonify({'error': str(e)}), 500


@app.route('/api/actions-excel')
def download_actions_excel():
    sess_data = get_session_data()
    analysis = sess_data['analysis']
    if analysis['engine'] is None:
        return jsonify({'error': 'No data loaded. Upload a file first.'}), 400
    if AdvancedHealthEngine is None:
        return jsonify({'error': 'advanced_health_engine.py is missing!'}), 500
    selected_standard = request.args.get('standard', 'all')
    severity_filter = request.args.get('severity', 'all').lower()
    severity_levels = {
        'critical': ['critical'], 'high': ['critical', 'high'],
        'medium': ['critical', 'high', 'medium'],
        'all': ['critical', 'high', 'medium', 'low', 'info']
    }
    allowed_severities = severity_levels.get(severity_filter, severity_levels['all'])
    try:
        import pandas as pd
        from io import BytesIO
        from openpyxl.styles import Font, PatternFill, Alignment
        health = AdvancedHealthEngine(analysis['engine'])
        results = health.run_all_checks(selected_standard=selected_standard)
        top_actions = results.get('top_actions', []) or []
        standards_data = results.get('standards', {}) or {}
        filtered_top_actions = [
            a for a in top_actions if (a.get('severity') or 'low').lower() in allowed_severities
        ]
        meta_rows = [
            ['Report Type', 'Schedule Health - Top Actions Export'],
            ['Selected Standard', selected_standard],
            ['Severity Filter', severity_filter.upper()],
            ['File Name', analysis.get('file_name', '')],
            ['Generated At', results.get('analysis_date', '')],
            ['', ''],
            ['Overall Score', results.get('overall_score', '')],
            ['Total Checks', results.get('total_checks', '')],
            ['Passed Checks', results.get('passed_checks', '')],
            ['Failed Checks', results.get('failed_checks', '')],
            ['Critical Failures', results.get('critical_failures', '')],
            ['High Failures', results.get('high_failures', '')],
            ['Pass Rate (%)', results.get('pass_rate', '')],
            ['', ''],
            ['Filtered Actions Count', len(filtered_top_actions)],
        ]
        top_summary_rows = []
        for idx, action in enumerate(filtered_top_actions, 1):
            top_summary_rows.append({
                'Rank': idx, 'Standard': action.get('standard', ''),
                'Check ID': action.get('id', ''), 'Check Name': action.get('name', ''),
                'Category': action.get('category', ''),
                'Severity': (action.get('severity') or '').upper(),
                'Affected Count': action.get('count', 0), 'Total': action.get('total', 0),
                'Percentage': action.get('percentage', 0), 'Threshold': action.get('threshold', ''),
                'Value': action.get('value', ''), 'Recommendation': action.get('recommendation', ''),
                'Description': action.get('description', ''),
            })
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            pd.DataFrame(meta_rows, columns=['Field', 'Value']).to_excel(writer, sheet_name='Report Info', index=False)
            if top_summary_rows:
                pd.DataFrame(top_summary_rows).to_excel(writer, sheet_name='Top Actions Summary', index=False)
            else:
                pd.DataFrame([{'Info': f'No actions matched severity filter: {severity_filter.upper()}'}]).to_excel(writer, sheet_name='Top Actions Summary', index=False)
            for std_name, std_data in standards_data.items():
                sheet_name = (std_name or 'Standard')[:31]
                rows = []
                failed_in_std = []
                for category in std_data.get('categories', []):
                    for check in category.get('checks', []):
                        if check.get('status') != 'fail': continue
                        if (check.get('severity') or 'low').lower() not in allowed_severities: continue
                        failed_in_std.append((category.get('name', ''), check))
                if not failed_in_std:
                    rows.append({'Section': 'No Matching Failures', 'Field': '', 'Value': f'No {severity_filter.upper()} failures found for {std_name}', 'Activity ID': '', 'Activity Name': '', 'WBS': ''})
                else:
                    for cat_name, check in failed_in_std:
                        rows.append({'Section': 'METRIC', 'Field': 'Check ID', 'Value': check.get('id', ''), 'Activity ID': '', 'Activity Name': '', 'WBS': ''})
                        rows.append({'Section': '', 'Field': 'Check Name', 'Value': check.get('name', ''), 'Activity ID': '', 'Activity Name': '', 'WBS': ''})
                        rows.append({'Section': '', 'Field': 'Category', 'Value': cat_name, 'Activity ID': '', 'Activity Name': '', 'WBS': ''})
                        rows.append({'Section': '', 'Field': 'Severity', 'Value': (check.get('severity') or '').upper(), 'Activity ID': '', 'Activity Name': '', 'WBS': ''})
                        rows.append({'Section': '', 'Field': 'Description', 'Value': check.get('description', ''), 'Activity ID': '', 'Activity Name': '', 'WBS': ''})
                        rows.append({'Section': '', 'Field': 'Threshold', 'Value': check.get('threshold', ''), 'Activity ID': '', 'Activity Name': '', 'WBS': ''})
                        if check.get('count') is not None:
                            rows.append({'Section': '', 'Field': 'Affected', 'Value': f"{check.get('count', 0)} of {check.get('total', 0)} ({check.get('percentage', 0)}%)", 'Activity ID': '', 'Activity Name': '', 'WBS': ''})
                        if check.get('value') is not None and check.get('value') != '':
                            rows.append({'Section': '', 'Field': 'Value', 'Value': check.get('value', ''), 'Activity ID': '', 'Activity Name': '', 'WBS': ''})
                        rows.append({'Section': '', 'Field': 'Recommendation', 'Value': check.get('recommendation', ''), 'Activity ID': '', 'Activity Name': '', 'WBS': ''})
                        items = check.get('failed_items', []) or []
                        rows.append({'Section': 'AFFECTED ACTIVITIES', 'Field': f'Total: {len(items)}', 'Value': '', 'Activity ID': 'Activity ID', 'Activity Name': 'Activity Name', 'WBS': 'WBS'})
                        if items:
                            for item in items:
                                rows.append({'Section': '', 'Field': '', 'Value': '', 'Activity ID': item.get('code', ''), 'Activity Name': item.get('name', ''), 'WBS': item.get('wbs', '')})
                        else:
                            rows.append({'Section': '', 'Field': '', 'Value': '(No activity list available)', 'Activity ID': '', 'Activity Name': '', 'WBS': ''})
                        rows.append({'Section': '', 'Field': '', 'Value': '', 'Activity ID': '', 'Activity Name': '', 'WBS': ''})
                pd.DataFrame(rows).to_excel(writer, sheet_name=sheet_name, index=False)
            workbook = writer.book
            header_font = Font(bold=True, color='FFFFFF', size=11)
            header_fill = PatternFill(start_color='1E40AF', end_color='1E40AF', fill_type='solid')
            metric_fill = PatternFill(start_color='FEF3C7', end_color='FEF3C7', fill_type='solid')
            activity_hdr_fill = PatternFill(start_color='DBEAFE', end_color='DBEAFE', fill_type='solid')
            bold_font = Font(bold=True)
            wrap_align = Alignment(wrap_text=True, vertical='top')
            for sheet_name in workbook.sheetnames:
                ws = workbook[sheet_name]
                for cell in ws[1]:
                    cell.font = header_font; cell.fill = header_fill
                    cell.alignment = Alignment(horizontal='left', vertical='center')
                for col in ws.columns:
                    max_len = 10; col_letter = col[0].column_letter
                    for cell in col:
                        try:
                            val = str(cell.value) if cell.value is not None else ''
                            if len(val) > max_len: max_len = len(val)
                        except Exception: pass
                    ws.column_dimensions[col_letter].width = min(max_len + 2, 60)
                ws.freeze_panes = 'A2'
                if sheet_name not in ['Report Info', 'Top Actions Summary']:
                    for row in ws.iter_rows(min_row=2):
                        section_cell = row[0]
                        if section_cell.value == 'METRIC':
                            for c in row: c.fill = metric_fill; c.font = bold_font
                        elif section_cell.value == 'AFFECTED ACTIVITIES':
                            for c in row: c.fill = activity_hdr_fill; c.font = bold_font
                        for c in row: c.alignment = wrap_align
        output.seek(0)
        filename = f"health_top_actions_{selected_standard}_{severity_filter}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return send_file(output, as_attachment=True, download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    except Exception as e:
        logger.exception("Actions Excel error")
        return jsonify({'error': str(e)}), 500


@app.route('/api/ai-narrative')
def get_ai_narrative():
    sess_data = get_session_data()
    analysis = sess_data['analysis']
    if analysis['engine'] is None:
        return jsonify({'error': 'No schedule loaded. Upload an XER file first.'}), 400
    if AINarrativeEngine is None:
        return jsonify({'error': 'ai_narrative_engine.py is missing!'}), 500
    try:
        health_data = {}
        if AdvancedHealthEngine is not None:
            try:
                health_gen = AdvancedHealthEngine(analysis['engine'])
                health_data = health_gen.run_all_checks('all')
            except Exception as he:
                logger.warning("Health data unavailable for narrative: %s", he)
        evm_data = {}
        if EVMEngine is not None:
            try:
                evm_data = EVMEngine(analysis['engine']).calculate()
            except Exception: pass
        comp_data = sess_data.get('comparison', {}).get('results', {}) or {}
        narrative_gen = AINarrativeEngine(health_data=health_data, comparison_data=comp_data, evm_data=evm_data)
        results = narrative_gen.generate_narrative()
        return jsonify({'success': True, 'file_name': analysis['file_name'], 'data': results})
    except Exception as e:
        logger.exception("AI Narrative generation error")
        return jsonify({'error': str(e)}), 500


if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    debug_mode = os.environ.get('FLASK_ENV') != 'production'
    logger.info("============================================================")
    logger.info("🚀 P6 ANALYZER - LONGEST PATH + AI + RESOURCES + FULL SUITE")
    logger.info("============================================================")
    logger.info("📌 Running on port %s", port)
    logger.info("👉 Open in browser: http://localhost:%s", port)
    app.run(debug=debug_mode, host='0.0.0.0', port=port)
'''

with open("app.py", "w", encoding="utf-8") as f:
    f.write(APP_CODE)
print("  ✅ Updated app.py")


# ==============================================================================
# FILE 3: data_engine.py -- Only patches get_gantt_data() to inject 
# longest_path flag into each task
# ==============================================================================

# Read existing data_engine.py, find get_gantt_data method, inject longest_path support
try:
    with open("data_engine.py", "r", encoding="utf-8") as f:
        de_content = f.read()

    # Add is_longest_path flag to task dict
    if "'is_longest_path'" not in de_content:
        # Insert into task dict in get_gantt_data after 'is_critical' line
        de_content = de_content.replace(
            "'is_critical': is_critical,",
            "'is_critical': is_critical,\n                'is_longest_path': str(act.get('task_id', '')) in getattr(self, 'longest_path_ids', set()),",
            1
        )
    
    # Also update wbs_task dict
    if "'longest_path_count'" not in de_content:
        de_content = de_content.replace(
            "'is_critical': (min_float <= 0) if min_float != float('inf') else False,",
            "'is_critical': (min_float <= 0) if min_float != float('inf') else False,\n                'longest_path_count': sum(1 for a in child_acts if str(a.get('task_id', '')) in getattr(self, 'longest_path_ids', set())),",
            1
        )
    
    # Add longest_path_count and longest_path_ids to gantt data return
    if "'longest_path_count'" not in de_content or "'longest_path_ids'" not in de_content.split("return {\n            'tasks'")[1][:500]:
        old_return = """return {
            'tasks': all_tasks, 'links': links,
            'total': len(tasks), 'wbs_summary_count': len(wbs_summary_tasks),
            'critical_count': sum(1 for t in tasks if t.get('is_critical')),
            'data_date': data_date.strftime('%Y-%m-%d') if data_date else '',
            'groupable_values': {},
        }"""
        new_return = """return {
            'tasks': all_tasks, 'links': links,
            'total': len(tasks), 'wbs_summary_count': len(wbs_summary_tasks),
            'critical_count': sum(1 for t in tasks if t.get('is_critical')),
            'longest_path_count': sum(1 for t in tasks if t.get('is_longest_path')),
            'longest_path_ids': list(getattr(self, 'longest_path_ids', set())),
            'data_date': data_date.strftime('%Y-%m-%d') if data_date else '',
            'groupable_values': {},
        }"""
        de_content = de_content.replace(old_return, new_return)
    
    with open("data_engine.py", "w", encoding="utf-8") as f:
        f.write(de_content)
    print("  ✅ Patched data_engine.py (added longest_path flag)")
except Exception as e:
    print(f"  ⚠️ Could not auto-patch data_engine.py: {e}")
    print(f"     You may need to manually add 'is_longest_path' to task dict in get_gantt_data()")


# ==============================================================================
# FILE 4: templates/gantt.html - Add "Longest Path Only" toggle button
# ==============================================================================

try:
    with open("templates/gantt.html", "r", encoding="utf-8") as f:
        gantt_html = f.read()
    
    # Insert Longest Path toggle before Critical Only button
    if 'longestPathBtn' not in gantt_html:
        old_snippet = '<button type="button" class="tb-btn" onclick="toggleCritical()">\n                    <span id="criticalBtn">🔴 Critical Only</span>\n                </button>'
        new_snippet = '''<button type="button" class="tb-btn" onclick="toggleLongestPath()">
                    <span id="longestPathBtn">🎯 Longest Path Only</span>
                </button>
                <button type="button" class="tb-btn" onclick="toggleCritical()">
                    <span id="criticalBtn">🔴 Critical (TF≤0) Only</span>
                </button>'''
        gantt_html = gantt_html.replace(old_snippet, new_snippet)
    
    # Add Longest Path bar color style
    if 'gantt-longest-path' not in gantt_html:
        old_css = '.gantt_task_line.gantt-critical { background-color: #dc2626 !important; border-color: #991b1b; }'
        new_css = '''.gantt_task_line.gantt-critical { background-color: #dc2626 !important; border-color: #991b1b; }
        .gantt_task_line.gantt-longest-path { background-color: #7c3aed !important; border-color: #5b21b6; }
        .gantt_row.longest-path-row { background-color: #f3e8ff !important; }'''
        gantt_html = gantt_html.replace(old_css, new_css)
    
    with open("templates/gantt.html", "w", encoding="utf-8") as f:
        f.write(gantt_html)
    print("  ✅ Patched templates/gantt.html (added Longest Path toggle)")
except Exception as e:
    print(f"  ⚠️ Could not auto-patch gantt.html: {e}")


# ==============================================================================
# FILE 5: static/gantt.js - Add toggleLongestPath() and rendering logic
# ==============================================================================

try:
    with open("static/gantt.js", "r", encoding="utf-8") as f:
        gantt_js = f.read()
    
    # Add showLongestPathOnly variable near showCriticalOnly
    if 'showLongestPathOnly' not in gantt_js:
        gantt_js = gantt_js.replace(
            'let showCriticalOnly = false;',
            'let showCriticalOnly = false;\nlet showLongestPathOnly = false;'
        )
    
    # Add toggleLongestPath function after toggleCritical
    if 'function toggleLongestPath' not in gantt_js:
        toggle_lp_function = '''

function toggleLongestPath() {
    showLongestPathOnly = !showLongestPathOnly;
    const el = document.getElementById('longestPathBtn');
    const btn = el ? el.closest('.tb-btn') : null;
    if (btn) btn.classList.toggle('active', showLongestPathOnly);
    if (el) el.textContent = showLongestPathOnly ? '🎯 Longest Path Only ✓' : '🎯 Longest Path Only';
    renderGantt();
}
'''
        gantt_js = gantt_js.replace(
            'function toggleWbsSummary()',
            toggle_lp_function + '\nfunction toggleWbsSummary()'
        )
    
    # Add longest path filter in renderGantt() alongside critical filter
    if 'showLongestPathOnly' in gantt_js and 'tasksToShow = tasksToShow.filter(function (t) {\n            return t.is_wbs_summary || t.is_longest_path;' not in gantt_js:
        old_filter = '''if (showCriticalOnly) {
        tasksToShow = tasksToShow.filter(function (t) {
            return t.is_wbs_summary || t.is_critical;
        });
    }'''
        new_filter = '''if (showCriticalOnly) {
        tasksToShow = tasksToShow.filter(function (t) {
            return t.is_wbs_summary || t.is_critical;
        });
    }
    if (showLongestPathOnly) {
        tasksToShow = tasksToShow.filter(function (t) {
            return t.is_wbs_summary || t.is_longest_path;
        });
    }'''
        gantt_js = gantt_js.replace(old_filter, new_filter)
    
    # Update task_class template to highlight longest path bars
    if "if (task.is_longest_path) return 'gantt-longest-path';" not in gantt_js:
        gantt_js = gantt_js.replace(
            "if (task.is_group) return 'gantt-summary-l' + (task.group_level || 1);\n        return task.custom_class || '';",
            "if (task.is_group) return 'gantt-summary-l' + (task.group_level || 1);\n        if (task.is_longest_path) return 'gantt-longest-path';\n        return task.custom_class || '';"
        )
    
    # Update grid_row_class for longest path row highlight
    if "if (task.is_longest_path) return 'longest-path-row';" not in gantt_js:
        gantt_js = gantt_js.replace(
            "if (task.is_critical) return 'critical-row';",
            "if (task.is_longest_path) return 'longest-path-row';\n        if (task.is_critical) return 'critical-row';"
        )
    
    # Update stats display to show longest path count
    if "Longest Path: " not in gantt_js:
        gantt_js = gantt_js.replace(
            "'📌 ' + (data.total || 0) + ' activities | 🌳 ' +\n                    (data.wbs_summary_count || 0) + ' WBS | 🔴 ' +\n                    (data.critical_count || 0) + ' critical | rows: ' + allTasks.length;",
            "'📌 ' + (data.total || 0) + ' activities | 🌳 ' +\n                    (data.wbs_summary_count || 0) + ' WBS | 🔴 ' +\n                    (data.critical_count || 0) + ' critical | 🎯 ' +\n                    (data.longest_path_count || 0) + ' longest path | rows: ' + allTasks.length;"
        )
    
    # Expose toggle to window
    if 'window.toggleLongestPath' not in gantt_js:
        gantt_js += '\n\nwindow.toggleLongestPath = toggleLongestPath;\n'
    
    with open("static/gantt.js", "w", encoding="utf-8") as f:
        f.write(gantt_js)
    print("  ✅ Patched static/gantt.js (added Longest Path toggle & rendering)")
except Exception as e:
    print(f"  ⚠️ Could not auto-patch gantt.js: {e}")


print("\n🎉 Phase 2 - Step 4 (Native Longest Path) Applied Successfully!")
print("✨ Restart Flask (python app.py), upload an XER, and go to /gantt.")
print("   Click '🎯 Longest Path Only' to see the true driving critical path (purple bars).")
print("   Access API results at: http://localhost:5000/api/longest-path")