"""
REPORT GENERATOR
================
Creates Excel reports from the Schedule Engine results
(basic dashboard DCMA + activities dump).

Advanced 622+ health Excel is handled separately in app.py
(/api/actions-excel).
"""

import os
from datetime import datetime
import logging

logger = logging.getLogger(__name__)


class ReportGenerator:
    """
    Generates Excel reports from schedule analysis results.

    USAGE:
        reporter = ReportGenerator(engine)
        reporter.generate_full_report("output/schedule_report.xlsx")
    """

    def __init__(self, engine):
        self.engine = engine

    def generate_full_report(self, output_path):
        """
        Create a multi-tab Excel workbook:
        Summary | DCMA 14-Point | Activities | Critical Path | Relationships | Issues
        """
        try:
            import pandas as pd
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError as e:
            logger.error("pandas/openpyxl required: %s", e)
            print("❌ pandas and openpyxl are required. Run: pip install pandas openpyxl")
            return None

        logger.info("📝 Generating report: %s", output_path)

        out_dir = os.path.dirname(os.path.abspath(output_path))
        if out_dir:
            os.makedirs(out_dir, exist_ok=True)

        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            self._write_summary_tab(writer, pd)
            self._write_dcma_tab(writer, pd)
            self._write_activities_tab(writer, pd)
            self._write_critical_path_tab(writer, pd)
            self._write_relationships_tab(writer, pd)
            self._write_issues_tab(writer, pd)

            # Format all sheets
            workbook = writer.book
            for sheet_name in workbook.sheetnames:
                self._style_worksheet(
                    workbook[sheet_name],
                    pass_fail_col=(3 if sheet_name == 'DCMA 14-Point' else None),
                )

        logger.info("✅ Report saved: %s", output_path)
        print(f"  ✅ Report saved: {output_path}")
        return output_path

    # ═══════════════════════════════════════════════════════
    # TABS
    # ═══════════════════════════════════════════════════════

    def _write_summary_tab(self, writer, pd):
        stats = self.engine.schedule_stats or {}
        info = {}
        if hasattr(self.engine, '_get_project_info'):
            info = self.engine._get_project_info() or {}

        rows = [
            ['Schedule Summary Report', ''],
            ['Generated', datetime.now().strftime('%Y-%m-%d %H:%M')],
            ['', ''],
            ['PROJECT', ''],
            ['Project Name', info.get('name', '')],
            ['Plan Start', info.get('start', '')],
            ['Plan Finish', info.get('finish', '')],
            ['Data Date', info.get('data_date', '')],
            ['', ''],
            ['ACTIVITY COUNTS', ''],
            ['Total Activities', stats.get('total_activities', 0)],
            ['  Not Started', stats.get('not_started', 0)],
            ['  In Progress', stats.get('in_progress', 0)],
            ['  Completed', stats.get('completed', 0)],
            ['', ''],
            ['ACTIVITY TYPES', ''],
            ['  Tasks', stats.get('tasks', 0)],
            ['  Milestones', stats.get('milestones', 0)],
            ['  Level of Effort', stats.get('loe', 0)],
            ['  WBS Summary', stats.get('wbs_summary', 0)],
            ['', ''],
            ['CRITICAL PATH', ''],
            ['  Critical Activities', stats.get('critical_count', 0)],
            ['  Negative Float', stats.get('negative_float', 0)],
            ['  High Float (>44d)', stats.get('high_float_gt_44d', 0)],
            ['', ''],
            ['NETWORK', ''],
            ['  Total Relationships', stats.get('total_relationships', 0)],
            ['  Calendars', stats.get('total_calendars', 0)],
            ['', ''],
            ['Note', 'Basic DCMA dashboard export. Use Health → Excel for 622+ standards.'],
        ]

        df = pd.DataFrame(rows, columns=['Metric', 'Value'])
        df.to_excel(writer, sheet_name='Summary', index=False)
        print("  📊 Summary tab created")

    def _write_dcma_tab(self, writer, pd):
        rows = []
        for check_name, result in (self.engine.dcma_results or {}).items():
            clean_name = check_name.replace('_', ' ')
            has_pct = 'pct' in result
            has_count = 'count' in result

            if result.get('pass') is True:
                status = 'PASS'
            elif result.get('pass') is False:
                status = 'FAIL'
            else:
                status = 'N/A'

            if has_pct:
                value_display = f"{result.get('pct', '')}%"
                count_display = result.get('count', '')
                total_display = result.get('total', '')
            elif 'value' in result:
                value_display = str(result.get('value', ''))
                count_display = ''
                total_display = ''
            else:
                value_display = ''
                count_display = result.get('count', '')
                total_display = result.get('total', '')

            rows.append({
                'Check': clean_name,
                'Value': value_display,
                'Count': count_display,
                'Total': total_display,
                'Threshold': result.get('threshold', ''),
                'Result': status,
            })

        if not rows:
            rows = [{'Check': 'No DCMA results', 'Value': '', 'Count': '', 'Total': '',
                     'Threshold': '', 'Result': 'N/A'}]

        df = pd.DataFrame(rows)
        # Column order: Result at index 5 for styling — use explicit order
        df = df[['Check', 'Value', 'Count', 'Total', 'Threshold', 'Result']]
        df.to_excel(writer, sheet_name='DCMA 14-Point', index=False)
        print("  🏥 DCMA tab created")

    def _write_activities_tab(self, writer, pd):
        df = None
        if hasattr(self.engine, 'get_activities_dataframe'):
            df = self.engine.get_activities_dataframe()

        if df is None or df.empty:
            # Fallback build
            cols = [
                'task_code', 'task_name', 'wbs_code', 'wbs_name', 'status_text', 'type_text',
                'original_duration_days', 'remaining_duration_days', 'total_float_days',
                'is_critical', 'early_start_date', 'early_end_date', 'phys_complete_pct'
            ]
            data = [{c: a.get(c, '') for c in cols} for a in self.engine.activities]
            df = pd.DataFrame(data)

        rename = {
            'task_code': 'Activity ID',
            'task_name': 'Activity Name',
            'wbs_code': 'WBS Code',
            'wbs_name': 'WBS',
            'status_text': 'Status',
            'type_text': 'Type',
            'original_duration_days': 'Original Duration (d)',
            'remaining_duration_days': 'Remaining Duration (d)',
            'total_float_days': 'Total Float (d)',
            'free_float_days': 'Free Float (d)',
            'is_critical': 'Critical',
            'early_start_date': 'Early Start',
            'early_end_date': 'Early Finish',
            'late_start_date': 'Late Start',
            'late_end_date': 'Late Finish',
            'target_start_date': 'Target Start',
            'target_end_date': 'Target Finish',
            'act_start_date': 'Actual Start',
            'act_end_date': 'Actual Finish',
            'phys_complete_pct': 'Physical %',
        }
        df = df.rename(columns={k: v for k, v in rename.items() if k in df.columns})
        df.to_excel(writer, sheet_name='Activities', index=False)
        print(f"  📌 Activities tab created ({len(df)} rows)")

    def _write_critical_path_tab(self, writer, pd):
        columns = [
            'task_code', 'task_name', 'wbs_name',
            'original_duration_days', 'total_float_days',
            'early_start_date', 'early_end_date', 'status_text'
        ]
        rename = {
            'task_code': 'Activity ID',
            'task_name': 'Activity Name',
            'wbs_name': 'WBS',
            'original_duration_days': 'Duration (d)',
            'total_float_days': 'Total Float (d)',
            'early_start_date': 'Early Start',
            'early_end_date': 'Early Finish',
            'status_text': 'Status',
        }

        crit = getattr(self.engine, 'critical_activities', []) or []
        rows = [{col: act.get(col, '') for col in columns} for act in crit]

        if not rows:
            df = pd.DataFrame([{'Info': 'No critical activities found'}])
        else:
            df = pd.DataFrame(rows).rename(columns=rename)

        df.to_excel(writer, sheet_name='Critical Path', index=False)
        print(f"  🔴 Critical Path tab created ({len(crit)} rows)")

    def _write_relationships_tab(self, writer, pd):
        columns = [
            'pred_code', 'pred_name', 'succ_code', 'succ_name',
            'type_text', 'lag_days'
        ]
        rename = {
            'pred_code': 'Predecessor ID',
            'pred_name': 'Predecessor Name',
            'succ_code': 'Successor ID',
            'succ_name': 'Successor Name',
            'type_text': 'Type',
            'lag_days': 'Lag (d)',
        }

        rels = getattr(self.engine, 'relationships', []) or []
        rows = [{col: rel.get(col, '') for col in columns} for rel in rels]

        if not rows:
            df = pd.DataFrame([{'Info': 'No relationships found'}])
        else:
            df = pd.DataFrame(rows).rename(columns=rename)

        df.to_excel(writer, sheet_name='Relationships', index=False)
        print(f"  🔗 Relationships tab created ({len(rels)} rows)")

    def _write_issues_tab(self, writer, pd):
        issues = []

        for check_name, result in (self.engine.dcma_results or {}).items():
            if result.get('pass') is not False:
                continue

            items = result.get('activities') or result.get('items') or []
            clean_check = check_name.replace('_', ' ')

            if not items:
                issues.append({
                    'Issue Type': clean_check,
                    'Activity ID': '',
                    'Activity Name': '(No item list)',
                    'WBS': '',
                    'Status': '',
                    'Total Float': '',
                    'Duration': '',
                    'Detail': f"Count={result.get('count', '')}",
                })
                continue

            for item in items[:100]:
                code, name, extra = self._format_issue_item(item)
                issues.append({
                    'Issue Type': clean_check,
                    'Activity ID': code,
                    'Activity Name': name,
                    'WBS': (item.get('wbs_name') or '')[:60] if isinstance(item, dict) else '',
                    'Status': item.get('status_text', '') if isinstance(item, dict) else '',
                    'Total Float': item.get('total_float_days', '') if isinstance(item, dict) else '',
                    'Duration': item.get('original_duration_days', '') if isinstance(item, dict) else '',
                    'Detail': extra,
                })

        if not issues:
            df = pd.DataFrame([{'Info': 'No failed DCMA checks'}])
        else:
            df = pd.DataFrame(issues)

        df.to_excel(writer, sheet_name='Issues', index=False)
        print(f"  ⚠️ Issues tab created ({len(issues)} rows)")

    def _format_issue_item(self, item):
        """
        Normalize activity or relationship dict into (id, name, detail).
        """
        if not isinstance(item, dict):
            return str(item), '', ''

        # Relationship-like
        if (
            'pred_code' in item
            or 'succ_code' in item
            or 'pred_task_id' in item
            or item.get('pred_type')
            or item.get('type_text')
        ):
            p = item.get('pred_code') or item.get('pred_task_id', '')
            s = item.get('succ_code') or item.get('task_code') or item.get('task_id', '')
            pn = item.get('pred_name', '')
            sn = item.get('succ_name', '')
            rt = item.get('type_text') or item.get('pred_type', '')
            lag = item.get('lag_days', '')
            return (
                f"{p} → {s}",
                f"{pn} → {sn}",
                f"{rt}, lag={lag}",
            )

        return (
            item.get('task_code', '') or item.get('wbs_short_name', ''),
            item.get('task_name', '') or item.get('wbs_name', ''),
            '',
        )

    # ═══════════════════════════════════════════════════════
    # STYLING
    # ═══════════════════════════════════════════════════════

    def _style_worksheet(self, ws, pass_fail_col=None):
        """Apply header style, freeze panes, column widths, optional PASS/FAIL colors."""
        from openpyxl.styles import Font, PatternFill, Alignment

        header_font = Font(bold=True, color='FFFFFF', size=11)
        header_fill = PatternFill(start_color='1E40AF', end_color='1E40AF', fill_type='solid')
        green = PatternFill(start_color='D1FAE5', end_color='D1FAE5', fill_type='solid')
        red = PatternFill(start_color='FEE2E2', end_color='FEE2E2', fill_type='solid')
        gray = PatternFill(start_color='F1F5F9', end_color='F1F5F9', fill_type='solid')

        if ws.max_row < 1:
            return

        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='left', vertical='center')

        ws.freeze_panes = 'A2'

        # Auto width (capped)
        for col in ws.columns:
            max_len = 8
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    val = str(cell.value) if cell.value is not None else ''
                    if len(val) > max_len:
                        max_len = len(val)
                except Exception:
                    pass
            ws.column_dimensions[col_letter].width = min(max_len + 2, 55)

        # PASS/FAIL coloring — find Result column by header name
        result_col_idx = None
        for idx, cell in enumerate(ws[1], start=1):
            if str(cell.value).strip().lower() == 'result':
                result_col_idx = idx
                break

        if result_col_idx:
            for row in ws.iter_rows(min_row=2, min_col=result_col_idx, max_col=result_col_idx):
                cell = row[0]
                val = str(cell.value or '').upper()
                if val == 'PASS':
                    cell.fill = green
                    cell.font = Font(bold=True, color='065F46')
                elif val == 'FAIL':
                    cell.fill = red
                    cell.font = Font(bold=True, color='991B1B')
                elif val == 'N/A':
                    cell.fill = gray