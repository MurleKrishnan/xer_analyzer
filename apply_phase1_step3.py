import os
import shutil
from datetime import datetime

print("🚀 Applying Phase 1 - Step 3: Resource Units & Cost Histograms...")

# 1. Create Backup Folder
timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
backup_dir = f"_backup_phase1_step3_{timestamp}"
os.makedirs(backup_dir, exist_ok=True)
print(f"📦 Created backup folder: {backup_dir}")

files_to_backup = [
    "app.py",
    "templates/evm.html",
    "static/evm.js",
]

for file_path in files_to_backup:
    if os.path.exists(file_path):
        dest = os.path.join(backup_dir, os.path.basename(file_path.replace("/", os.sep)))
        shutil.copy2(file_path, dest)
        print(f"   Backed up {file_path}")


# ==============================================================================
# FILE 1: resource_engine.py (NEW MODULE - Extract Resource Analytics)
# ==============================================================================

RESOURCE_ENGINE_CODE = '''"""
RESOURCE ANALYTICS ENGINE
==========================
Extracts detailed resource loading data from XER files:
- Monthly man-hours histograms (stacked by resource type)
- Monthly cost burn rates (with cumulative curves)
- Top resources by units and cost
- Planned vs Actual comparisons
"""

from datetime import datetime, timedelta
from collections import defaultdict
import logging

logger = logging.getLogger(__name__)


class ResourceEngine:
    """Extracts resource loading histograms and cost curves from the ScheduleEngine."""

    def __init__(self, engine):
        self.engine = engine
        self.resource_data = {}
        
        # Build resource name lookup
        self.resource_names = {}
        self.resource_types = {}
        for r in engine.raw_tables.get('RSRC', {}).get('rows', []):
            rid = str(r.get('rsrc_id', ''))
            if rid:
                name = r.get('rsrc_name', '') or r.get('rsrc_short_name', '') or 'Unnamed'
                self.resource_names[rid] = name
                # Classify by resource type
                rtype = r.get('rsrc_type', '')
                type_label = {
                    'RT_Labor': 'Labor',
                    'RT_Equip': 'Equipment',
                    'RT_Mat': 'Material',
                    'RT_Nonlabor': 'Non-Labor',
                }.get(rtype, 'Other')
                self.resource_types[rid] = type_label

    def calculate(self):
        """Run all resource extractions."""
        logger.info("👷 Calculating resource analytics...")
        try:
            self._extract_monthly_units()
            self._extract_monthly_costs()
            self._extract_top_resources()
            logger.info("  ✅ Resource analytics complete")
        except Exception as e:
            logger.exception("Resource extraction error: %s", e)
            self.resource_data = {'error': str(e)}
        return self.resource_data

    def _extract_monthly_units(self):
        """
        Spread activity resource hours across their duration into monthly buckets.
        Groups by resource type (Labor / Equipment / Material / etc).
        """
        activities = {str(a.get('task_id', '')): a for a in self.engine.activities}
        
        # {(year, month): {resource_type: units}}
        monthly_units_by_type = defaultdict(lambda: defaultdict(float))
        
        # Overall cumulative
        monthly_totals = defaultdict(float)
        
        for res in self.engine.resources:
            task_id = str(res.get('task_id', ''))
            rsrc_id = str(res.get('rsrc_id', ''))
            
            act = activities.get(task_id)
            if not act:
                continue
            
            total_units = self._to_float(res.get('target_qty', '0'))
            if total_units <= 0:
                continue
            
            rtype = self.resource_types.get(rsrc_id, 'Other')
            
            # Get activity time window
            start = act.get('target_start_date_parsed') or act.get('early_start_date_parsed')
            end = act.get('target_end_date_parsed') or act.get('early_end_date_parsed')
            if not start or not end or end <= start:
                continue
            
            # Spread units evenly by day across activity duration
            total_days = max(1, (end - start).days)
            units_per_day = total_units / total_days
            
            # Walk through each day and bucket into month
            current = start
            while current <= end:
                key = (current.year, current.month)
                monthly_units_by_type[key][rtype] += units_per_day
                monthly_totals[key] += units_per_day
                current += timedelta(days=1)
        
        # Convert to sorted month timeline
        if not monthly_totals:
            self.resource_data['monthly_units'] = {'labels': [], 'datasets': [], 'total_curve': []}
            return
        
        sorted_months = sorted(monthly_totals.keys())
        labels = [f"{y}-{m:02d}" for y, m in sorted_months]
        
        # Collect all resource types present
        all_types = set()
        for month_data in monthly_units_by_type.values():
            all_types.update(month_data.keys())
        
        type_colors = {
            'Labor': '#3b82f6',
            'Equipment': '#f59e0b',
            'Material': '#10b981',
            'Non-Labor': '#8b5cf6',
            'Other': '#64748b',
        }
        
        datasets = []
        for rtype in sorted(all_types):
            data_points = []
            for month_key in sorted_months:
                data_points.append(round(monthly_units_by_type[month_key].get(rtype, 0), 1))
            datasets.append({
                'label': rtype,
                'data': data_points,
                'backgroundColor': type_colors.get(rtype, '#64748b'),
                'stack': 'units',
            })
        
        # Cumulative line
        cumulative = []
        running = 0.0
        for month_key in sorted_months:
            running += monthly_totals[month_key]
            cumulative.append(round(running, 1))
        
        self.resource_data['monthly_units'] = {
            'labels': labels,
            'datasets': datasets,
            'cumulative_curve': cumulative,
            'peak_month': labels[monthly_totals.values().__iter__().__next__() and 
                                 max(range(len(sorted_months)), key=lambda i: monthly_totals[sorted_months[i]])] if sorted_months else '',
            'peak_units': round(max(monthly_totals.values()), 1) if monthly_totals else 0,
            'total_units': round(sum(monthly_totals.values()), 1),
        }

    def _extract_monthly_costs(self):
        """
        Spread activity resource costs across their duration into monthly buckets.
        Also separates planned (target_cost) vs actual (act_reg_cost + act_ot_cost).
        """
        activities = {str(a.get('task_id', '')): a for a in self.engine.activities}
        
        monthly_planned_cost = defaultdict(float)
        monthly_actual_cost = defaultdict(float)
        
        for res in self.engine.resources:
            task_id = str(res.get('task_id', ''))
            act = activities.get(task_id)
            if not act:
                continue
            
            planned_cost = self._to_float(res.get('target_cost', '0'))
            actual_cost = (
                self._to_float(res.get('act_reg_cost', '0'))
                + self._to_float(res.get('act_ot_cost', '0'))
                + self._to_float(res.get('act_cost', '0'))
            )
            
            if planned_cost <= 0 and actual_cost <= 0:
                continue
            
            start = act.get('target_start_date_parsed') or act.get('early_start_date_parsed')
            end = act.get('target_end_date_parsed') or act.get('early_end_date_parsed')
            if not start or not end or end <= start:
                continue
            
            total_days = max(1, (end - start).days)
            planned_per_day = planned_cost / total_days
            actual_per_day = actual_cost / total_days if actual_cost > 0 else 0
            
            current = start
            while current <= end:
                key = (current.year, current.month)
                monthly_planned_cost[key] += planned_per_day
                monthly_actual_cost[key] += actual_per_day
                current += timedelta(days=1)
        
        if not monthly_planned_cost and not monthly_actual_cost:
            self.resource_data['monthly_cost'] = {'labels': [], 'planned': [], 'actual': [], 'cumulative_planned': [], 'cumulative_actual': []}
            return
        
        all_months = sorted(set(list(monthly_planned_cost.keys()) + list(monthly_actual_cost.keys())))
        labels = [f"{y}-{m:02d}" for y, m in all_months]
        
        planned = [round(monthly_planned_cost.get(m, 0), 2) for m in all_months]
        actual = [round(monthly_actual_cost.get(m, 0), 2) for m in all_months]
        
        cumulative_planned = []
        cumulative_actual = []
        running_p = 0.0
        running_a = 0.0
        for i, _ in enumerate(all_months):
            running_p += planned[i]
            running_a += actual[i]
            cumulative_planned.append(round(running_p, 2))
            cumulative_actual.append(round(running_a, 2))
        
        # Total actuals count only up to data date
        has_actuals = any(a > 0 for a in actual)
        
        self.resource_data['monthly_cost'] = {
            'labels': labels,
            'planned': planned,
            'actual': actual if has_actuals else [],
            'cumulative_planned': cumulative_planned,
            'cumulative_actual': cumulative_actual if has_actuals else [],
            'total_planned': round(sum(planned), 2),
            'total_actual': round(sum(actual), 2) if has_actuals else 0,
            'peak_month_planned': labels[planned.index(max(planned))] if planned else '',
            'peak_planned_cost': round(max(planned), 2) if planned else 0,
        }

    def _extract_top_resources(self):
        """Top 20 resources by total man-hours and cost."""
        # {rsrc_id: {'units': X, 'cost': Y, 'name': N, 'type': T}}
        resource_totals = defaultdict(lambda: {'units': 0.0, 'cost': 0.0})
        
        for res in self.engine.resources:
            rid = str(res.get('rsrc_id', ''))
            if not rid:
                continue
            resource_totals[rid]['units'] += self._to_float(res.get('target_qty', '0'))
            resource_totals[rid]['cost'] += self._to_float(res.get('target_cost', '0'))
        
        # Convert to list with names
        resource_list = []
        for rid, totals in resource_totals.items():
            if totals['units'] > 0 or totals['cost'] > 0:
                resource_list.append({
                    'id': rid,
                    'name': self.resource_names.get(rid, f'Resource {rid}'),
                    'type': self.resource_types.get(rid, 'Other'),
                    'units': round(totals['units'], 1),
                    'cost': round(totals['cost'], 2),
                })
        
        # Sort by units (descending), take top 20
        top_by_units = sorted(resource_list, key=lambda x: x['units'], reverse=True)[:20]
        top_by_cost = sorted(resource_list, key=lambda x: x['cost'], reverse=True)[:20]
        
        self.resource_data['top_resources'] = {
            'by_units': top_by_units,
            'by_cost': top_by_cost,
            'total_count': len(resource_list),
        }

    def _to_float(self, value):
        try:
            return float(value)
        except (ValueError, TypeError):
            return 0.0
'''

with open("resource_engine.py", "w", encoding="utf-8") as f:
    f.write(RESOURCE_ENGINE_CODE)
print("  ✅ Created resource_engine.py")


# ==============================================================================
# FILE 2: app.py (Add /api/resource-data endpoint)
# ==============================================================================

APP_CODE = '''"""
P6 SCHEDULE ANALYZER - MAIN WEB APPLICATION (app.py)
=====================================================
Now with Resource Analytics Endpoint (Phase 1, Step 3)
"""

from flask import Flask, render_template, request, jsonify, send_file, session
from flask_cors import CORS
from werkzeug.utils import secure_filename
import os
import uuid
import time
import logging
from datetime import datetime

try:
    from config import (
        get_config, MAX_UPLOAD_SIZE_MB, SECRET_KEY, SESSION_LIFETIME_HOURS,
    )
except ImportError:
    MAX_UPLOAD_SIZE_MB = 100
    SECRET_KEY = 'dev-only-CHANGE-ME'
    SESSION_LIFETIME_HOURS = 24
    def get_config():
        return {
            'company_name': 'MK Constructions', 'app_title': 'P6 Schedule Analyzer',
            'app_subtitle': 'DCMA 14-Point Check & Analytics',
            'use_logo_image': False, 'theme': {'primary': '#1e40af', 'accent': '#3b82f6'},
            'features': {'gantt': True, 'comparison': True, 'evm': True, 'export': True, 'health': True}
        }

from parser import XERParser
from data_engine import ScheduleEngine
from reports import ReportGenerator

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

try:
    from comparison_engine import ScheduleComparator
    logger.info("✅ ScheduleComparator imported")
except Exception as e:
    ScheduleComparator = None
    logger.warning("❌ ScheduleComparator import failed: %s", e)

try:
    from evm_engine import EVMEngine
    logger.info("✅ EVMEngine imported")
except Exception as e:
    EVMEngine = None
    logger.warning("❌ EVMEngine import failed: %s", e)

try:
    from advanced_health_engine import AdvancedHealthEngine
    logger.info("✅ AdvancedHealthEngine imported")
except Exception as e:
    AdvancedHealthEngine = None
    logger.warning("❌ AdvancedHealthEngine import failed: %s", e)

try:
    from pdf_report_generator import PDFReportGenerator
    logger.info("✅ PDFReportGenerator imported")
except Exception as e:
    PDFReportGenerator = None
    logger.warning("❌ PDFReportGenerator import failed: %s", e)

try:
    from ai_narrative_engine import AINarrativeEngine
    logger.info("✅ AINarrativeEngine imported")
except Exception as e:
    AINarrativeEngine = None
    logger.warning("❌ AINarrativeEngine import failed: %s", e)

try:
    from resource_engine import ResourceEngine
    logger.info("✅ ResourceEngine imported")
except Exception as e:
    ResourceEngine = None
    logger.warning("❌ ResourceEngine import failed: %s", e)


app = Flask(__name__)
app.secret_key = SECRET_KEY
CORS(app)

UPLOAD_FOLDER = 'uploads'
OUTPUT_FOLDER = 'output'
ALLOWED_EXTENSIONS = {'xer'}

app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['OUTPUT_FOLDER'] = OUTPUT_FOLDER
app.config['MAX_CONTENT_LENGTH'] = MAX_UPLOAD_SIZE_MB * 1024 * 1024

os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(OUTPUT_FOLDER, exist_ok=True)


def cleanup_old_files(folder, max_age_hours=SESSION_LIFETIME_HOURS):
    if not os.path.exists(folder):
        return
    cutoff = time.time() - (max_age_hours * 3600)
    for fname in os.listdir(folder):
        fpath = os.path.join(folder, fname)
        try:
            if os.path.isfile(fpath) and os.path.getmtime(fpath) < cutoff:
                os.remove(fpath)
                logger.info("🧹 Cleaned up old file: %s", fname)
        except Exception as e:
            logger.warning("Could not delete %s: %s", fname, e)

cleanup_old_files(UPLOAD_FOLDER)
cleanup_old_files(OUTPUT_FOLDER)


SESSION_STORAGE = {}

def get_session_data():
    sid = session.get('sid')
    if not sid or sid not in SESSION_STORAGE:
        sid = uuid.uuid4().hex
        session['sid'] = sid
        SESSION_STORAGE[sid] = {
            'analysis': {'engine': None, 'dashboard_data': None, 'file_name': None, 'analyzed_at': None},
            'comparison': {'comparator': None, 'results': None, 'baseline_file': None, 'current_file': None},
            'health_cache': {},
        }
    return SESSION_STORAGE[sid]


def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


def analyze_xer_file(file_path_or_stream, original_filename, session_data):
    logger.info("🔍 Analyzing XER: %s", original_filename)
    parser = XERParser()
    tables = parser.parse(file_path_or_stream)
    if tables is None or not tables:
        return {'error': 'Failed to parse XER file or file is empty.'}
    engine = ScheduleEngine()
    engine.load_data(tables)
    engine.analyze()
    dashboard_data = engine.get_dashboard_data()
    analysis = session_data['analysis']
    analysis['engine'] = engine
    analysis['dashboard_data'] = dashboard_data
    analysis['file_name'] = original_filename
    analysis['analyzed_at'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    session_data['health_cache'] = {}
    return dashboard_data


@app.context_processor
def inject_config():
    return {'config': get_config()}


@app.errorhandler(413)
def request_entity_too_large(error):
    return jsonify({'error': f'File size exceeds maximum limit of {MAX_UPLOAD_SIZE_MB} MB.'}), 413


@app.route('/')
def home():
    return render_template('index.html')


@app.route('/api/upload', methods=['POST'])
def upload_file():
    if 'file' not in request.files:
        return jsonify({'error': 'No file provided'}), 400
    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'No file selected'}), 400
    if not allowed_file(file.filename):
        return jsonify({'error': 'File must be a .xer file'}), 400
    original_name = secure_filename(file.filename)
    unique_name = f"{uuid.uuid4().hex}_{original_name}"
    file_path = os.path.join(app.config['UPLOAD_FOLDER'], unique_name)
    file.save(file_path)
    sess_data = get_session_data()
    try:
        dashboard_data = analyze_xer_file(file_path, original_name, sess_data)
        if 'error' in dashboard_data:
            return jsonify(dashboard_data), 400
        analysis = sess_data['analysis']
        return jsonify({
            'success': True, 'file_name': analysis['file_name'],
            'analyzed_at': analysis['analyzed_at'], 'data': dashboard_data
        })
    except Exception as e:
        logger.exception("Upload analysis error")
        return jsonify({'error': str(e)}), 500


@app.route('/api/dashboard')
def get_dashboard():
    sess_data = get_session_data()
    analysis = sess_data['analysis']
    if analysis['dashboard_data'] is None:
        return jsonify({'has_data': False})
    return jsonify({
        'has_data': True, 'file_name': analysis['file_name'],
        'analyzed_at': analysis['analyzed_at'], 'data': analysis['dashboard_data']
    })


@app.route('/api/load-sample')
def load_sample():
    sample_path = os.path.join('input', 'sample.xer')
    if not os.path.exists(sample_path):
        return jsonify({'error': 'sample.xer not found in input/ folder'}), 404
    sess_data = get_session_data()
    try:
        dashboard_data = analyze_xer_file(sample_path, 'sample.xer', sess_data)
        analysis = sess_data['analysis']
        return jsonify({
            'success': True, 'file_name': 'sample.xer',
            'analyzed_at': analysis['analyzed_at'], 'data': dashboard_data
        })
    except Exception as e:
        logger.exception("Sample load error")
        return jsonify({'error': str(e)}), 500


@app.route('/api/export-excel')
def export_excel():
    sess_data = get_session_data()
    engine = sess_data['analysis']['engine']
    if engine is None:
        return jsonify({'error': 'No schedule loaded. Please upload an XER file first.'}), 400
    out_name = f"schedule_report_{uuid.uuid4().hex[:8]}.xlsx"
    output_path = os.path.join(app.config['OUTPUT_FOLDER'], out_name)
    reporter = ReportGenerator(engine)
    res_path = reporter.generate_full_report(output_path)
    if not res_path or not os.path.exists(res_path):
        return jsonify({'error': 'Failed to generate Excel report.'}), 500
    return send_file(
        res_path, as_attachment=True,
        download_name=f"schedule_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


@app.route('/gantt')
def gantt_view():
    return render_template('gantt.html')


@app.route('/api/gantt-data')
def get_gantt_data():
    sess_data = get_session_data()
    analysis = sess_data['analysis']
    if analysis['engine'] is None:
        return jsonify({'error': 'No data loaded. Please upload an XER file first.'}), 400
    max_acts = request.args.get('max', 2000, type=int)
    gantt_data = analysis['engine'].get_gantt_data(max_activities=max_acts)
    return jsonify({
        'success': True, 'file_name': analysis['file_name'], 'data': gantt_data
    })


@app.route('/comparison')
def comparison_view():
    return render_template('comparison.html')


@app.route('/api/compare', methods=['POST'])
def compare_schedules():
    if ScheduleComparator is None:
        return jsonify({'error': 'comparison_engine.py is missing!'}), 500
    if 'baseline' not in request.files or 'current' not in request.files:
        return jsonify({'error': 'Both baseline and current files required'}), 400
    baseline_file = request.files['baseline']
    current_file = request.files['current']
    if not (allowed_file(baseline_file.filename) and allowed_file(current_file.filename)):
        return jsonify({'error': 'Both files must be .xer files'}), 400
    baseline_name = secure_filename(baseline_file.filename)
    current_name = secure_filename(current_file.filename)
    baseline_path = os.path.join(app.config['UPLOAD_FOLDER'], f"bl_{uuid.uuid4().hex[:8]}_{baseline_name}")
    current_path = os.path.join(app.config['UPLOAD_FOLDER'], f"cur_{uuid.uuid4().hex[:8]}_{current_name}")
    baseline_file.save(baseline_path)
    current_file.save(current_path)
    sess_data = get_session_data()
    try:
        comparator = ScheduleComparator()
        comparator.load_baseline(baseline_path)
        comparator.load_current(current_path)
        results = comparator.compare()
        comp = sess_data['comparison']
        comp['comparator'] = comparator
        comp['results'] = results
        comp['baseline_file'] = baseline_name
        comp['current_file'] = current_name
        return jsonify({
            'success': True, 'baseline_file': baseline_name,
            'current_file': current_name, 'results': results
        })
    except Exception as e:
        logger.exception("Comparison error")
        return jsonify({'error': str(e)}), 500


@app.route('/api/comparison-data')
def get_comparison_data():
    sess_data = get_session_data()
    comp = sess_data['comparison']
    if comp['results'] is None:
        return jsonify({'has_data': False})
    return jsonify({
        'has_data': True, 'baseline_file': comp['baseline_file'],
        'current_file': comp['current_file'], 'results': comp['results']
    })


@app.route('/evm')
def evm_view():
    return render_template('evm.html')


@app.route('/api/evm-data')
def get_evm_data():
    sess_data = get_session_data()
    analysis = sess_data['analysis']
    if analysis['engine'] is None:
        return jsonify({'error': 'No data loaded. Upload an XER file first.'}), 400
    if EVMEngine is None:
        return jsonify({'error': 'evm_engine.py is missing!'}), 500
    try:
        evm = EVMEngine(analysis['engine'])
        results = evm.calculate()
        return jsonify({
            'success': True, 'file_name': analysis['file_name'], 'data': results
        })
    except Exception as e:
        logger.exception("EVM calculation error")
        return jsonify({'error': str(e)}), 500


# ════════════════════════════════════════════
# NEW ROUTE: RESOURCE ANALYTICS
# ════════════════════════════════════════════

@app.route('/api/resource-data')
def get_resource_data():
    sess_data = get_session_data()
    analysis = sess_data['analysis']
    if analysis['engine'] is None:
        return jsonify({'error': 'No data loaded. Upload an XER file first.'}), 400
    if ResourceEngine is None:
        return jsonify({'error': 'resource_engine.py is missing!'}), 500
    try:
        engine = ResourceEngine(analysis['engine'])
        results = engine.calculate()
        return jsonify({
            'success': True, 'file_name': analysis['file_name'], 'data': results
        })
    except Exception as e:
        logger.exception("Resource analytics error")
        return jsonify({'error': str(e)}), 500


@app.route('/health')
def health_view():
    return render_template('health.html')


@app.route('/api/health-data')
def get_health_data():
    sess_data = get_session_data()
    analysis = sess_data['analysis']
    if analysis['engine'] is None:
        return jsonify({'error': 'No data loaded. Upload a file first.'}), 400
    if AdvancedHealthEngine is None:
        return jsonify({'error': 'advanced_health_engine.py is missing!'}), 500
    selected_standard = request.args.get('standard', 'all')
    try:
        health = AdvancedHealthEngine(analysis['engine'])
        results = health.run_all_checks(selected_standard=selected_standard)
        return jsonify({
            'success': True, 'file_name': analysis['file_name'], 'data': results
        })
    except Exception as e:
        logger.exception("Health analysis error")
        return jsonify({'error': str(e)}), 500


@app.route('/api/executive-pdf')
def download_executive_pdf():
    sess_data = get_session_data()
    analysis = sess_data['analysis']
    if analysis['engine'] is None:
        return jsonify({'error': 'No data loaded'}), 400
    if PDFReportGenerator is None or AdvancedHealthEngine is None:
        return jsonify({'error': 'PDF or Health engine module missing!'}), 500
    selected_standard = request.args.get('standard', 'all')
    severity_filter = request.args.get('severity', 'all')
    try:
        health = AdvancedHealthEngine(analysis['engine'])
        results = health.run_all_checks(selected_standard=selected_standard)
        generator = PDFReportGenerator(results, analysis['file_name'], severity_filter=severity_filter)
        pdf_buffer = generator.generate_executive_report()
        return send_file(
            pdf_buffer, as_attachment=True,
            download_name=f"executive_report_{selected_standard}_{severity_filter}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf",
            mimetype='application/pdf'
        )
    except Exception as e:
        logger.exception("PDF generation error")
        return jsonify({'error': str(e)}), 500


@app.route('/api/actions-pdf')
def download_actions_pdf():
    sess_data = get_session_data()
    analysis = sess_data['analysis']
    if analysis['engine'] is None:
        return jsonify({'error': 'No data loaded'}), 400
    if PDFReportGenerator is None or AdvancedHealthEngine is None:
        return jsonify({'error': 'PDF or Health engine module missing!'}), 500
    selected_standard = request.args.get('standard', 'all')
    severity_filter = request.args.get('severity', 'all')
    try:
        health = AdvancedHealthEngine(analysis['engine'])
        results = health.run_all_checks(selected_standard=selected_standard)
        generator = PDFReportGenerator(results, analysis['file_name'], severity_filter=severity_filter)
        pdf_buffer = generator.generate_actions_report()
        return send_file(
            pdf_buffer, as_attachment=True,
            download_name=f"action_list_{selected_standard}_{severity_filter}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf",
            mimetype='application/pdf'
        )
    except Exception as e:
        logger.exception("PDF actions error")
        return jsonify({'error': str(e)}), 500


@app.route('/api/actions-excel')
def download_actions_excel():
    sess_data = get_session_data()
    analysis = sess_data['analysis']
    if analysis['engine'] is None:
        return jsonify({'error': 'No data loaded. Upload a file first.'}), 400
    if AdvancedHealthEngine is None:
        return jsonify({'error': 'advanced_health_engine.py is missing!'}), 500
    selected_standard = request.args.get('standard', 'all')
    severity_filter = request.args.get('severity', 'all').lower()
    severity_levels = {
        'critical': ['critical'], 'high': ['critical', 'high'],
        'medium': ['critical', 'high', 'medium'],
        'all': ['critical', 'high', 'medium', 'low', 'info']
    }
    allowed_severities = severity_levels.get(severity_filter, severity_levels['all'])
    try:
        import pandas as pd
        from io import BytesIO
        from openpyxl.styles import Font, PatternFill, Alignment
        health = AdvancedHealthEngine(analysis['engine'])
        results = health.run_all_checks(selected_standard=selected_standard)
        top_actions = results.get('top_actions', []) or []
        standards_data = results.get('standards', {}) or {}
        filtered_top_actions = [
            a for a in top_actions
            if (a.get('severity') or 'low').lower() in allowed_severities
        ]
        meta_rows = [
            ['Report Type', 'Schedule Health - Top Actions Export'],
            ['Selected Standard', selected_standard],
            ['Severity Filter', severity_filter.upper()],
            ['File Name', analysis.get('file_name', '')],
            ['Generated At', results.get('analysis_date', '')],
            ['', ''],
            ['Overall Score', results.get('overall_score', '')],
            ['Total Checks', results.get('total_checks', '')],
            ['Passed Checks', results.get('passed_checks', '')],
            ['Failed Checks', results.get('failed_checks', '')],
            ['Critical Failures', results.get('critical_failures', '')],
            ['High Failures', results.get('high_failures', '')],
            ['Pass Rate (%)', results.get('pass_rate', '')],
            ['', ''],
            ['Filtered Actions Count', len(filtered_top_actions)],
        ]
        top_summary_rows = []
        for idx, action in enumerate(filtered_top_actions, 1):
            top_summary_rows.append({
                'Rank': idx, 'Standard': action.get('standard', ''),
                'Check ID': action.get('id', ''), 'Check Name': action.get('name', ''),
                'Category': action.get('category', ''),
                'Severity': (action.get('severity') or '').upper(),
                'Affected Count': action.get('count', 0), 'Total': action.get('total', 0),
                'Percentage': action.get('percentage', 0), 'Threshold': action.get('threshold', ''),
                'Value': action.get('value', ''), 'Recommendation': action.get('recommendation', ''),
                'Description': action.get('description', ''),
            })
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            pd.DataFrame(meta_rows, columns=['Field', 'Value']).to_excel(writer, sheet_name='Report Info', index=False)
            if top_summary_rows:
                pd.DataFrame(top_summary_rows).to_excel(writer, sheet_name='Top Actions Summary', index=False)
            else:
                pd.DataFrame([{'Info': f'No actions matched severity filter: {severity_filter.upper()}'}]).to_excel(writer, sheet_name='Top Actions Summary', index=False)
            for std_name, std_data in standards_data.items():
                sheet_name = (std_name or 'Standard')[:31]
                rows = []
                failed_in_std = []
                for category in std_data.get('categories', []):
                    for check in category.get('checks', []):
                        if check.get('status') != 'fail': continue
                        if (check.get('severity') or 'low').lower() not in allowed_severities: continue
                        failed_in_std.append((category.get('name', ''), check))
                if not failed_in_std:
                    rows.append({'Section': 'No Matching Failures', 'Field': '', 'Value': f'No {severity_filter.upper()} failures found for {std_name}', 'Activity ID': '', 'Activity Name': '', 'WBS': ''})
                else:
                    for cat_name, check in failed_in_std:
                        rows.append({'Section': 'METRIC', 'Field': 'Check ID', 'Value': check.get('id', ''), 'Activity ID': '', 'Activity Name': '', 'WBS': ''})
                        rows.append({'Section': '', 'Field': 'Check Name', 'Value': check.get('name', ''), 'Activity ID': '', 'Activity Name': '', 'WBS': ''})
                        rows.append({'Section': '', 'Field': 'Category', 'Value': cat_name, 'Activity ID': '', 'Activity Name': '', 'WBS': ''})
                        rows.append({'Section': '', 'Field': 'Severity', 'Value': (check.get('severity') or '').upper(), 'Activity ID': '', 'Activity Name': '', 'WBS': ''})
                        rows.append({'Section': '', 'Field': 'Description', 'Value': check.get('description', ''), 'Activity ID': '', 'Activity Name': '', 'WBS': ''})
                        rows.append({'Section': '', 'Field': 'Threshold', 'Value': check.get('threshold', ''), 'Activity ID': '', 'Activity Name': '', 'WBS': ''})
                        if check.get('count') is not None:
                            rows.append({'Section': '', 'Field': 'Affected', 'Value': f"{check.get('count', 0)} of {check.get('total', 0)} ({check.get('percentage', 0)}%)", 'Activity ID': '', 'Activity Name': '', 'WBS': ''})
                        if check.get('value') is not None and check.get('value') != '':
                            rows.append({'Section': '', 'Field': 'Value', 'Value': check.get('value', ''), 'Activity ID': '', 'Activity Name': '', 'WBS': ''})
                        rows.append({'Section': '', 'Field': 'Recommendation', 'Value': check.get('recommendation', ''), 'Activity ID': '', 'Activity Name': '', 'WBS': ''})
                        items = check.get('failed_items', []) or []
                        rows.append({'Section': 'AFFECTED ACTIVITIES', 'Field': f'Total: {len(items)}', 'Value': '', 'Activity ID': 'Activity ID', 'Activity Name': 'Activity Name', 'WBS': 'WBS'})
                        if items:
                            for item in items:
                                rows.append({'Section': '', 'Field': '', 'Value': '', 'Activity ID': item.get('code', ''), 'Activity Name': item.get('name', ''), 'WBS': item.get('wbs', '')})
                        else:
                            rows.append({'Section': '', 'Field': '', 'Value': '(No activity list available)', 'Activity ID': '', 'Activity Name': '', 'WBS': ''})
                        rows.append({'Section': '', 'Field': '', 'Value': '', 'Activity ID': '', 'Activity Name': '', 'WBS': ''})
                pd.DataFrame(rows).to_excel(writer, sheet_name=sheet_name, index=False)
            workbook = writer.book
            header_font = Font(bold=True, color='FFFFFF', size=11)
            header_fill = PatternFill(start_color='1E40AF', end_color='1E40AF', fill_type='solid')
            metric_fill = PatternFill(start_color='FEF3C7', end_color='FEF3C7', fill_type='solid')
            activity_hdr_fill = PatternFill(start_color='DBEAFE', end_color='DBEAFE', fill_type='solid')
            bold_font = Font(bold=True)
            wrap_align = Alignment(wrap_text=True, vertical='top')
            for sheet_name in workbook.sheetnames:
                ws = workbook[sheet_name]
                for cell in ws[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal='left', vertical='center')
                for col in ws.columns:
                    max_len = 10
                    col_letter = col[0].column_letter
                    for cell in col:
                        try:
                            val = str(cell.value) if cell.value is not None else ''
                            if len(val) > max_len: max_len = len(val)
                        except Exception: pass
                    ws.column_dimensions[col_letter].width = min(max_len + 2, 60)
                ws.freeze_panes = 'A2'
                if sheet_name not in ['Report Info', 'Top Actions Summary']:
                    for row in ws.iter_rows(min_row=2):
                        section_cell = row[0]
                        if section_cell.value == 'METRIC':
                            for c in row: c.fill = metric_fill; c.font = bold_font
                        elif section_cell.value == 'AFFECTED ACTIVITIES':
                            for c in row: c.fill = activity_hdr_fill; c.font = bold_font
                        for c in row: c.alignment = wrap_align
        output.seek(0)
        filename = f"health_top_actions_{selected_standard}_{severity_filter}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return send_file(output, as_attachment=True, download_name=filename, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    except Exception as e:
        logger.exception("Actions Excel error")
        return jsonify({'error': str(e)}), 500


@app.route('/api/ai-narrative')
def get_ai_narrative():
    sess_data = get_session_data()
    analysis = sess_data['analysis']
    if analysis['engine'] is None:
        return jsonify({'error': 'No schedule loaded. Upload an XER file first.'}), 400
    if AINarrativeEngine is None:
        return jsonify({'error': 'ai_narrative_engine.py is missing!'}), 500
    try:
        health_data = {}
        if AdvancedHealthEngine is not None:
            try:
                health_gen = AdvancedHealthEngine(analysis['engine'])
                health_data = health_gen.run_all_checks('all')
            except Exception as he:
                logger.warning("Health data unavailable for narrative: %s", he)
        evm_data = {}
        if EVMEngine is not None:
            try:
                evm_data = EVMEngine(analysis['engine']).calculate()
            except Exception: pass
        comp_data = sess_data.get('comparison', {}).get('results', {}) or {}
        narrative_gen = AINarrativeEngine(health_data=health_data, comparison_data=comp_data, evm_data=evm_data)
        results = narrative_gen.generate_narrative()
        return jsonify({'success': True, 'file_name': analysis['file_name'], 'data': results})
    except Exception as e:
        logger.exception("AI Narrative generation error")
        return jsonify({'error': str(e)}), 500


if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    debug_mode = os.environ.get('FLASK_ENV') != 'production'
    logger.info("============================================================")
    logger.info("🚀 P6 SCHEDULE ANALYZER - AI + RESOURCES + FULL SUITE READY")
    logger.info("============================================================")
    logger.info("📌 Running on port %s", port)
    logger.info("🔧 Debug mode: %s", debug_mode)
    logger.info("👉 Open in browser: http://localhost:%s", port)
    app.run(debug=debug_mode, host='0.0.0.0', port=port)
'''

with open("app.py", "w", encoding="utf-8") as f:
    f.write(APP_CODE)
print("  ✅ Updated app.py")


# ==============================================================================
# FILE 3: templates/evm.html (Full Rewrite with Resource Analytics Section)
# ==============================================================================

EVM_HTML_CODE = '''<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>EVM &amp; S-Curves | {{ config.app_title }}</title>
    
    <script src="https://cdn.jsdelivr.net/npm/chart.js"></script>

    <style>
        :root {
            --color-primary: {{ config.theme.primary }};
            --color-primary-dark: {{ config.theme.primary_dark }};
            --color-accent: {{ config.theme.accent }};
            --color-success: {{ config.theme.success }};
            --color-warning: {{ config.theme.warning }};
            --color-danger: {{ config.theme.danger }};
            --color-bg: {{ config.theme.bg }};
            --color-surface: {{ config.theme.surface }};
            --color-text: {{ config.theme.text }};
            --color-muted: {{ config.theme.muted }};
            --color-border: {{ config.theme.border }};
        }
    </style>
    
    <link rel="stylesheet" href="{{ url_for('static', filename='style.css') }}">
    
    <style>
        .evm-metrics {
            display: grid; grid-template-columns: repeat(auto-fit, minmax(200px, 1fr));
            gap: 1rem; margin-bottom: 2rem;
        }
        .metric-card {
            background: var(--color-surface); padding: 1.25rem;
            border-radius: var(--radius-lg, 12px); box-shadow: var(--shadow-md, 0 2px 4px rgba(0,0,0,0.05));
            border-top: 4px solid var(--color-accent);
            border-left: 1px solid var(--color-border); border-right: 1px solid var(--color-border);
            border-bottom: 1px solid var(--color-border);
        }
        .metric-card.good { border-top-color: var(--color-success); }
        .metric-card.warning { border-top-color: var(--color-warning); }
        .metric-card.bad { border-top-color: var(--color-danger); }
        .metric-card.neutral { border-top-color: var(--color-accent); }
        .metric-label { color: var(--color-muted); font-size: 0.85rem; margin-bottom: 0.5rem; }
        .metric-value { font-size: 1.75rem; font-weight: 700; color: var(--color-text); }
        .metric-subtitle { font-size: 0.85rem; color: var(--color-muted); margin-top: 0.25rem; }
        .status-good { color: var(--color-success); font-weight: 600; }
        .status-warning { color: var(--color-warning); font-weight: 600; }
        .status-bad { color: var(--color-danger); font-weight: 600; }
        .status-neutral { color: var(--color-muted); }
        
        a.btn { text-decoration: none; }
        .app-header .btn-secondary[aria-current="page"] {
            background: rgba(255, 255, 255, 0.4); border-color: #fff; font-weight: 700;
        }

        /* Resource Analytics Section */
        .resource-section {
            background: var(--color-surface); border: 1px solid var(--color-border);
            border-radius: 12px; padding: 1.5rem; margin-bottom: 2rem;
        }
        .resource-grid {
            display: grid; grid-template-columns: 1fr 1fr; gap: 1.5rem; margin-top: 1rem;
        }
        @media (max-width: 900px) {
            .resource-grid { grid-template-columns: 1fr; }
        }
        .resource-stat-box {
            display: grid; grid-template-columns: repeat(auto-fit, minmax(140px, 1fr));
            gap: 0.75rem; margin: 1rem 0;
        }
        .resource-stat {
            background: #f8fafc; border: 1px solid var(--color-border);
            border-radius: 8px; padding: 0.75rem; text-align: center;
        }
        .resource-stat .val {
            font-size: 1.25rem; font-weight: 700; color: var(--color-primary);
        }
        .resource-stat .lbl {
            font-size: 0.7rem; color: var(--color-muted);
            text-transform: uppercase; letter-spacing: 0.02em; margin-top: 0.25rem;
        }
        .chart-wrapper {
            position: relative; height: 350px; padding: 1rem;
            background: #fff; border: 1px solid var(--color-border); border-radius: 8px;
        }
        .chart-wrapper h4 {
            font-size: 0.95rem; margin-bottom: 0.5rem; color: var(--color-text);
        }
        .top-resource-list {
            max-height: 400px; overflow-y: auto;
            background: #fff; border: 1px solid var(--color-border);
            border-radius: 8px; padding: 0.5rem;
        }
        .top-resource-item {
            display: flex; justify-content: space-between; align-items: center;
            padding: 0.5rem; border-bottom: 1px solid var(--color-border);
            font-size: 0.85rem;
        }
        .top-resource-item:last-child { border-bottom: none; }
        .top-resource-item .name { font-weight: 600; color: var(--color-text); }
        .top-resource-item .type {
            font-size: 0.7rem; padding: 0.15rem 0.4rem; border-radius: 999px;
            background: #e0e7ff; color: #3730a3; margin-left: 0.5rem;
        }
        .top-resource-item .val { font-weight: 700; color: var(--color-primary); }
    </style>
</head>
<body>
    <header class="app-header">
        <div class="header-content">
            <div class="logo-section">
                {% if config.use_logo_image %}
                    <img src="{{ url_for('static', filename=config.logo_path) }}" alt="Logo" style="height: 45px;">
                {% else %}
                    <span class="logo-icon">📈</span>
                {% endif %}
                <div>
                    <h1>{{ config.app_title }} — EVM</h1>
                    <p class="subtitle">Earned Value Management &amp; Resource Analytics</p>
                </div>
            </div>
            <div class="header-actions">
                <a href="/" class="btn btn-secondary">📊 Dashboard</a>
                <a href="/gantt" class="btn btn-secondary">📅 Gantt</a>
                <a href="/comparison" class="btn btn-secondary">🔄 Compare</a>
                <a href="/evm" class="btn btn-secondary" aria-current="page">📈 EVM</a>
                <a href="/health" class="btn btn-secondary">🏥 Health</a>
            </div>
        </div>
    </header>

    <main class="app-main">
        <div id="loadingMessage" class="loading-screen" style="display:flex;">
            <div class="spinner"></div>
            <p>Loading EVM data...</p>
        </div>

        <div id="evmContent" style="display:none;">
            <div class="file-info-bar">
                <span>📁 <strong id="fileName">--</strong></span>
                <span>📅 Data Date: <strong id="dataDate">--</strong></span>
            </div>

            <h2 style="margin-bottom:1rem;">📈 Performance Indicators</h2>
            <div class="evm-metrics" id="performanceMetrics"></div>

            <h2 style="margin-bottom:1rem;">💰 Financial Metrics</h2>
            <div class="evm-metrics" id="financialMetrics"></div>

            <div class="chart-card" style="margin-bottom:2rem;">
                <h3>📈 Project S-Curve</h3>
                <p style="color:var(--color-muted); font-size:0.85rem; margin-bottom:1rem;">
                    Cumulative value over time showing Planned Value (PV), Earned Value (EV), and Actual Cost (AC) through Data Date.
                </p>
                <div id="scurveError" style="display:none; color:var(--color-danger); padding:0.5rem 0;"></div>
                <div style="height:400px; position:relative;">
                    <canvas id="scurveChart"></canvas>
                </div>
            </div>

            <!-- ═══════════════════════════════════════ -->
            <!-- 👷 RESOURCE ANALYTICS SECTION -->
            <!-- ═══════════════════════════════════════ -->
            <div class="resource-section" id="resourceSection" style="display:none;">
                <div style="display: flex; justify-content: space-between; align-items: center; margin-bottom: 1rem; flex-wrap: wrap; gap: 0.5rem;">
                    <h2 style="margin: 0;">👷 Resource &amp; Labor Analytics</h2>
                    <button type="button" class="btn btn-primary" onclick="loadResourceData(true)">
                        🔄 Refresh
                    </button>
                </div>
                <p style="color:var(--color-muted); font-size:0.85rem; margin-bottom:1rem;">
                    Monthly labor hours (stacked by resource type), cost burn rate, and top resource drivers.
                </p>

                <div id="resourceErrorBox" style="display:none; padding: 1rem; background: #fef2f2; border-radius: 8px; color: var(--color-danger); margin-bottom: 1rem;"></div>

                <div class="resource-stat-box" id="resourceStats"></div>

                <!-- Monthly Man-Hours Histogram -->
                <div class="chart-wrapper" style="margin-bottom: 1.5rem;">
                    <h4>📊 Monthly Man-Hours by Resource Type</h4>
                    <canvas id="monthlyUnitsChart"></canvas>
                </div>

                <!-- Monthly Cost Burn Rate -->
                <div class="chart-wrapper" style="margin-bottom: 1.5rem;">
                    <h4>💰 Monthly Cost Burn Rate (Planned vs Actual)</h4>
                    <canvas id="monthlyCostChart"></canvas>
                </div>

                <div class="resource-grid">
                    <!-- Cumulative Units Curve -->
                    <div class="chart-wrapper">
                        <h4>📈 Cumulative Man-Hours Curve</h4>
                        <canvas id="cumulativeUnitsChart"></canvas>
                    </div>
                    <!-- Top Resources by Units -->
                    <div>
                        <h4 style="font-size: 0.95rem; margin-bottom: 0.5rem; color: var(--color-text);">
                            🏆 Top 20 Resources by Man-Hours
                        </h4>
                        <div class="top-resource-list" id="topResourcesUnits"></div>
                    </div>
                </div>
            </div>

            <div class="dcma-section">
                <h3>📖 What Do These Metrics Mean?</h3>
                <div style="display:grid; grid-template-columns:repeat(auto-fit, minmax(260px, 1fr)); gap:1.25rem; margin-top:1rem;">
                    <div>
                        <h4>SPI (Schedule Performance Index)</h4>
                        <p style="font-size:0.875rem; color:var(--color-muted); margin-top:0.25rem;">
                            <strong>≥ 1.0:</strong> Ahead of schedule ✅<br>
                            <strong>0.95 – 0.99:</strong> Slightly behind ⚠️<br>
                            <strong>&lt; 0.95:</strong> Significantly behind ❌
                        </p>
                    </div>
                    <div>
                        <h4>CPI (Cost Performance Index)</h4>
                        <p style="font-size:0.875rem; color:var(--color-muted); margin-top:0.25rem;">
                            <strong>≥ 1.0:</strong> Under budget ✅<br>
                            <strong>0.95 – 0.99:</strong> Slightly over budget ⚠️<br>
                            <strong>&lt; 0.95:</strong> Significantly over budget ❌
                        </p>
                    </div>
                    <div>
                        <h4>Peak Labor Month</h4>
                        <p style="font-size:0.875rem; color:var(--color-muted); margin-top:0.25rem;">
                            The month with the highest planned man-hour demand. Useful for resource leveling and crew planning.
                        </p>
                    </div>
                    <div>
                        <h4>Cumulative Curves</h4>
                        <p style="font-size:0.875rem; color:var(--color-muted); margin-top:0.25rem;">
                            Show cumulative labor buildup over time — indicator of ramp-up, peak execution, and demobilization phases.
                        </p>
                    </div>
                </div>
            </div>
        </div>
    </main>

    <footer class="app-footer">
        <p>&copy; {{ config.footer_year }} {{ config.company_name }} | {{ config.footer_text }}</p>
    </footer>

    <script src="{{ url_for('static', filename='evm.js') }}"></script>
</body>
</html>
'''

os.makedirs("templates", exist_ok=True)
with open("templates/evm.html", "w", encoding="utf-8") as f:
    f.write(EVM_HTML_CODE)
print("  ✅ Updated templates/evm.html")


# ==============================================================================
# FILE 4: static/evm.js (Full Rewrite with Resource Chart Renderers)
# ==============================================================================

EVM_JS_CODE = '''/*
    EVM PAGE LOGIC + RESOURCE ANALYTICS (Phase 1, Step 3)
    ======================================================
*/

let scurveChart = null;
let monthlyUnitsChart = null;
let monthlyCostChart = null;
let cumulativeUnitsChart = null;

document.addEventListener('DOMContentLoaded', function () {
    loadEVM();
});

function esc(s) {
    if (s == null) return '';
    return String(s)
        .replace(/&/g, '&amp;').replace(/</g, '&lt;')
        .replace(/>/g, '&gt;').replace(/"/g, '&quot;')
        .replace(/'/g, '&#39;');
}

function statusClass(statusObj) {
    const st = (statusObj && statusObj.status) ? String(statusObj.status) : 'neutral';
    if (st === 'good' || st === 'warning' || st === 'bad') return st;
    return 'neutral';
}

function fmtMoney(val) {
    const n = Number(val);
    if (!Number.isFinite(n)) return '—';
    try {
        return new Intl.NumberFormat('en-US', {
            style: 'currency', currency: 'USD', maximumFractionDigits: 0
        }).format(n);
    } catch (e) {
        return '$' + Math.round(n).toLocaleString('en-US');
    }
}

function fmtNum(val) {
    const n = Number(val);
    if (!Number.isFinite(n)) return '—';
    return n.toLocaleString('en-US', { maximumFractionDigits: 0 });
}

function fmtIndex(val) {
    const n = Number(val);
    if (!Number.isFinite(n)) return 'N/A';
    return n.toFixed(3);
}

function fmtPct(val) {
    const n = Number(val);
    if (!Number.isFinite(n)) return '—';
    return n.toFixed(1) + '%';
}

function normalizeSeries(arr) {
    return (arr || []).map(function (v) {
        if (v === null || v === undefined) return null;
        const n = Number(v);
        return Number.isFinite(n) ? n : null;
    });
}

function showEvmError(msg) {
    const loading = document.getElementById('loadingMessage');
    const content = document.getElementById('evmContent');
    if (content) content.style.display = 'none';
    if (loading) {
        loading.style.display = 'block';
        loading.innerHTML =
            '<div style="text-align:center;padding:2rem;">' +
            '<p style="color:#dc2626;">❌ ' + esc(msg) + '</p>' +
            '<p style="color:#64748b;margin-top:0.5rem;">Upload an XER on the Dashboard first.</p>' +
            '<a href="/" class="btn btn-primary" style="margin-top:1rem;display:inline-flex;">← Dashboard</a>' +
            '</div>';
    }
}

async function loadEVM() {
    const loading = document.getElementById('loadingMessage');
    if (loading) {
        loading.style.display = 'block';
        loading.innerHTML = '<p style="text-align:center;padding:2rem;">Loading EVM data...</p>';
    }

    try {
        const res = await fetch('/api/evm-data');
        let response = {};
        try {
            response = await res.json();
        } catch (e) {
            throw new Error(res.ok ? 'Invalid JSON from server' : ('Request failed (' + res.status + ')'));
        }
        if (!res.ok || response.error) {
            throw new Error(response.error || ('Failed to load EVM (' + res.status + ')'));
        }
        const metrics = response.data && response.data.metrics;
        const scurve = (response.data && response.data.scurve) || {};
        if (!metrics) throw new Error('Invalid EVM payload (missing metrics)');

        if (loading) loading.style.display = 'none';
        const content = document.getElementById('evmContent');
        if (content) content.style.display = 'block';

        const fileName = document.getElementById('fileName');
        const dataDate = document.getElementById('dataDate');
        if (fileName) fileName.textContent = response.file_name || '—';
        if (dataDate) dataDate.textContent = metrics.data_date || '—';

        renderPerformanceMetrics(metrics);
        renderFinancialMetrics(metrics);
        renderScurve(scurve, metrics);
        
        // Load resource analytics after EVM
        loadResourceData(false);
    } catch (err) {
        console.error(err);
        showEvmError(err.message || 'Failed to load EVM data');
    }
}

function renderPerformanceMetrics(m) {
    const container = document.getElementById('performanceMetrics');
    if (!container) return;
    const sSt = statusClass(m.schedule_status);
    const cSt = statusClass(m.cost_status);
    const sText = (m.schedule_status && m.schedule_status.text) ? m.schedule_status.text : '';
    const cText = (m.cost_status && m.cost_status.text) ? m.cost_status.text : '';

    container.innerHTML =
        '<div class="metric-card ' + sSt + '">' +
        '<div class="metric-label">SPI (Schedule)</div>' +
        '<div class="metric-value">' + esc(fmtIndex(m.spi)) + '</div>' +
        '<div class="metric-subtitle status-' + sSt + '">' + esc(sText) + '</div></div>' +

        '<div class="metric-card ' + cSt + '">' +
        '<div class="metric-label">CPI (Cost)</div>' +
        '<div class="metric-value">' +
        (m.is_cost_loaded === false ? 'N/A*' : esc(fmtIndex(m.cpi))) +
        '</div>' +
        '<div class="metric-subtitle status-' + cSt + '">' + esc(cText) + '</div></div>' +

        '<div class="metric-card neutral">' +
        '<div class="metric-label">% Complete</div>' +
        '<div class="metric-value">' + esc(fmtPct(m.pct_complete)) + '</div>' +
        '<div class="metric-subtitle">Earned / BAC</div></div>' +

        '<div class="metric-card neutral">' +
        '<div class="metric-label">% Spent</div>' +
        '<div class="metric-value">' + esc(fmtPct(m.pct_spent)) + '</div>' +
        '<div class="metric-subtitle">AC / BAC</div></div>';
}

function renderFinancialMetrics(m) {
    const container = document.getElementById('financialMetrics');
    if (!container) return;
    const svClass = Number(m.sv) >= 0 ? 'good' : 'bad';
    const cvClass = Number(m.cv) >= 0 ? 'good' : 'bad';
    const vacClass = Number(m.vac) >= 0 ? 'good' : 'bad';
    const cvCardClass = m.is_cost_loaded === false ? 'neutral' : cvClass;
    const vacCardClass = m.is_cost_loaded === false ? 'neutral' : vacClass;

    container.innerHTML =
        '<div class="metric-card neutral"><div class="metric-label">BAC</div><div class="metric-value">' + esc(fmtMoney(m.bac)) + '</div><div class="metric-subtitle">Budget at Completion</div></div>' +
        '<div class="metric-card neutral"><div class="metric-label">PV</div><div class="metric-value">' + esc(fmtMoney(m.pv)) + '</div><div class="metric-subtitle">Planned Value</div></div>' +
        '<div class="metric-card neutral"><div class="metric-label">EV</div><div class="metric-value">' + esc(fmtMoney(m.ev)) + '</div><div class="metric-subtitle">Earned Value</div></div>' +
        '<div class="metric-card neutral"><div class="metric-label">AC</div><div class="metric-value">' + esc(fmtMoney(m.ac)) + '</div><div class="metric-subtitle">Actual Cost</div></div>' +
        '<div class="metric-card ' + svClass + '"><div class="metric-label">SV</div><div class="metric-value">' + esc(fmtMoney(m.sv)) + '</div><div class="metric-subtitle">EV − PV</div></div>' +
        '<div class="metric-card ' + cvCardClass + '"><div class="metric-label">CV</div><div class="metric-value">' + esc(fmtMoney(m.cv)) + '</div><div class="metric-subtitle">EV − AC</div></div>' +
        '<div class="metric-card neutral"><div class="metric-label">EAC</div><div class="metric-value">' + esc(fmtMoney(m.eac)) + '</div><div class="metric-subtitle">Est. at Completion</div></div>' +
        '<div class="metric-card neutral"><div class="metric-label">ETC</div><div class="metric-value">' + esc(fmtMoney(m.etc)) + '</div><div class="metric-subtitle">Est. to Complete</div></div>' +
        '<div class="metric-card ' + vacCardClass + '"><div class="metric-label">VAC</div><div class="metric-value">' + esc(fmtMoney(m.vac)) + '</div><div class="metric-subtitle">BAC − EAC</div></div>';
}

function renderScurve(scurveData, metrics) {
    const canvas = document.getElementById('scurveChart');
    const errNode = document.getElementById('scurveError');

    if (errNode) { errNode.style.display = 'none'; errNode.textContent = ''; }

    if (!scurveData || scurveData.error) {
        if (errNode) {
            errNode.style.display = 'block';
            errNode.textContent = (scurveData && scurveData.error) ? String(scurveData.error) : 'No S-curve data available';
        }
        if (scurveChart) { scurveChart.destroy(); scurveChart = null; }
        return;
    }
    if (!canvas) return;
    const ctx = canvas.getContext('2d');
    if (scurveChart) { scurveChart.destroy(); scurveChart = null; }

    const labels = scurveData.labels || [];
    const pv = normalizeSeries(scurveData.planned_value);
    const ev = normalizeSeries(scurveData.earned_value);
    const ac = normalizeSeries(scurveData.actual_cost);

    scurveChart = new Chart(ctx, {
        type: 'line',
        data: {
            labels: labels,
            datasets: [
                { label: 'Planned Value (PV)', data: pv, borderColor: '#3b82f6', borderWidth: 2, fill: false, tension: 0.1, spanGaps: false, pointRadius: 0, pointHoverRadius: 4 },
                { label: 'Earned Value (EV)', data: ev, borderColor: '#10b981', borderWidth: 3, fill: false, tension: 0.1, spanGaps: false, pointRadius: 0, pointHoverRadius: 4 },
                { label: 'Actual Cost (AC)', data: ac, borderColor: '#dc2626', borderWidth: 2, borderDash: [5, 5], fill: false, tension: 0.1, spanGaps: false, pointRadius: 0, pointHoverRadius: 4 }
            ]
        },
        options: {
            responsive: true, maintainAspectRatio: false,
            interaction: { mode: 'index', intersect: false },
            plugins: {
                legend: { position: 'bottom' },
                tooltip: {
                    callbacks: {
                        label: function (ctx) {
                            const val = ctx.parsed && ctx.parsed.y;
                            if (val === null || val === undefined || !Number.isFinite(val)) return null;
                            return ctx.dataset.label + ': ' + fmtMoney(val);
                        }
                    }
                }
            },
            scales: {
                y: {
                    beginAtZero: true,
                    ticks: { callback: function (v) { return '$' + Number(v).toLocaleString('en-US'); } }
                },
                x: { ticks: { maxTicksLimit: 20, maxRotation: 45 } }
            }
        }
    });
}

// ═══════════════════════════════════════════
// 👷 RESOURCE ANALYTICS RENDERER
// ═══════════════════════════════════════════

async function loadResourceData(forceRefresh) {
    const section = document.getElementById('resourceSection');
    const errorBox = document.getElementById('resourceErrorBox');
    if (!section) return;

    section.style.display = 'block';

    try {
        const res = await fetch('/api/resource-data');
        const data = await res.json().catch(function () { return {}; });

        if (!res.ok || data.error) {
            throw new Error(data.error || 'Failed to load resource data');
        }

        const rData = data.data || {};

        if (rData.error) {
            if (errorBox) {
                errorBox.style.display = 'block';
                errorBox.textContent = '⚠️ ' + rData.error;
            }
            return;
        }

        if (errorBox) errorBox.style.display = 'none';

        renderResourceStats(rData);
        renderMonthlyUnitsChart(rData.monthly_units || {});
        renderMonthlyCostChart(rData.monthly_cost || {});
        renderCumulativeUnitsChart(rData.monthly_units || {});
        renderTopResources(rData.top_resources || {});
    } catch (err) {
        console.error('Resource load error:', err);
        if (errorBox) {
            errorBox.style.display = 'block';
            errorBox.textContent = '❌ ' + (err.message || 'Failed to load resource data');
        }
    }
}

function renderResourceStats(rData) {
    const container = document.getElementById('resourceStats');
    if (!container) return;

    const units = rData.monthly_units || {};
    const cost = rData.monthly_cost || {};
    const top = rData.top_resources || {};

    const stats = [
        { label: 'Total Man-Hours', value: fmtNum(units.total_units || 0) },
        { label: 'Peak Month (Hrs)', value: (units.peak_month || '—') + ' (' + fmtNum(units.peak_units || 0) + ')' },
        { label: 'Total Planned Cost', value: fmtMoney(cost.total_planned || 0) },
        { label: 'Total Actual Cost', value: fmtMoney(cost.total_actual || 0) },
        { label: 'Unique Resources', value: fmtNum(top.total_count || 0) },
    ];

    container.innerHTML = stats.map(function (s) {
        return '<div class="resource-stat"><div class="val">' + esc(s.value) + '</div><div class="lbl">' + esc(s.label) + '</div></div>';
    }).join('');
}

function renderMonthlyUnitsChart(unitsData) {
    const canvas = document.getElementById('monthlyUnitsChart');
    if (!canvas) return;
    const ctx = canvas.getContext('2d');
    if (monthlyUnitsChart) { monthlyUnitsChart.destroy(); monthlyUnitsChart = null; }

    if (!unitsData.labels || !unitsData.labels.length) {
        ctx.fillStyle = '#64748b';
        ctx.font = '14px sans-serif';
        ctx.textAlign = 'center';
        ctx.fillText('No resource unit data available', canvas.width / 2, canvas.height / 2);
        return;
    }

    monthlyUnitsChart = new Chart(ctx, {
        type: 'bar',
        data: {
            labels: unitsData.labels,
            datasets: unitsData.datasets || []
        },
        options: {
            responsive: true, maintainAspectRatio: false,
            plugins: {
                legend: { position: 'bottom' },
                tooltip: {
                    callbacks: {
                        label: function (ctx) {
                            return ctx.dataset.label + ': ' + fmtNum(ctx.parsed.y) + ' hrs';
                        }
                    }
                }
            },
            scales: {
                x: { stacked: true, ticks: { maxTicksLimit: 24, maxRotation: 45 } },
                y: { 
                    stacked: true, 
                    beginAtZero: true,
                    ticks: { callback: function (v) { return fmtNum(v) + ' hrs'; } }
                }
            }
        }
    });
}

function renderMonthlyCostChart(costData) {
    const canvas = document.getElementById('monthlyCostChart');
    if (!canvas) return;
    const ctx = canvas.getContext('2d');
    if (monthlyCostChart) { monthlyCostChart.destroy(); monthlyCostChart = null; }

    if (!costData.labels || !costData.labels.length) {
        ctx.fillStyle = '#64748b';
        ctx.font = '14px sans-serif';
        ctx.textAlign = 'center';
        ctx.fillText('No cost data available', canvas.width / 2, canvas.height / 2);
        return;
    }

    const datasets = [
        {
            label: 'Planned Cost (Monthly)',
            data: costData.planned || [],
            backgroundColor: 'rgba(59, 130, 246, 0.6)',
            borderColor: '#3b82f6',
            type: 'bar',
            order: 2
        }
    ];

    if (costData.actual && costData.actual.length) {
        datasets.push({
            label: 'Actual Cost (Monthly)',
            data: costData.actual,
            backgroundColor: 'rgba(220, 38, 38, 0.6)',
            borderColor: '#dc2626',
            type: 'bar',
            order: 2
        });
    }

    // Cumulative overlay lines
    if (costData.cumulative_planned && costData.cumulative_planned.length) {
        datasets.push({
            label: 'Cumulative Planned',
            data: costData.cumulative_planned,
            borderColor: '#1e40af',
            backgroundColor: 'transparent',
            type: 'line',
            yAxisID: 'y1',
            borderWidth: 2,
            fill: false,
            tension: 0.1,
            pointRadius: 0,
            order: 1
        });
    }
    if (costData.cumulative_actual && costData.cumulative_actual.length) {
        datasets.push({
            label: 'Cumulative Actual',
            data: costData.cumulative_actual,
            borderColor: '#7f1d1d',
            backgroundColor: 'transparent',
            type: 'line',
            yAxisID: 'y1',
            borderWidth: 2,
            borderDash: [3, 3],
            fill: false,
            tension: 0.1,
            pointRadius: 0,
            order: 1
        });
    }

    monthlyCostChart = new Chart(ctx, {
        data: {
            labels: costData.labels,
            datasets: datasets
        },
        options: {
            responsive: true, maintainAspectRatio: false,
            interaction: { mode: 'index', intersect: false },
            plugins: {
                legend: { position: 'bottom' },
                tooltip: {
                    callbacks: {
                        label: function (ctx) {
                            return ctx.dataset.label + ': ' + fmtMoney(ctx.parsed.y);
                        }
                    }
                }
            },
            scales: {
                x: { ticks: { maxTicksLimit: 24, maxRotation: 45 } },
                y: {
                    type: 'linear', position: 'left', beginAtZero: true,
                    title: { display: true, text: 'Monthly Cost' },
                    ticks: { callback: function (v) { return '$' + Number(v).toLocaleString('en-US'); } }
                },
                y1: {
                    type: 'linear', position: 'right', beginAtZero: true,
                    title: { display: true, text: 'Cumulative Cost' },
                    grid: { drawOnChartArea: false },
                    ticks: { callback: function (v) { return '$' + Number(v).toLocaleString('en-US'); } }
                }
            }
        }
    });
}

function renderCumulativeUnitsChart(unitsData) {
    const canvas = document.getElementById('cumulativeUnitsChart');
    if (!canvas) return;
    const ctx = canvas.getContext('2d');
    if (cumulativeUnitsChart) { cumulativeUnitsChart.destroy(); cumulativeUnitsChart = null; }

    if (!unitsData.labels || !unitsData.labels.length) {
        ctx.fillStyle = '#64748b';
        ctx.font = '14px sans-serif';
        ctx.textAlign = 'center';
        ctx.fillText('No cumulative data available', canvas.width / 2, canvas.height / 2);
        return;
    }

    cumulativeUnitsChart = new Chart(ctx, {
        type: 'line',
        data: {
            labels: unitsData.labels,
            datasets: [{
                label: 'Cumulative Man-Hours',
                data: unitsData.cumulative_curve || [],
                borderColor: '#10b981',
                backgroundColor: 'rgba(16, 185, 129, 0.15)',
                borderWidth: 3,
                fill: true,
                tension: 0.2,
                pointRadius: 0,
                pointHoverRadius: 4
            }]
        },
        options: {
            responsive: true, maintainAspectRatio: false,
            plugins: {
                legend: { position: 'bottom' },
                tooltip: {
                    callbacks: {
                        label: function (ctx) {
                            return 'Cumulative: ' + fmtNum(ctx.parsed.y) + ' hrs';
                        }
                    }
                }
            },
            scales: {
                x: { ticks: { maxTicksLimit: 20, maxRotation: 45 } },
                y: {
                    beginAtZero: true,
                    ticks: { callback: function (v) { return fmtNum(v) + ' hrs'; } }
                }
            }
        }
    });
}

function renderTopResources(topData) {
    const container = document.getElementById('topResourcesUnits');
    if (!container) return;

    const list = topData.by_units || [];
    if (!list.length) {
        container.innerHTML = '<p style="text-align:center;padding:1rem;color:#64748b;">No resources with man-hours found.</p>';
        return;
    }

    container.innerHTML = list.map(function (r, idx) {
        return '<div class="top-resource-item">' +
            '<div>' +
            '<span style="color:#94a3b8;font-weight:600;">' + (idx + 1) + '.</span> ' +
            '<span class="name">' + esc(r.name) + '</span>' +
            '<span class="type">' + esc(r.type) + '</span>' +
            '</div>' +
            '<div class="val">' + fmtNum(r.units) + ' hrs</div>' +
            '</div>';
    }).join('');
}
'''

os.makedirs("static", exist_ok=True)
with open("static/evm.js", "w", encoding="utf-8") as f:
    f.write(EVM_JS_CODE)
print("  ✅ Updated static/evm.js")

print("\n🎉 Phase 1 - Step 3 (Resource Analytics) Applied Successfully!")
print("✨ Restart Flask (python app.py), navigate to /evm, and see the new Resource Analytics section!")