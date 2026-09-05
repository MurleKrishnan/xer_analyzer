import os
import shutil
from datetime import datetime

print("🚀 Starting Patch Application via Python (Part 3/3)...")

# 1. Create Backup Folder
timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
backup_dir = f"_backup_part3_{timestamp}"
os.makedirs(backup_dir, exist_ok=True)
print(f"📦 Created backup folder: {backup_dir}")

files_to_backup = [
    "health_standards/nasa_checks.py",
    "health_standards/aace_checks.py",
    "health_standards/industry_checks.py",
    "evm_engine.py",
    "comparison_engine.py",
    "reports.py",
    "pdf_report_generator.py",
    "app.py",
]

for file_path in files_to_backup:
    if os.path.exists(file_path):
        dest = os.path.join(
            backup_dir, os.path.basename(file_path.replace("/", os.sep))
        )
        shutil.copy2(file_path, dest)
        print(f"   Backed up {file_path}")

os.makedirs("health_standards", exist_ok=True)


# ------------------------------------------------------------------------------
# 1. health_standards/nasa_checks.py
# ------------------------------------------------------------------------------
NASA_CHECKS_CODE = '''"""
NASA NPR 7120.5 SCHEDULE MANAGEMENT
====================================
"""

from health_standards.base_checker import BaseChecker
from collections import Counter, defaultdict
from datetime import datetime
import statistics


class NASAChecks(BaseChecker):
    """NASA schedule management check suite."""

    def run_checks(self):
        return {
            'name': 'NASA NPR 7120.5 & Handbook',
            'description': 'NASA program and project management schedule requirements',
            'categories': [
                self._schedule_structure(),
                self._logic_integrity(),
                self._duration_analysis(),
                self._milestone_management(),
                self._critical_path_analysis(),
                self._risk_maturity(),
            ]
        }

    def _wbs_maps(self):
        by_id = {str(w.get('wbs_id', '')): w for w in self.wbs_nodes if w.get('wbs_id')}

        def parent_of(w):
            for k in ('parent_wbs_id', 'parent_id', 'parent_wbs', 'wbs_parent_id'):
                v = w.get(k, '')
                if v not in (None, '', '0', 0):
                    return str(v)
            return ''

        children = defaultdict(list)
        for wid, w in by_id.items():
            p = parent_of(w)
            if p:
                children[p].append(wid)

        acts_by_wbs = defaultdict(list)
        for a in self.activities:
            wid = str(a.get('wbs_id', '') or '')
            if wid:
                acts_by_wbs[wid].append(a)

        def depth(wid, seen=None):
            seen = seen or set()
            if not wid or wid in seen:
                return 0
            seen.add(wid)
            w = by_id.get(wid)
            if not w:
                return 0
            p = parent_of(w)
            return 1 + (depth(p, seen) if p in by_id else 0)

        leaf_ids = [wid for wid in by_id if not children.get(wid)]
        leaves_with_acts = [wid for wid in leaf_ids if acts_by_wbs.get(wid)]
        max_depth = max((depth(wid) for wid in by_id), default=0)

        return {
            'by_id': by_id,
            'children': children,
            'acts_by_wbs': acts_by_wbs,
            'leaf_ids': leaf_ids,
            'leaves_with_acts': leaves_with_acts,
            'max_depth': max_depth,
        }

    def _schedule_structure(self):
        checks = []
        total = len(self.activities) or 1
        wbs = self._wbs_maps()

        checks.append(self.make_metric(
            'NASA-101', 'IMS Activity Count',
            f'{len(self.activities)} activities',
            len(self.activities), 'NASA', 'Structure',
            threshold_min=100, severity='low', info_only=True,
            recommendation='IMS should contain sufficient detail for management.'
        ))

        leaf_total = len(wbs['leaf_ids']) or 1
        leaf_covered = len(wbs['leaves_with_acts'])
        coverage = leaf_covered / leaf_total * 100
        checks.append(self.make_metric(
            'NASA-102', 'Leaf WBS Coverage',
            f'{coverage:.1f}% of leaf WBS nodes have activities ({leaf_covered}/{leaf_total})',
            coverage, 'NASA', 'Structure',
            threshold_min=80, severity='medium',
            recommendation='Ensure leaf WBS elements map to planned work (parent headers may be empty).'
        ))

        no_name = [a for a in self.activities if not str(a.get('task_name', '')).strip()]
        checks.append(self.make_check(
            'NASA-103', 'Missing Activity Names',
            'All activities must have descriptive names',
            len(no_name), total, 0, 'NASA', 'critical', 'Structure',
            'Add descriptive names to all activities.',
            no_name
        ))

        short_names = [
            a for a in self.activities
            if 0 < len(str(a.get('task_name', '')).strip()) < 10
        ]
        checks.append(self.make_check(
            'NASA-104', 'Very Short Activity Names (<10 chars)',
            'NASA recommends descriptive verb-based names',
            len(short_names), total, 3, 'NASA', 'medium', 'Structure',
            'Use verb-based descriptive names (e.g., "Design Panel", "Test Interface").',
            short_names
        ))

        name_counts = Counter(
            str(a.get('task_name', '')).strip()
            for a in self.activities if str(a.get('task_name', '')).strip()
        )
        dup_names = {n for n, c in name_counts.items() if c > 1}
        dup_acts = [a for a in self.activities if str(a.get('task_name', '')).strip() in dup_names]
        checks.append(self.make_check(
            'NASA-105', 'Duplicate Activity Names',
            'Activity names should be unique for clarity',
            len(dup_acts), total, 5, 'NASA', 'medium', 'Structure',
            'Add distinguishing context to duplicate names.',
            dup_acts
        ))

        id_lengths = [len(a.get('task_code', '')) for a in self.activities if a.get('task_code')]
        unique_lengths = len(set(id_lengths)) if id_lengths else 0
        checks.append(self.make_metric(
            'NASA-106', 'Activity ID Format Consistency',
            f'{unique_lengths} different ID lengths',
            unique_lengths, 'NASA', 'Structure',
            threshold_max=3, severity='low',
            recommendation='Use consistent ID formatting across all activities.'
        ))

        checks.append(self.make_metric(
            'NASA-107', 'Maximum WBS Depth',
            f'{wbs["max_depth"]} levels (parent hierarchy)',
            wbs['max_depth'], 'NASA', 'Structure',
            threshold_min=2, threshold_max=8, severity='low', info_only=True
        ))

        return {'name': 'Schedule Structure', 'checks': checks}

    def _logic_integrity(self):
        checks = []
        total = len(self.incomplete) or 1
        active_rels = self.active_relationships()
        rel_total = len(active_rels) or 1
        dens_denom = len(self.real_activities) or 1

        density = len(self.relationships) / dens_denom
        checks.append(self.make_metric(
            'NASA-201', 'Logic Density',
            f'{density:.2f} relationships per real activity',
            density, 'NASA', 'Logic',
            threshold_min=1.5, threshold_max=3.5, severity='medium',
            recommendation='NASA guideline: roughly 1.5-3.5 relationships per activity.'
        ))

        open_start = self.open_start_activities()
        open_end = self.open_end_activities()

        checks.append(self.make_check(
            'NASA-202', 'Open Start Activities',
            'Incomplete non-milestone activities without predecessors',
            len(open_start), total, 1, 'NASA', 'high', 'Logic',
            'Add predecessors to eliminate dangling activities.',
            open_start
        ))

        checks.append(self.make_check(
            'NASA-203', 'Open End Activities',
            'Incomplete non-milestone activities without successors',
            len(open_end), total, 1, 'NASA', 'high', 'Logic',
            'Add successors to eliminate dangling activities.',
            open_end
        ))

        fs_rels = [r for r in active_rels if r.get('pred_type') == 'PR_FS']
        fs_pct = len(fs_rels) / rel_total * 100
        checks.append(self.make_metric(
            'NASA-204', 'Finish-to-Start Percentage',
            f'{fs_pct:.1f}% are FS (active logic)',
            fs_pct, 'NASA', 'Logic',
            threshold_min=85, severity='medium',
            recommendation='NASA prefers 85%+ FS relationships.'
        ))

        cross_wbs = 0
        for r in active_rels:
            pred = self.engine.activity_by_id.get(str(r.get('pred_task_id', '')), {})
            succ = self.engine.activity_by_id.get(str(r.get('task_id', '')), {})
            if pred.get('wbs_id') and succ.get('wbs_id') and pred.get('wbs_id') != succ.get('wbs_id'):
                cross_wbs += 1
        cross_pct = cross_wbs / rel_total * 100
        checks.append(self.make_metric(
            'NASA-205', 'Cross-WBS Relationships',
            f'{cross_pct:.1f}% of active ties cross WBS boundaries',
            cross_pct, 'NASA', 'Logic',
            severity='low', info_only=True,
            recommendation='Cross-WBS logic is normal but should be reviewed for interfaces.'
        ))

        checks.append(self.make_metric(
            'NASA-206', 'Total Relationships',
            f'{len(self.relationships)} logic ties',
            len(self.relationships), 'NASA', 'Logic',
            severity='low', info_only=True
        ))

        too_many_preds = [
            a for a in self.incomplete
            if len(self.engine.predecessors.get(str(a.get('task_id', '')), [])) > 10
        ]
        checks.append(self.make_check(
            'NASA-207', 'Activities with >10 Predecessors',
            'Complex fan-in may need decomposition',
            len(too_many_preds), total, 3, 'NASA', 'low', 'Logic',
            'Consider summary milestones to reduce complexity.',
            too_many_preds
        ))

        too_many_succs = [
            a for a in self.incomplete
            if len(self.engine.successors.get(str(a.get('task_id', '')), [])) > 10
        ]
        checks.append(self.make_check(
            'NASA-208', 'Activities with >10 Successors',
            'High fan-out = risk concentration',
            len(too_many_succs), total, 3, 'NASA', 'low', 'Logic',
            'Review activities driving many successors.',
            too_many_succs
        ))

        fs_lag = self.fs_with_lag()
        checks.append(self.make_check(
            'NASA-FS-LAG', 'FS + Lag Relationships',
            'NASA guidance discourages FS with lag',
            len(fs_lag), rel_total, 3, 'NASA', 'medium', 'Logic',
            'Replace lag with a schedule activity for transparency.',
            fs_lag
        ))

        leads = [r for r in active_rels if r.get('lag_days', 0) < 0]
        checks.append(self.make_check(
            'NASA-209', 'Negative Lags (Leads)',
            'Leads distort network logic',
            len(leads), rel_total, 0, 'NASA', 'high', 'Logic',
            'Remove negative lags.',
            leads
        ))

        return {'name': 'Logic & Network Integrity', 'checks': checks}

    def _duration_analysis(self):
        checks = []
        total = len(self.incomplete) or 1

        durations = [
            a.get('original_duration_days', 0) for a in self.real_activities
            if a.get('original_duration_days', 0) > 0
        ]

        if durations:
            mean_dur = statistics.mean(durations)
            median_dur = statistics.median(durations)
            checks.append(self.make_metric(
                'NASA-301', 'Average Duration',
                f'{mean_dur:.1f} days',
                mean_dur, 'NASA', 'Duration',
                threshold_min=1, threshold_max=30, severity='low', info_only=True,
                recommendation='NASA typical: 5-20 days average.'
            ))
            checks.append(self.make_metric(
                'NASA-302', 'Median Duration',
                f'{median_dur:.1f} days',
                median_dur, 'NASA', 'Duration',
                threshold_min=1, threshold_max=25, severity='low', info_only=True
            ))
            if len(durations) > 1:
                checks.append(self.make_metric(
                    'NASA-303', 'Duration Std Deviation',
                    f'{statistics.stdev(durations):.1f} days',
                    statistics.stdev(durations), 'NASA', 'Duration',
                    severity='low', info_only=True
                ))

        very_short = [
            a for a in self.incomplete
            if 0 < a.get('original_duration_days', 0) < 2 and not self.is_milestone(a)
        ]
        checks.append(self.make_check(
            'NASA-304', 'Very Short Activities (<2 days)',
            'May need consolidation',
            len(very_short), total, 5, 'NASA', 'low', 'Duration',
            'Consider consolidating micro-activities.',
            very_short
        ))

        excessive = [
            a for a in self.incomplete
            if a.get('original_duration_days', 0) > 88
            and not self.is_milestone(a)
            and a.get('task_type') != 'TT_LOE'
        ]
        checks.append(self.make_check(
            'NASA-305', 'Excessive Duration (>88 days)',
            'NASA handbook-aligned max ~88-day detailed activities',
            len(excessive), total, 2, 'NASA', 'medium', 'Duration',
            'Break down activities over 88 days.',
            excessive
        ))

        return {'name': 'Duration Analysis', 'checks': checks}

    def _milestone_management(self):
        checks = []
        total = len(self.activities) or 1
        mile_total = max(len(self.milestones), 1)

        mile_pct = len(self.milestones) / total * 100 if total else 0
        checks.append(self.make_metric(
            'NASA-401', 'Milestone Percentage',
            f'{mile_pct:.1f}% are milestones',
            mile_pct, 'NASA', 'Milestones',
            threshold_min=3, threshold_max=15, severity='medium',
            recommendation='NASA guideline: roughly 3-15% milestones.'
        ))

        mile_with_dur = [
            a for a in self.milestones
            if a.get('original_duration_days', 0) > 0
        ]
        checks.append(self.make_check(
            'NASA-402', 'Milestones with Duration',
            'Milestones must have zero duration',
            len(mile_with_dur), mile_total, 0, 'NASA', 'critical', 'Milestones',
            'Set milestone durations to zero.',
            mile_with_dur
        ))

        finish_miles = [
            a for a in self.milestones
            if a.get('task_type') == 'TT_FinMile'
            or (
                a.get('task_type') == 'TT_Mile'
                and str(a.get('task_id', '')) not in self.engine.successors
                and a.get('status_code') != 'TK_Complete'
            )
        ]
        mile_no_pred = [
            a for a in finish_miles
            if str(a.get('task_id', '')) not in self.engine.predecessors
        ]
        checks.append(self.make_check(
            'NASA-403', 'Finish Milestones Without Predecessors',
            'Finish milestones must have predecessors',
            len(mile_no_pred), max(len(finish_miles), 1), 0, 'NASA', 'high', 'Milestones',
            'Add predecessors to finish milestones.',
            mile_no_pred
        ))

        checks.append(self.make_metric(
            'NASA-404', 'Total Milestones',
            f'{len(self.milestones)} milestones',
            len(self.milestones), 'NASA', 'Milestones',
            threshold_min=5, severity='medium',
            recommendation='NASA projects typically need multiple key milestones.'
        ))

        return {'name': 'Milestone Management', 'checks': checks}

    def _critical_path_analysis(self):
        checks = []
        total_inc = len(self.incomplete) or 1
        crit_inc = [a for a in self.incomplete if a.get('is_critical')]
        cp_count = len(crit_inc)
        cp_pct = cp_count / total_inc * 100

        checks.append(self.make_metric(
            'NASA-501', 'Critical Path % (Incomplete)',
            f'{cp_count} activities ({cp_pct:.1f}% of incomplete)',
            cp_pct, 'NASA', 'Critical Path',
            threshold_min=5, threshold_max=25, severity='medium',
            recommendation='NASA guideline: CP roughly 5-25% of remaining activities.'
        ))

        checks.append(self.make_boolean(
            'NASA-502', 'Critical Path Exists',
            'Schedule must have a critical path on remaining work',
            cp_count > 0, 'NASA', 'Critical Path',
            severity='critical',
            recommendation='Must have valid CPM critical path.'
        ))

        continuity = self._cp_continuity()
        checks.append(self.make_metric(
            'NASA-502b', 'Critical Path Continuity',
            'Fraction of critical acts linked to other critical acts',
            continuity, 'NASA', 'Critical Path',
            threshold_min=0.85, severity='high',
            recommendation='Critical path should form a continuous chain to project finish.'
        ))

        near = [a for a in self.incomplete if 0 < a.get('total_float_days', 0) <= 5]
        near_pct = len(near) / total_inc * 100
        checks.append(self.make_metric(
            'NASA-503', 'Near-Critical (<5 days float)',
            f'{near_pct:.1f}% near-critical',
            near_pct, 'NASA', 'Critical Path',
            threshold_max=15, severity='medium',
            recommendation='High near-critical percentage = high schedule risk.'
        ))

        floats = [a.get('total_float_days', 0) for a in self.incomplete]
        if floats:
            checks.append(self.make_metric(
                'NASA-504', 'Float Range (max)',
                f'{min(floats):.0f} to {max(floats):.0f} days',
                max(floats), 'NASA', 'Critical Path',
                threshold_max=200, severity='low', info_only=True
            ))

        return {'name': 'Critical Path Analysis', 'checks': checks}

    def _cp_continuity(self):
        try:
            crit_ids = {str(a.get('task_id', '')) for a in self.incomplete if a.get('is_critical')}
            if not crit_ids:
                return None
            connected = 0
            for cid in crit_ids:
                succs = self.engine.successors.get(cid, [])
                preds = self.engine.predecessors.get(cid, [])
                if any(str(s.get('task_id')) in crit_ids for s in succs):
                    connected += 1
                elif any(str(p.get('task_id')) in crit_ids for p in preds):
                    connected += 1
            return round(connected / len(crit_ids), 3)
        except Exception:
            return None

    def _risk_maturity(self):
        checks = []
        total = len(self.activities) or 1

        if self.data_date:
            age = (datetime.now() - self.data_date).days
            checks.append(self.make_metric(
                'NASA-601', 'Data Date Age',
                f'{age} days old',
                age, 'NASA', 'Maturity',
                threshold_max=30, severity='medium',
                info_only=(age > 365),
                recommendation='Data date should be updated regularly (≤30 days) for live projects.'
            ))

        complete_pct = len(self.completed) / total * 100 if total else 0
        checks.append(self.make_metric(
            'NASA-602', 'Overall Completion (activity count)',
            f'{complete_pct:.1f}% complete',
            complete_pct, 'NASA', 'Maturity',
            severity='low', info_only=True
        ))

        checks.append(self.make_metric(
            'NASA-603', 'Active Activities',
            f'{len(self.in_progress)} in progress',
            len(self.in_progress), 'NASA', 'Maturity',
            severity='low', info_only=True
        ))

        ns_pct = len(self.not_started) / total * 100 if total else 0
        checks.append(self.make_metric(
            'NASA-604', 'Not Started Percentage',
            f'{ns_pct:.1f}%',
            ns_pct, 'NASA', 'Maturity',
            severity='low', info_only=True
        ))

        pool = self.real_activities
        pool_n = len(pool) or 1
        with_bl = sum(1 for a in pool if a.get('target_start_date', ''))
        bl_pct = with_bl / pool_n * 100
        checks.append(self.make_metric(
            'NASA-605', 'Target/Planned Date Coverage',
            f'{bl_pct:.1f}% have target start (baseline proxy)',
            bl_pct, 'NASA', 'Maturity',
            threshold_min=100, severity='high',
            recommendation='All discrete activities should have planned/target dates; true PMB may be a separate baseline.'
        ))

        hard = [a for a in self.incomplete if self.has_hard_constraint(a)]
        checks.append(self.make_check(
            'NASA-606', 'Hard Constraints',
            'Hard constraints override CPM',
            len(hard), len(self.incomplete) or 1, 5, 'NASA', 'high', 'Maturity',
            'Minimize MSO/MEO/Mandatory constraints.',
            hard
        ))

        return {'name': 'Schedule Maturity & Risk', 'checks': checks}
'''

nasa_path = os.path.join("health_standards", "nasa_checks.py")
with open(nasa_path, "w", encoding="utf-8") as f:
    f.write(NASA_CHECKS_CODE)
print("  ✅ Updated health_standards/nasa_checks.py")


# ------------------------------------------------------------------------------
# 2. health_standards/aace_checks.py
# ------------------------------------------------------------------------------
AACE_CHECKS_CODE = '''"""
AACE INTERNATIONAL SCHEDULE PRACTICES
======================================
"""

from health_standards.base_checker import BaseChecker
from collections import Counter, defaultdict
import statistics


class AACEChecks(BaseChecker):
    """AACE International check suite."""

    def run_checks(self):
        return {
            'name': 'AACE International RPs',
            'description': 'AACE International Recommended Practices for scheduling',
            'categories': [
                self._activity_definition(),
                self._duration_estimation(),
                self._logic_relationships(),
                self._resource_loading(),
                self._forensic_readiness(),
                self._schedule_level_detail(),
            ]
        }

    def _wbs_maps(self):
        by_id = {str(w.get('wbs_id', '')): w for w in self.wbs_nodes if w.get('wbs_id')}

        def parent_of(w):
            for k in ('parent_wbs_id', 'parent_id', 'parent_wbs', 'wbs_parent_id'):
                v = w.get(k, '')
                if v not in (None, '', '0', 0):
                    return str(v)
            return ''

        def depth(wid, seen=None):
            seen = seen or set()
            if not wid or wid in seen:
                return 0
            seen.add(wid)
            w = by_id.get(wid)
            if not w:
                return 0
            p = parent_of(w)
            return 1 + (depth(p, seen) if p in by_id else 0)

        max_depth = max((depth(wid) for wid in by_id), default=0)
        return max_depth

    def _activity_definition(self):
        checks = []
        total = len(self.activities) or 1
        
        checks.append(self.make_metric(
            'AACE-101', 'Discrete Activity Count',
            f'{len(self.real_activities)} discrete activities',
            len(self.real_activities), 'AACE', 'Activity Definition',
            threshold_min=10, severity='low', info_only=True
        ))
        
        loe = sum(1 for a in self.activities if a.get('task_type') == 'TT_LOE')
        loe_pct = loe / total * 100
        checks.append(self.make_metric(
            'AACE-102', 'LOE Percentage',
            f'{loe_pct:.1f}% LOE',
            loe_pct, 'AACE', 'Activity Definition',
            threshold_max=10, severity='medium',
            recommendation='AACE recommends LOE < 10% of activities.'
        ))
        
        generic_names = {'TASK', 'WORK', 'ACTIVITY', 'ITEM', 'TEST', 'DO', 'COMPLETE', 'NEW TASK'}
        generic = [
            a for a in self.activities 
            if str(a.get('task_name', '')).strip().upper() in generic_names
        ]
        checks.append(self.make_check(
            'AACE-103', 'Generic Activity Names',
            'Names should be specific and meaningful',
            len(generic), total, 1, 'AACE', 'low', 'Activity Definition',
            'Use descriptive, action-oriented names.',
            generic
        ))
        
        start_mile = sum(1 for a in self.activities if a.get('task_type') == 'TT_Mile')
        finish_mile = sum(1 for a in self.activities if a.get('task_type') == 'TT_FinMile')
        checks.append(self.make_metric(
            'AACE-104', 'Start vs Finish Milestones',
            f'{start_mile} start, {finish_mile} finish',
            start_mile + finish_mile, 'AACE', 'Activity Definition',
            severity='low', info_only=True
        ))
        
        return {'name': 'Activity Definition (RP 32R-04)', 'checks': checks}

    def _duration_estimation(self):
        checks = []
        total_inc = len(self.incomplete) or 1
        
        durations = [
            a.get('original_duration_days', 0) for a in self.real_activities 
            if a.get('original_duration_days', 0) > 0
        ]
        
        if durations:
            avg = statistics.mean(durations)
            median = statistics.median(durations)
            
            checks.append(self.make_metric(
                'AACE-201', 'Average Duration',
                f'{avg:.1f} days',
                avg, 'AACE', 'Duration Estimation',
                threshold_min=1, threshold_max=20, severity='low', info_only=True,
                recommendation='AACE typical: 5-15 days average.'
            ))
            
            checks.append(self.make_metric(
                'AACE-202', 'Median Duration',
                f'{median:.1f} days',
                median, 'AACE', 'Duration Estimation',
                threshold_min=1, threshold_max=20, severity='low', info_only=True
            ))
        
        if durations:
            v_short = sum(1 for d in durations if d < 1)
            short = sum(1 for d in durations if 1 <= d < 5)
            med = sum(1 for d in durations if 5 <= d < 20)
            long_d = sum(1 for d in durations if 20 <= d < 44)
            v_long = sum(1 for d in durations if d >= 44)
            
            checks.append(self.make_metric(
                'AACE-203', 'Duration Distribution',
                f'{v_short} <1d | {short} 1-5d | {med} 5-20d | {long_d} 20-44d | {v_long} ≥44d',
                len(durations), 'AACE', 'Duration Estimation',
                severity='low', info_only=True
            ))
        
        long_acts = [
            a for a in self.incomplete 
            if a.get('original_duration_days', 0) > 44
            and not self.is_milestone(a) and a.get('task_type') != 'TT_LOE'
        ]
        checks.append(self.make_check(
            'AACE-204', 'Activities Over AACE Max (44 days)',
            'Break down activities per AACE guidance',
            len(long_acts), total_inc, 5, 'AACE', 'medium', 'Duration Estimation',
            'Decompose activities exceeding 44-day threshold.',
            long_acts
        ))
        
        round_5 = sum(1 for d in durations if d > 0 and d % 5 == 0)
        round_pct = round_5 / len(durations) * 100 if durations else 0
        checks.append(self.make_metric(
            'AACE-205', 'Round Durations (multiples of 5)',
            f'{round_pct:.1f}%',
            round_pct, 'AACE', 'Duration Estimation',
            severity='low', info_only=True,
            recommendation='High round percentage may indicate rough estimation.'
        ))
        
        return {'name': 'Duration Estimation (RP 32R-04)', 'checks': checks}

    def _logic_relationships(self):
        checks = []
        total_inc = len(self.incomplete) or 1
        active_rels = self.active_relationships()
        rel_total = len(active_rels) or 1
        
        type_counts = Counter(r.get('pred_type', '') for r in active_rels)
        fs = type_counts.get('PR_FS', 0)
        fs_pct = fs / rel_total * 100
        
        checks.append(self.make_metric(
            'AACE-301', 'FS Percentage',
            f'{fs_pct:.1f}% (active logic)',
            fs_pct, 'AACE', 'Logic',
            threshold_min=80, severity='medium',
            recommendation='AACE recommends 80%+ FS relationships.'
        ))
        
        ss_pairs = set((str(r.get('pred_task_id')), str(r.get('task_id'))) 
                      for r in active_rels if r.get('pred_type') == 'PR_SS')
        ff_pairs = set((str(r.get('pred_task_id')), str(r.get('task_id'))) 
                      for r in active_rels if r.get('pred_type') == 'PR_FF')
        ladder = ss_pairs & ff_pairs
        
        checks.append(self.make_metric(
            'AACE-302', 'Ladder Logic Pairs (SS+FF)',
            f'{len(ladder)} active pairs',
            len(ladder), 'AACE', 'Logic',
            severity='low', info_only=True,
            recommendation='Ladder logic is acceptable for overlapping activities.'
        ))
        
        lags = [r.get('lag_days', 0) for r in active_rels if r.get('lag_days', 0) != 0]
        if lags:
            avg_lag = statistics.mean([abs(l) for l in lags])
            checks.append(self.make_metric(
                'AACE-303', 'Average Lag Magnitude',
                f'{avg_lag:.1f} days',
                avg_lag, 'AACE', 'Logic',
                threshold_max=10, severity='medium',
                recommendation='High avg lag indicates over-reliance on lags.'
            ))
        
        sf_rels = [r for r in active_rels if r.get('pred_type') == 'PR_SF']
        checks.append(self.make_check(
            'AACE-304', 'SF Relationships',
            'Start-to-Finish rarely correct',
            len(sf_rels), rel_total, 0, 'AACE', 'high', 'Logic',
            'Verify each SF relationship - usually a mistake.',
            sf_rels
        ))
        
        open_start = self.open_start_activities()
        open_end = self.open_end_activities()
        
        checks.append(self.make_check(
            'AACE-OPEN-01', 'Open Start Activities',
            'Incomplete non-milestone activities without predecessors',
            len(open_start), total_inc, 1, 'AACE', 'high', 'Logic',
            'AACE requires proper logic ties. Only start milestones should have no predecessors.',
            open_start
        ))
        
        checks.append(self.make_check(
            'AACE-OPEN-02', 'Open End Activities',
            'Incomplete non-milestone activities without successors',
            len(open_end), total_inc, 1, 'AACE', 'high', 'Logic',
            'AACE requires closed network. Only finish milestones should have no successors.',
            open_end
        ))
        
        fs_lag = self.fs_with_lag()
        checks.append(self.make_check(
            'AACE-FS-LAG', 'FS + Lag Relationships',
            'AACE discourages FS relationships with lag',
            len(fs_lag), rel_total, 3, 'AACE', 'medium', 'Logic',
            'Replace lag with a schedule activity for transparency (per AACE RP 38R-06).',
            fs_lag
        ))
        
        return {'name': 'Logic Relationships', 'checks': checks}

    def _resource_loading(self):
        checks = []
        
        if not self.resources:
            checks.append(self.make_metric(
                'AACE-400', 'Resource Loading Present',
                'No resource assignments in XER',
                None, 'AACE', 'Resources', info_only=True,
                recommendation='Resource loading enables forensic delay and cost analysis.'
            ))
            return {'name': 'Resource Loading', 'checks': checks}
            
        tasks_with_res = set(str(r.get('task_id')) for r in self.resources)
        work_acts = [
            a for a in self.incomplete 
            if not self.is_milestone(a) and a.get('task_type') != 'TT_LOE'
        ]
        
        no_res = [a for a in work_acts if str(a.get('task_id', '')) not in tasks_with_res]
        no_res_pct = len(no_res) / max(len(work_acts), 1) * 100
        
        checks.append(self.make_metric(
            'AACE-401', 'Unresourced Work %',
            f'{no_res_pct:.1f}%',
            no_res_pct, 'AACE', 'Resources',
            threshold_max=10, severity='medium',
            recommendation='AACE recommends <10% unresourced work.'
        ))
        
        with_cost = sum(1 for r in self.resources if self.to_float(r.get('target_cost', '0')) > 0)
        cost_pct = with_cost / max(len(self.resources), 1) * 100
        checks.append(self.make_metric(
            'AACE-402', 'Costed Resources',
            f'{cost_pct:.1f}%',
            cost_pct, 'AACE', 'Resources',
            threshold_min=90, severity='medium',
            recommendation='Cost loading enables forensic analysis.'
        ))
        
        return {'name': 'Resource Loading', 'checks': checks}

    def _forensic_readiness(self):
        checks = []
        total = len(self.activities) or 1
        
        with_bl = sum(1 for a in self.real_activities if a.get('target_start_date', ''))
        bl_pct = with_bl / max(len(self.real_activities), 1) * 100
        checks.append(self.make_metric(
            'AACE-501', 'Target Date Coverage (Baseline Proxy)',
            f'{bl_pct:.1f}% have target start',
            bl_pct, 'AACE', 'Forensic Readiness',
            threshold_min=100, severity='high',
            recommendation='Complete baseline is critical for forensic delay analysis.'
        ))
        
        with_actuals = sum(
            1 for a in self.activities 
            if a.get('act_start_date', '') or a.get('act_end_date', '')
        )
        actual_pct = with_actuals / total * 100
        checks.append(self.make_metric(
            'AACE-502', 'Activities with Actual Dates',
            f'{actual_pct:.1f}%',
            actual_pct, 'AACE', 'Forensic Readiness',
            severity='low', info_only=True
        ))
        
        checks.append(self.make_boolean(
            'AACE-503', 'Data Date Set',
            'Required for time-slice analysis',
            self.data_date is not None, 'AACE', 'Forensic Readiness',
            severity='high',
            recommendation='Data date required for forensic time-slice analysis.'
        ))
        
        checks.append(self.make_metric(
            'AACE-504', 'Calendar Definitions',
            f'{len(self.calendars)} calendars',
            len(self.calendars), 'AACE', 'Forensic Readiness',
            threshold_min=1, severity='medium',
            recommendation='Calendars needed for accurate CPM analysis.'
        ))
        
        return {'name': 'Forensic Readiness (RP 29R-03)', 'checks': checks}

    def _schedule_level_detail(self):
        checks = []
        
        act_count = len(self.real_activities)
        if act_count < 100:
            level = 'Level 1 (Summary)'
        elif act_count < 500:
            level = 'Level 2 (Summary)'
        elif act_count < 2000:
            level = 'Level 3 (Working)'
        elif act_count < 5000:
            level = 'Level 4 (Detail)'
        else:
            level = 'Level 5 (Highly Detailed)'
        
        checks.append(self.make_metric(
            'AACE-601', 'AACE Schedule Level',
            level,
            act_count, 'AACE', 'Level of Detail',
            severity='low', info_only=True
        ))
        
        max_depth = self._wbs_maps()
        checks.append(self.make_metric(
            'AACE-602', 'WBS Depth',
            f'{max_depth} levels',
            max_depth, 'AACE', 'Level of Detail',
            threshold_min=2, threshold_max=8, severity='medium',
            recommendation='AACE recommends 2-8 WBS levels.'
        ))
        
        wbs_counts = Counter(
            str(a.get('wbs_id', '')) for a in self.activities if a.get('wbs_id')
        )
        if wbs_counts:
            values = list(wbs_counts.values())
            if len(values) > 1:
                mean_val = statistics.mean(values)
                if mean_val > 0:
                    cv = statistics.stdev(values) / mean_val
                    checks.append(self.make_metric(
                        'AACE-603', 'WBS Coefficient of Variation',
                        f'{cv:.2f}',
                        cv, 'AACE', 'Level of Detail',
                        threshold_max=2.0, severity='low',
                        recommendation='High variation (>2.0) indicates inconsistent schedule detail across WBS branches.'
                    ))
        
        return {'name': 'Level of Detail (RP 37R-06)', 'checks': checks}
'''

aace_path = os.path.join("health_standards", "aace_checks.py")
with open(aace_path, "w", encoding="utf-8") as f:
    f.write(AACE_CHECKS_CODE)
print("  ✅ Updated health_standards/aace_checks.py")


# ------------------------------------------------------------------------------
# 3. health_standards/industry_checks.py
# ------------------------------------------------------------------------------
INDUSTRY_CHECKS_CODE = '''"""
INDUSTRY BEST PRACTICES
========================
"""

from health_standards.base_checker import BaseChecker
from collections import Counter, defaultdict, deque
from datetime import datetime
import statistics


class IndustryChecks(BaseChecker):
    """Industry best practices check suite."""

    def run_checks(self):
        return {
            'name': 'Industry Best Practices',
            'description': 'Consensus best practices from PMI, CII, ISO, and industry',
            'categories': [
                self._schedule_completeness(),
                self._logic_quality(),
                self._resource_realism(),
                self._progress_transparency(),
                self._schedule_optimization(),
                self._maintainability(),
            ]
        }

    def _wbs_max_depth(self):
        by_id = {str(w.get('wbs_id', '')): w for w in self.wbs_nodes if w.get('wbs_id')}

        def parent_of(w):
            for k in ('parent_wbs_id', 'parent_id', 'parent_wbs', 'wbs_parent_id'):
                v = w.get(k, '')
                if v not in (None, '', '0', 0):
                    return str(v)
            return ''

        def depth(wid, seen=None):
            seen = seen or set()
            if not wid or wid in seen:
                return 0
            seen.add(wid)
            w = by_id.get(wid)
            if not w:
                return 0
            p = parent_of(w)
            return 1 + (depth(p, seen) if p in by_id else 0)

        return max((depth(wid) for wid in by_id), default=0)

    def _has_circular_logic(self):
        graph = defaultdict(list)
        in_degree = defaultdict(int)
        nodes = set()

        for a in self.activities:
            tid = str(a.get('task_id', '') or '')
            if tid:
                nodes.add(tid)
                in_degree.setdefault(tid, 0)

        for r in self.relationships:
            p = str(r.get('pred_task_id', '') or '')
            s = str(r.get('task_id', '') or '')
            if not p or not s:
                continue
            if p not in nodes:
                nodes.add(p)
                in_degree.setdefault(p, 0)
            if s not in nodes:
                nodes.add(s)
                in_degree.setdefault(s, 0)
            graph[p].append(s)
            in_degree[s] += 1

        if not nodes:
            return False

        q = deque([n for n in nodes if in_degree[n] == 0])
        visited = 0
        while q:
            n = q.popleft()
            visited += 1
            for m in graph[n]:
                in_degree[m] -= 1
                if in_degree[m] == 0:
                    q.append(m)

        return visited != len(nodes)

    def _schedule_completeness(self):
        checks = []
        total = len(self.activities) or 1

        checks.append(self.make_metric(
            'IND-101', 'Project Activities Defined',
            f'{len(self.activities)}',
            len(self.activities), 'Industry', 'Completeness',
            threshold_min=1, severity='critical',
            recommendation='Schedule must contain activities.'
        ))

        no_type = [a for a in self.activities if not a.get('task_type', '')]
        checks.append(self.make_check(
            'IND-102', 'Untyped Activities',
            'All activities need type classification',
            len(no_type), total, 0, 'Industry', 'high', 'Completeness',
            'Assign Task / Milestone / LOE / etc.',
            no_type
        ))

        no_wbs = [a for a in self.activities if not a.get('wbs_id', '')]
        checks.append(self.make_check(
            'IND-103', 'Missing WBS Assignment',
            'Activities need WBS assignment',
            len(no_wbs), total, 0, 'Industry', 'high', 'Completeness',
            'Assign activities to WBS nodes.',
            no_wbs
        ))

        no_cal = [a for a in self.activities if not a.get('clndr_id', '')]
        checks.append(self.make_check(
            'IND-104', 'Missing Calendar Assignment',
            'All activities need a calendar',
            len(no_cal), total, 0, 'Industry', 'high', 'Completeness',
            'Assign calendars to avoid default/silent calendar issues.',
            no_cal
        ))

        checks.append(self.make_metric(
            'IND-105', 'Milestone Presence',
            f'{len(self.milestones)} milestones',
            len(self.milestones), 'Industry', 'Completeness',
            threshold_min=3, severity='medium',
            recommendation='Projects should have key milestones for tracking.'
        ))

        return {'name': 'Schedule Completeness', 'checks': checks}

    def _logic_quality(self):
        checks = []
        total_inc = len(self.incomplete) or 1
        active_rels = self.active_relationships()
        rel_total = len(active_rels) or 1
        dens_denom = len(self.real_activities) or 1

        density = len(self.relationships) / dens_denom
        checks.append(self.make_metric(
            'IND-201', 'Logic Density',
            f'{density:.2f} rel / real activity',
            density, 'Industry', 'Logic',
            threshold_min=1.5, threshold_max=4.0, severity='medium',
            recommendation='Industry norm: roughly 1.5-4.0 relationships per activity.'
        ))

        fs = sum(1 for r in active_rels if r.get('pred_type') == 'PR_FS')
        fs_pct = fs / rel_total * 100
        checks.append(self.make_metric(
            'IND-202', 'FS Percentage (Active Logic)',
            f'{fs_pct:.1f}%',
            fs_pct, 'Industry', 'Logic',
            threshold_min=80, severity='medium',
            recommendation='Industry norm: 80%+ FS relationships.'
        ))

        open_start = self.open_start_activities()
        open_end = self.open_end_activities()
        dangling_ids = {str(a.get('task_id')) for a in open_start} | {
            str(a.get('task_id')) for a in open_end
        }
        dangling = [a for a in self.incomplete if str(a.get('task_id')) in dangling_ids]

        checks.append(self.make_check(
            'IND-203', 'Dangling Activities',
            'Incomplete non-milestone activities missing pred and/or succ',
            len(dangling), total_inc, 2, 'Industry', 'high', 'Logic',
            'Close the network: every work activity needs proper logic.',
            dangling
        ))

        has_cycle = self._has_circular_logic()
        checks.append(self.make_boolean(
            'IND-204', 'Circular Logic Free',
            'No circular dependencies detected in the network',
            not has_cycle, 'Industry', 'Logic',
            severity='critical',
            recommendation='Break cycle(s) — schedule will not calculate correctly.'
        ))

        open_s = open_start
        open_e = open_end
        checks.append(self.make_check(
            'IND-205', 'Open Start Activities',
            'Incomplete non-milestones without predecessors',
            len(open_s), total_inc, 1, 'Industry', 'high', 'Logic',
            'Only start milestones should lack predecessors.',
            open_s
        ))
        checks.append(self.make_check(
            'IND-206', 'Open End Activities',
            'Incomplete non-milestones without successors',
            len(open_e), total_inc, 1, 'Industry', 'high', 'Logic',
            'Only finish milestones should lack successors.',
            open_e
        ))

        leads = [r for r in active_rels if r.get('lag_days', 0) < 0]
        checks.append(self.make_check(
            'IND-207', 'Negative Lags (Leads)',
            'Leads distort CPM',
            len(leads), rel_total, 0, 'Industry', 'high', 'Logic',
            'Remove negative lags.',
            leads
        ))

        fs_lag = self.fs_with_lag()
        checks.append(self.make_check(
            'IND-208', 'FS + Lag Relationships',
            'Prefer explicit wait activities over FS lag',
            len(fs_lag), rel_total, 5, 'Industry', 'medium', 'Logic',
            'Replace lag with visible schedule activities where practical.',
            fs_lag
        ))

        return {'name': 'Logic Quality', 'checks': checks}

    def _resource_realism(self):
        checks = []

        checks.append(self.make_metric(
            'IND-301', 'Resource Assignments',
            f'{len(self.resources)}',
            len(self.resources), 'Industry', 'Resources',
            severity='low', info_only=True
        ))

        if not self.resources:
            checks.append(self.make_metric(
                'IND-300', 'Schedule Resource Loaded',
                'No TASKRSRC assignments found',
                None, 'Industry', 'Resources', info_only=True,
                recommendation='Resource loading is preferred for cost/crew realism; N/A if pure logic IMS.'
            ))
            return {'name': 'Resource Realism', 'checks': checks}

        tasks_with_res = set(str(r.get('task_id')) for r in self.resources)
        work_acts = [
            a for a in self.incomplete
            if not self.is_milestone(a) and a.get('task_type') != 'TT_LOE'
        ]
        no_res = [a for a in work_acts if str(a.get('task_id', '')) not in tasks_with_res]
        no_res_pct = len(no_res) / max(len(work_acts), 1) * 100

        checks.append(self.make_metric(
            'IND-302', 'Unresourced Work',
            f'{no_res_pct:.1f}%',
            no_res_pct, 'Industry', 'Resources',
            threshold_max=15, severity='medium',
            recommendation='Industry norm: <15% unresourced work when loading is required.'
        ))
        checks.append(self.make_check(
            'IND-302b', 'Unresourced Work Activities (List)',
            'Work activities without resource assignments',
            len(no_res), max(len(work_acts), 1), 15, 'Industry', 'medium', 'Resources',
            'Assign resources where required by contract/controls process.',
            no_res
        ))

        cost_loaded = sum(1 for r in self.resources if self.to_float(r.get('target_cost', '0')) > 0)
        cost_pct = cost_loaded / max(len(self.resources), 1) * 100
        checks.append(self.make_metric(
            'IND-303', 'Cost-Loaded Assignments',
            f'{cost_pct:.1f}%',
            cost_pct, 'Industry', 'Resources',
            threshold_min=85, severity='medium',
            recommendation='Cost loading enables budget and EVM analysis.'
        ))

        return {'name': 'Resource Realism', 'checks': checks}

    def _progress_transparency(self):
        checks = []
        total = len(self.activities) or 1

        prog_issues = [
            a for a in self.activities
            if self.to_float(a.get('phys_complete_pct', '0')) > 0
            and not a.get('act_start_date', '')
        ]
        checks.append(self.make_check(
            'IND-401', 'Inconsistent Progress',
            'Progress without actual start',
            len(prog_issues), total, 0, 'Industry', 'critical', 'Progress',
            'Enter actual start when progress is claimed.',
            prog_issues
        ))

        if self.data_date:
            age = (datetime.now() - self.data_date).days
            checks.append(self.make_metric(
                'IND-402', 'Data Date Age',
                f'{age} days',
                age, 'Industry', 'Progress',
                threshold_max=30, severity='medium',
                info_only=(age > 365),
                recommendation='Industry norm: update monthly for live control.'
            ))

        comp_missing = [
            a for a in self.completed
            if not a.get('act_start_date', '') or not a.get('act_end_date', '')
        ]
        checks.append(self.make_check(
            'IND-403', 'Completed with Missing Actuals',
            'Complete activities need both actual start and finish',
            len(comp_missing), max(len(self.completed), 1), 0, 'Industry', 'high', 'Progress',
            'Backfill actual start/finish on completed work.',
            comp_missing
        ))

        finish_no_100 = [
            a for a in self.activities
            if a.get('act_end_date', '')
            and self.to_float(a.get('phys_complete_pct', '0')) < 100
        ]
        checks.append(self.make_check(
            'IND-404', 'Actual Finish Without 100%',
            'Finished activities should be 100% complete',
            len(finish_no_100), total, 0, 'Industry', 'high', 'Progress',
            'Set physical % to 100 when actual finish is recorded.',
            finish_no_100
        ))

        return {'name': 'Progress Transparency', 'checks': checks}

    def _schedule_optimization(self):
        checks = []
        total_inc = len(self.incomplete) or 1

        constrained = [a for a in self.incomplete if self.has_hard_constraint(a)]
        checks.append(self.make_check(
            'IND-501', 'Hard-Constrained Activities',
            'Hard constraints override CPM optimization',
            len(constrained), total_inc, 5, 'Industry', 'medium', 'Optimization',
            'Minimize MSO/MEO/Mandatory constraints for better CPM analysis.',
            constrained
        ))

        rel_tuples = [
            (str(r.get('pred_task_id', '')), str(r.get('task_id', '')), r.get('pred_type', ''))
            for r in self.relationships
        ]
        dup_rels = len(rel_tuples) - len(set(rel_tuples))
        checks.append(self.make_metric(
            'IND-502', 'Duplicate Relationships',
            f'{dup_rels} exact duplicate ties (same pred/succ/type)',
            dup_rels, 'Industry', 'Optimization',
            threshold_max=0, severity='low',
            recommendation='Remove duplicate identical relationships. SS+FF ladder pairs are not duplicates.'
        ))

        floats = [a.get('total_float_days', 0) for a in self.incomplete]
        if floats:
            avg = statistics.mean(floats)
            checks.append(self.make_metric(
                'IND-503', 'Average Float',
                f'{avg:.1f} days',
                avg, 'Industry', 'Optimization',
                threshold_min=5, threshold_max=60, severity='low', info_only=True,
                recommendation='Balance schedule pressure vs contingency.'
            ))

        alap = [
            a for a in self.incomplete
            if a.get('cstr_type') == 'CS_ALAP' or a.get('cstr_type2') == 'CS_ALAP'
        ]
        checks.append(self.make_check(
            'IND-504', 'ALAP Constraints',
            'ALAP consumes float and can hide risk',
            len(alap), total_inc, 2, 'Industry', 'medium', 'Optimization',
            'Minimize ALAP usage.',
            alap
        ))

        crit_inc = [a for a in self.incomplete if a.get('is_critical')]
        cp_pct = len(crit_inc) / total_inc * 100
        checks.append(self.make_metric(
            'IND-505', 'Critical Path % (Incomplete)',
            f'{cp_pct:.1f}%',
            cp_pct, 'Industry', 'Optimization',
            threshold_min=5, threshold_max=30, severity='medium',
            recommendation='Very high critical density often indicates over-constrained or poorly linked networks.'
        ))

        return {'name': 'Optimization Opportunities', 'checks': checks}

    def _maintainability(self):
        checks = []
        total = len(self.activities) or 1

        no_name = [a for a in self.activities if not str(a.get('task_name', '')).strip()]
        checks.append(self.make_check(
            'IND-601', 'Missing Names',
            'All activities need clear names',
            len(no_name), total, 0, 'Industry', 'critical', 'Maintainability',
            'Name every activity clearly.',
            no_name
        ))

        id_lengths = [len(a.get('task_code', '')) for a in self.activities if a.get('task_code')]
        if id_lengths:
            unique = len(set(id_lengths))
            checks.append(self.make_metric(
                'IND-602', 'ID Format Consistency',
                f'{unique} different lengths',
                unique, 'Industry', 'Maintainability',
                threshold_max=3, severity='low',
                recommendation='Use consistent activity ID format.'
            ))

        max_depth = self._wbs_max_depth()
        checks.append(self.make_metric(
            'IND-603', 'WBS Depth',
            f'{max_depth} levels (parent hierarchy)',
            max_depth, 'Industry', 'Maintainability',
            threshold_max=8, severity='low',
            recommendation='WBS deeper than 8 levels is often overly complex.'
        ))

        name_counts = Counter(
            str(a.get('task_name', '')).strip()
            for a in self.activities if str(a.get('task_name', '')).strip()
        )
        dup_names = {n for n, c in name_counts.items() if c > 1}
        dup_acts = [a for a in self.activities if str(a.get('task_name', '')).strip() in dup_names]
        checks.append(self.make_check(
            'IND-604', 'Duplicate Activity Names',
            'Names should be unique where practical',
            len(dup_acts), total, 5, 'Industry', 'low', 'Maintainability',
            'Add distinguishing context to duplicate names.',
            dup_acts
        ))

        code_counts = Counter(
            a.get('task_code', '') for a in self.activities if a.get('task_code')
        )
        dup_codes = {c for c, n in code_counts.items() if n > 1}
        dup_code_acts = [a for a in self.activities if a.get('task_code') in dup_codes]
        checks.append(self.make_check(
            'IND-605', 'Duplicate Activity IDs',
            'Activity IDs should be unique',
            len(dup_code_acts), total, 0, 'Industry', 'critical', 'Maintainability',
            'Fix duplicate IDs (or separate multi-project exports).',
            dup_code_acts
        ))

        return {'name': 'Maintainability', 'checks': checks}
'''

industry_path = os.path.join("health_standards", "industry_checks.py")
with open(industry_path, "w", encoding="utf-8") as f:
    f.write(INDUSTRY_CHECKS_CODE)
print("  ✅ Updated health_standards/industry_checks.py")


# ------------------------------------------------------------------------------
# 4. evm_engine.py
# ------------------------------------------------------------------------------
EVM_ENGINE_CODE = '''"""
EVM (EARNED VALUE MANAGEMENT) & S-CURVE ENGINE
================================================
Calculates project performance metrics and generates S-curve data.
"""

from datetime import datetime, timedelta
from collections import defaultdict
import logging

logger = logging.getLogger(__name__)


class EVMEngine:
    FALLBACK_COST_PER_DAY = 1000.0

    def __init__(self, engine):
        self.engine = engine
        self.evm_data = {}
        self.scurve_data = {}

        self.resource_costs = defaultdict(lambda: {'budget': 0.0, 'actual': 0.0})
        for r in getattr(engine, 'resources', []) or []:
            tid = str(r.get('task_id', '') or '')
            if not tid:
                continue
            self.resource_costs[tid]['budget'] += self._to_float(r.get('target_cost', '0'))
            self.resource_costs[tid]['actual'] += (
                self._to_float(r.get('act_reg_cost', '0'))
                + self._to_float(r.get('act_ot_cost', '0'))
                + self._to_float(r.get('act_cost', '0'))
            )

        total_budget = sum(c['budget'] for c in self.resource_costs.values())
        total_actual = sum(c['actual'] for c in self.resource_costs.values())
        self.has_real_costs = total_budget > 0 or total_actual > 0

    def calculate(self):
        logger.info("📊 Calculating EVM metrics...")
        self._calculate_evm_metrics()
        self._generate_scurve_data()
        logger.info("  ✅ EVM calculations complete (cost_loaded=%s)", self.has_real_costs)
        return {
            'metrics': self.evm_data,
            'scurve': self.scurve_data,
        }

    def _activity_budget_and_actual(self, act):
        tid = str(act.get('task_id', '') or '')
        duration = float(act.get('original_duration_days', 0) or 0)
        rc = self.resource_costs.get(tid, {'budget': 0.0, 'actual': 0.0})

        if self.has_real_costs:
            budget = float(rc['budget'] or 0.0)
            actual = float(rc['actual'] or 0.0)
            return budget, actual, False

        if duration <= 0:
            return 0.0, None, True
        budget = duration * self.FALLBACK_COST_PER_DAY
        return budget, None, True

    def _calculate_evm_metrics(self):
        activities = self.engine.activities
        data_date = self._get_data_date()

        total_bac = 0.0
        total_pv = 0.0
        total_ev = 0.0
        total_ac = 0.0
        ac_known = False

        for act in activities:
            if act.get('task_type') in ('TT_WBS', 'TT_LOE'):
                continue

            budget, actual, _ = self._activity_budget_and_actual(act)
            if budget <= 0 and (actual is None or actual <= 0):
                continue

            total_bac += budget

            progress = self._to_float(act.get('phys_complete_pct', '0')) / 100.0
            progress = max(0.0, min(1.0, progress))
            activity_ev = budget * progress
            total_ev += activity_ev

            if actual is not None and self.has_real_costs:
                total_ac += actual
                if actual > 0:
                    ac_known = True
            else:
                total_ac += activity_ev

            start_date = (
                act.get('target_start_date_parsed')
                or act.get('early_start_date_parsed')
            )
            end_date = (
                act.get('target_end_date_parsed')
                or act.get('early_end_date_parsed')
            )

            if start_date and end_date and data_date and budget > 0:
                if data_date < start_date:
                    planned_progress = 0.0
                elif data_date >= end_date:
                    planned_progress = 1.0
                else:
                    total_days = (end_date - start_date).days
                    days_done = (data_date - start_date).days
                    planned_progress = (days_done / total_days) if total_days > 0 else 0.0
                    planned_progress = max(0.0, min(1.0, planned_progress))
                total_pv += budget * planned_progress

        sv = total_ev - total_pv
        cv = total_ev - total_ac

        spi = (total_ev / total_pv) if total_pv > 0 else 0.0
        cpi = (total_ev / total_ac) if total_ac > 0 else 0.0
        eac = (total_bac / cpi) if cpi > 0 else total_bac

        etc = max(0.0, eac - total_ac)
        vac = total_bac - eac
        pct_complete = (total_ev / total_bac * 100.0) if total_bac > 0 else 0.0
        pct_spent = (total_ac / total_bac * 100.0) if total_bac > 0 else 0.0

        schedule_status = self._interpret_spi(spi if total_pv > 0 else 0.0)
        if not self.has_real_costs:
            cost_status = {
                'status': 'unknown',
                'text': 'N/A — schedule not cost-loaded',
            }
            cpi_out = cpi
        else:
            cost_status = self._interpret_cpi(cpi if total_ac > 0 else 0.0)
            cpi_out = cpi

        self.evm_data = {
            'bac': round(total_bac, 2),
            'pv': round(total_pv, 2),
            'ev': round(total_ev, 2),
            'ac': round(total_ac, 2),
            'sv': round(sv, 2),
            'cv': round(cv, 2),
            'spi': round(spi, 3),
            'cpi': round(cpi_out, 3),
            'eac': round(eac, 2),
            'etc': round(etc, 2),
            'vac': round(vac, 2),
            'pct_complete': round(pct_complete, 1),
            'pct_spent': round(pct_spent, 1),
            'is_cost_loaded': bool(self.has_real_costs),
            'ac_from_actuals': bool(self.has_real_costs and ac_known),
            'estimate_method': 'resource_cost' if self.has_real_costs else 'duration_proxy',
            'schedule_status': schedule_status,
            'cost_status': cost_status,
            'data_date': data_date.strftime('%Y-%m-%d') if data_date else 'Unknown',
        }

    def _generate_scurve_data(self):
        activities = [
            a for a in self.engine.activities
            if a.get('task_type') not in ('TT_WBS', 'TT_LOE')
        ]

        all_dates = []
        for act in activities:
            start = act.get('target_start_date_parsed') or act.get('early_start_date_parsed')
            end = act.get('target_end_date_parsed') or act.get('early_end_date_parsed')
            if start:
                all_dates.append(start)
            if end:
                all_dates.append(end)

        if not all_dates:
            self.scurve_data = {'error': 'No valid dates found'}
            return

        proj_start = min(all_dates)
        proj_end = max(all_dates)
        data_date = self._get_data_date()

        buckets = []
        current = proj_start
        max_buckets = 520
        while current <= proj_end and len(buckets) < max_buckets:
            buckets.append(current)
            current += timedelta(days=7)
        if not buckets or buckets[-1] < proj_end:
            buckets.append(proj_end)

        pv_curve = []
        ev_curve = []
        ac_curve = []
        labels = []

        for bucket_date in buckets:
            planned_value = 0.0
            earned_value = 0.0
            actual_cost = 0.0

            for act in activities:
                budget, actual, _ = self._activity_budget_and_actual(act)
                if budget <= 0:
                    continue

                start = act.get('target_start_date_parsed') or act.get('early_start_date_parsed')
                end = act.get('target_end_date_parsed') or act.get('early_end_date_parsed')
                if not start or not end:
                    continue

                if bucket_date >= end:
                    planned_value += budget
                elif bucket_date > start:
                    total_days = (end - start).days
                    days_done = (bucket_date - start).days
                    if total_days > 0:
                        planned_value += budget * min(1.0, max(0.0, days_done / total_days))

            if data_date and bucket_date <= data_date:
                for act in activities:
                    budget, actual, _ = self._activity_budget_and_actual(act)
                    if budget <= 0:
                        continue
                    start = act.get('target_start_date_parsed') or act.get('early_start_date_parsed')
                    end = act.get('target_end_date_parsed') or act.get('early_end_date_parsed')
                    if not start or not end:
                        continue

                    progress = max(0.0, min(1.0, self._to_float(act.get('phys_complete_pct', '0')) / 100.0))

                    if bucket_date < start:
                        portion = 0.0
                    elif bucket_date >= end:
                        portion = 1.0
                    else:
                        td = (end - start).days
                        portion = ((bucket_date - start).days / td) if td > 0 else 0.0
                        portion = max(0.0, min(1.0, portion))

                    if progress >= 1.0:
                        earned_value += budget * portion
                    else:
                        earned_value += budget * progress * portion

                    if self.has_real_costs and actual is not None:
                        actual_cost += float(actual) * portion
                    else:
                        actual_cost += budget * progress * portion

            labels.append(bucket_date.strftime('%Y-%m-%d'))
            pv_curve.append(round(planned_value, 2))

            if data_date and bucket_date <= data_date:
                ev_curve.append(round(earned_value, 2))
                ac_curve.append(round(actual_cost, 2))
            else:
                ev_curve.append(None)
                ac_curve.append(None)

        self.scurve_data = {
            'labels': labels,
            'planned_value': pv_curve,
            'earned_value': ev_curve,
            'actual_cost': ac_curve,
            'data_date': data_date.strftime('%Y-%m-%d') if data_date else None,
            'bac': self.evm_data.get('bac', 0),
            'is_cost_loaded': bool(self.has_real_costs),
        }

    def _get_data_date(self):
        if not getattr(self.engine, 'projects', None):
            return datetime.now()
        proj = self.engine.projects[0]
        date_str = proj.get('last_recalc_date', '')
        parsed = self.engine._parse_date(date_str)
        return parsed if parsed else datetime.now()

    def _interpret_spi(self, spi):
        if spi == 0:
            return {'status': 'unknown', 'text': 'Insufficient data'}
        if spi >= 1.0:
            return {'status': 'good', 'text': 'Ahead of schedule ✅'}
        if spi >= 0.95:
            return {'status': 'warning', 'text': 'Slightly behind ⚠️'}
        return {'status': 'bad', 'text': 'Behind schedule ❌'}

    def _interpret_cpi(self, cpi):
        if cpi == 0:
            return {'status': 'unknown', 'text': 'Insufficient data'}
        if cpi >= 1.0:
            return {'status': 'good', 'text': 'Under budget ✅'}
        if cpi >= 0.95:
            return {'status': 'warning', 'text': 'Slightly over budget ⚠️'}
        return {'status': 'bad', 'text': 'Over budget ❌'}

    def _to_float(self, value):
        try:
            return float(value)
        except (ValueError, TypeError):
            return 0.0
'''

with open("evm_engine.py", "w", encoding="utf-8") as f:
    f.write(EVM_ENGINE_CODE)
print("  ✅ Updated evm_engine.py")


# ------------------------------------------------------------------------------
# 5. comparison_engine.py
# ------------------------------------------------------------------------------
COMPARISON_ENGINE_CODE = '''"""
XER COMPARISON ENGINE
=====================
Compares two XER files (e.g., Baseline vs Current) and identifies all differences.
"""

from parser import XERParser
from data_engine import ScheduleEngine
from datetime import datetime
import logging

logger = logging.getLogger(__name__)


class ScheduleComparator:
    def __init__(self):
        self.baseline_engine = None
        self.current_engine = None
        self.comparison_results = {}

    def load_baseline(self, file_path_or_stream):
        logger.info(f"📊 Loading Baseline: {file_path_or_stream}")
        parser = XERParser()
        tables = parser.parse(file_path_or_stream)
        
        if tables is None:
            raise Exception("Failed to parse baseline schedule file.")
        
        self.baseline_engine = ScheduleEngine()
        self.baseline_engine.load_data(tables)
        self.baseline_engine.analyze()

    def load_current(self, file_path_or_stream):
        logger.info(f"📊 Loading Current: {file_path_or_stream}")
        parser = XERParser()
        tables = parser.parse(file_path_or_stream)
        
        if tables is None:
            raise Exception("Failed to parse current schedule file.")
        
        self.current_engine = ScheduleEngine()
        self.current_engine.load_data(tables)
        self.current_engine.analyze()

    def compare(self):
        if not self.baseline_engine or not self.current_engine:
            raise Exception("Both baseline and current schedules must be loaded first.")
        
        logger.info("🔍 Comparing schedules...")
        
        baseline_acts = {
            a.get('task_code', ''): a 
            for a in self.baseline_engine.activities
            if a.get('task_code')
        }
        current_acts = {
            a.get('task_code', ''): a 
            for a in self.current_engine.activities
            if a.get('task_code')
        }
        
        baseline_codes = set(baseline_acts.keys())
        current_codes = set(current_acts.keys())
        
        added_codes = current_codes - baseline_codes
        deleted_codes = baseline_codes - current_codes
        common_codes = baseline_codes & current_codes
        
        added = [self._format_activity(current_acts[code], 'added') for code in added_codes]
        deleted = [self._format_activity(baseline_acts[code], 'deleted') for code in deleted_codes]
        
        changed = []
        unchanged_count = 0
        
        for code in common_codes:
            baseline_act = baseline_acts[code]
            current_act = current_acts[code]
            
            changes = self._detect_changes(baseline_act, current_act)
            
            if changes:
                changed.append({
                    'code': code,
                    'name': current_act.get('task_name', ''),
                    'wbs': current_act.get('wbs_name', ''),
                    'changes': changes,
                    'baseline': self._format_activity(baseline_act, 'baseline'),
                    'current': self._format_activity(current_act, 'current'),
                })
            else:
                unchanged_count += 1
        
        summary = self._calculate_summary(added, deleted, changed, unchanged_count)
        critical_changes = self._analyze_critical_path_changes(baseline_acts, current_acts, common_codes)
        relationship_changes = self._compare_relationships()
        
        self.comparison_results = {
            'summary': summary,
            'added': added,
            'deleted': deleted,
            'changed': changed,
            'critical_changes': critical_changes,
            'relationship_changes': relationship_changes,
            'baseline_info': self._get_schedule_info(self.baseline_engine),
            'current_info': self._get_schedule_info(self.current_engine),
        }
        
        return self.comparison_results

    def _detect_changes(self, baseline, current):
        changes = []
        
        base_dur = float(baseline.get('original_duration_days', 0) or 0)
        curr_dur = float(current.get('original_duration_days', 0) or 0)
        dur_diff = curr_dur - base_dur
        if abs(dur_diff) > 0.1:
            changes.append({
                'field': 'Duration',
                'baseline': f"{base_dur:.1f}d",
                'current': f"{curr_dur:.1f}d",
                'delta': f"{dur_diff:+.1f}d",
                'delta_days': dur_diff,
                'severity': 'high' if abs(dur_diff) > 5 else 'medium'
            })
        
        base_start = self._get_best_start_date(baseline)
        curr_start = self._get_best_start_date(current)
        if base_start and curr_start and base_start != curr_start:
            delta_days = (curr_start - base_start).days
            if delta_days != 0:
                changes.append({
                    'field': 'Start Date',
                    'baseline': base_start.strftime('%Y-%m-%d'),
                    'current': curr_start.strftime('%Y-%m-%d'),
                    'delta': f"{delta_days:+d}d",
                    'delta_days': delta_days,
                    'severity': 'high' if abs(delta_days) > 7 else 'medium'
                })
        
        base_end = self._get_best_finish_date(baseline)
        curr_end = self._get_best_finish_date(current)
        if base_end and curr_end and base_end != curr_end:
            delta_days = (curr_end - base_end).days
            if delta_days != 0:
                changes.append({
                    'field': 'Finish Date',
                    'baseline': base_end.strftime('%Y-%m-%d'),
                    'current': curr_end.strftime('%Y-%m-%d'),
                    'delta': f"{delta_days:+d}d",
                    'delta_days': delta_days,
                    'severity': 'high' if abs(delta_days) > 7 else 'medium'
                })
        
        base_float = float(baseline.get('total_float_days', 0) or 0)
        curr_float = float(current.get('total_float_days', 0) or 0)
        float_diff = curr_float - base_float
        if abs(float_diff) > 0.5:
            changes.append({
                'field': 'Total Float',
                'baseline': f"{base_float:.1f}d",
                'current': f"{curr_float:.1f}d",
                'delta': f"{float_diff:+.1f}d",
                'delta_days': float_diff,
                'severity': 'high' if curr_float < 0 and base_float >= 0 else 'medium'
            })
        
        base_status = baseline.get('status_text', '')
        curr_status = current.get('status_text', '')
        if base_status != curr_status:
            changes.append({
                'field': 'Status',
                'baseline': base_status,
                'current': curr_status,
                'delta': '→',
                'delta_days': 0,
                'severity': 'low'
            })
        
        base_prog = self._to_float(baseline.get('phys_complete_pct', '0'))
        curr_prog = self._to_float(current.get('phys_complete_pct', '0'))
        prog_diff = curr_prog - base_prog
        if abs(prog_diff) > 0.5:
            changes.append({
                'field': 'Progress',
                'baseline': f"{base_prog:.0f}%",
                'current': f"{curr_prog:.0f}%",
                'delta': f"{prog_diff:+.0f}%",
                'delta_days': prog_diff,
                'severity': 'low'
            })
        
        return changes

    def _compare_relationships(self):
        base_rels = {
            f"{r.get('pred_code')}->{r.get('succ_code')}": r
            for r in self.baseline_engine.relationships
            if r.get('pred_code') and r.get('succ_code')
        }
        curr_rels = {
            f"{r.get('pred_code')}->{r.get('succ_code')}": r
            for r in self.current_engine.relationships
            if r.get('pred_code') and r.get('succ_code')
        }
        
        added_keys = set(curr_rels.keys()) - set(base_rels.keys())
        deleted_keys = set(base_rels.keys()) - set(curr_rels.keys())
        common_keys = set(base_rels.keys()) & set(curr_rels.keys())
        
        modified_logic = []
        for k in common_keys:
            b = base_rels[k]
            c = curr_rels[k]
            
            type_changed = b.get('pred_type') != c.get('pred_type')
            lag_diff = float(c.get('lag_days', 0) or 0) - float(b.get('lag_days', 0) or 0)
            
            if type_changed or abs(lag_diff) > 0.1:
                modified_logic.append({
                    'tie': k,
                    'pred_code': b.get('pred_code'),
                    'succ_code': b.get('succ_code'),
                    'pred_name': c.get('pred_name', ''),
                    'succ_name': c.get('succ_name', ''),
                    'baseline_type': b.get('type_text', ''),
                    'current_type': c.get('type_text', ''),
                    'baseline_lag': round(float(b.get('lag_days', 0) or 0), 1),
                    'current_lag': round(float(c.get('lag_days', 0) or 0), 1),
                    'lag_delta': f"{lag_diff:+.1f}d"
                })

        return {
            'added_count': len(added_keys),
            'deleted_count': len(deleted_keys),
            'modified_count': len(modified_logic),
            'modified_details': modified_logic
        }

    def _analyze_critical_path_changes(self, baseline_acts, current_acts, common):
        newly_critical = []
        no_longer_critical = []
        
        for code in common:
            base_crit = bool(baseline_acts[code].get('is_critical', False))
            curr_crit = bool(current_acts[code].get('is_critical', False))
            
            if not base_crit and curr_crit:
                newly_critical.append({
                    'code': code,
                    'name': current_acts[code].get('task_name', ''),
                    'wbs': current_acts[code].get('wbs_name', ''),
                    'float': round(float(current_acts[code].get('total_float_days', 0) or 0), 1),
                })
            elif base_crit and not curr_crit:
                no_longer_critical.append({
                    'code': code,
                    'name': current_acts[code].get('task_name', ''),
                    'wbs': current_acts[code].get('wbs_name', ''),
                    'float': round(float(current_acts[code].get('total_float_days', 0) or 0), 1),
                })
        
        return {
            'newly_critical': newly_critical,
            'no_longer_critical': no_longer_critical,
        }

    def _calculate_summary(self, added, deleted, changed, unchanged):
        slipped = 0
        improved = 0
        
        for change_item in changed:
            for c in change_item['changes']:
                if c['field'] == 'Finish Date':
                    delta_days = c.get('delta_days', 0)
                    if delta_days > 0:
                        slipped += 1
                    elif delta_days < 0:
                        improved += 1
                    break
        
        return {
            'total_baseline': len(deleted) + len(changed) + unchanged,
            'total_current': len(added) + len(changed) + unchanged,
            'added_count': len(added),
            'deleted_count': len(deleted),
            'changed_count': len(changed),
            'unchanged_count': unchanged,
            'slipped_count': slipped,
            'improved_count': improved,
        }

    def _format_activity(self, act, source):
        start = self._get_best_start_date(act)
        finish = self._get_best_finish_date(act)
        return {
            'code': act.get('task_code', ''),
            'name': act.get('task_name', ''),
            'wbs': act.get('wbs_name', ''),
            'duration': round(float(act.get('original_duration_days', 0) or 0), 1),
            'float': round(float(act.get('total_float_days', 0) or 0), 1),
            'start': start.strftime('%Y-%m-%d') if start else '',
            'finish': finish.strftime('%Y-%m-%d') if finish else '',
            'status': act.get('status_text', ''),
            'critical': bool(act.get('is_critical', False)),
            'source': source,
        }

    def _get_best_start_date(self, act):
        return act.get('act_start_date_parsed') or \\
               act.get('early_start_date_parsed') or \\
               act.get('target_start_date_parsed')

    def _get_best_finish_date(self, act):
        return act.get('act_end_date_parsed') or \\
               act.get('early_end_date_parsed') or \\
               act.get('target_end_date_parsed')

    def _get_schedule_info(self, engine):
        info = engine._get_project_info() if hasattr(engine, '_get_project_info') else {}
        return {
            'name': info.get('name', 'Unknown'),
            'total_activities': len(engine.activities),
            'critical_count': len(engine.critical_activities),
            'total_relationships': len(engine.relationships),
        }

    def _to_float(self, value):
        try:
            return float(value)
        except (ValueError, TypeError):
            return 0.0
'''

with open("comparison_engine.py", "w", encoding="utf-8") as f:
    f.write(COMPARISON_ENGINE_CODE)
print("  ✅ Updated comparison_engine.py")


# ------------------------------------------------------------------------------
# 6. reports.py
# ------------------------------------------------------------------------------
REPORTS_CODE = '''"""
REPORT GENERATOR
================
Creates Excel reports from the Schedule Engine results.
"""

import os
from datetime import datetime
import logging

logger = logging.getLogger(__name__)


class ReportGenerator:
    def __init__(self, engine):
        self.engine = engine

    def generate_full_report(self, output_path):
        try:
            import pandas as pd
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError as e:
            logger.error("pandas/openpyxl required: %s", e)
            print("❌ pandas and openpyxl are required. Run: pip install pandas openpyxl")
            return None

        logger.info("📝 Generating report: %s", output_path)

        out_dir = os.path.dirname(os.path.abspath(output_path))
        if out_dir:
            os.makedirs(out_dir, exist_ok=True)

        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            self._write_summary_tab(writer, pd)
            self._write_dcma_tab(writer, pd)
            self._write_activities_tab(writer, pd)
            self._write_critical_path_tab(writer, pd)
            self._write_relationships_tab(writer, pd)
            self._write_issues_tab(writer, pd)

            workbook = writer.book
            for sheet_name in workbook.sheetnames:
                self._style_worksheet(
                    workbook[sheet_name],
                    pass_fail_col=(3 if sheet_name == 'DCMA 14-Point' else None),
                )

        logger.info("✅ Report saved: %s", output_path)
        print(f"  ✅ Report saved: {output_path}")
        return output_path

    def _write_summary_tab(self, writer, pd):
        stats = self.engine.schedule_stats or {}
        info = {}
        if hasattr(self.engine, '_get_project_info'):
            info = self.engine._get_project_info() or {}

        rows = [
            ['Schedule Summary Report', ''],
            ['Generated', datetime.now().strftime('%Y-%m-%d %H:%M')],
            ['', ''],
            ['PROJECT', ''],
            ['Project Name', info.get('name', '')],
            ['Plan Start', info.get('start', '')],
            ['Plan Finish', info.get('finish', '')],
            ['Data Date', info.get('data_date', '')],
            ['', ''],
            ['ACTIVITY COUNTS', ''],
            ['Total Activities', stats.get('total_activities', 0)],
            ['  Not Started', stats.get('not_started', 0)],
            ['  In Progress', stats.get('in_progress', 0)],
            ['  Completed', stats.get('completed', 0)],
            ['', ''],
            ['ACTIVITY TYPES', ''],
            ['  Tasks', stats.get('tasks', 0)],
            ['  Milestones', stats.get('milestones', 0)],
            ['  Level of Effort', stats.get('loe', 0)],
            ['  WBS Summary', stats.get('wbs_summary', 0)],
            ['', ''],
            ['CRITICAL PATH', ''],
            ['  Critical Activities', stats.get('critical_count', 0)],
            ['  Negative Float', stats.get('negative_float', 0)],
            ['  High Float (>44d)', stats.get('high_float_gt_44d', 0)],
            ['', ''],
            ['NETWORK', ''],
            ['  Total Relationships', stats.get('total_relationships', 0)],
            ['  Calendars', stats.get('total_calendars', 0)],
            ['', ''],
            ['Note', 'Basic DCMA dashboard export. Use Health → Excel for 622+ standards.'],
        ]

        df = pd.DataFrame(rows, columns=['Metric', 'Value'])
        df.to_excel(writer, sheet_name='Summary', index=False)

    def _write_dcma_tab(self, writer, pd):
        rows = []
        for check_name, result in (self.engine.dcma_results or {}).items():
            clean_name = check_name.replace('_', ' ')
            has_pct = 'pct' in result

            if result.get('pass') is True:
                status = 'PASS'
            elif result.get('pass') is False:
                status = 'FAIL'
            else:
                status = 'N/A'

            if has_pct:
                value_display = f"{result.get('pct', '')}%"
                count_display = result.get('count', '')
                total_display = result.get('total', '')
            elif 'value' in result:
                value_display = str(result.get('value', ''))
                count_display = ''
                total_display = ''
            else:
                value_display = ''
                count_display = result.get('count', '')
                total_display = result.get('total', '')

            rows.append({
                'Check': clean_name,
                'Value': value_display,
                'Count': count_display,
                'Total': total_display,
                'Threshold': result.get('threshold', ''),
                'Result': status,
            })

        if not rows:
            rows = [{'Check': 'No DCMA results', 'Value': '', 'Count': '', 'Total': '', 'Threshold': '', 'Result': 'N/A'}]

        df = pd.DataFrame(rows)[['Check', 'Value', 'Count', 'Total', 'Threshold', 'Result']]
        df.to_excel(writer, sheet_name='DCMA 14-Point', index=False)

    def _write_activities_tab(self, writer, pd):
        df = None
        if hasattr(self.engine, 'get_activities_dataframe'):
            df = self.engine.get_activities_dataframe()

        if df is None or df.empty:
            cols = ['task_code', 'task_name', 'wbs_code', 'wbs_name', 'status_text', 'type_text',
                    'original_duration_days', 'remaining_duration_days', 'total_float_days',
                    'is_critical', 'early_start_date', 'early_end_date', 'phys_complete_pct']
            data = [{c: a.get(c, '') for c in cols} for a in self.engine.activities]
            df = pd.DataFrame(data)

        rename = {
            'task_code': 'Activity ID', 'task_name': 'Activity Name', 'wbs_code': 'WBS Code',
            'wbs_name': 'WBS', 'status_text': 'Status', 'type_text': 'Type',
            'original_duration_days': 'Original Duration (d)', 'remaining_duration_days': 'Remaining Duration (d)',
            'total_float_days': 'Total Float (d)', 'free_float_days': 'Free Float (d)',
            'is_critical': 'Critical', 'early_start_date': 'Early Start', 'early_end_date': 'Early Finish',
            'late_start_date': 'Late Start', 'late_end_date': 'Late Finish', 'target_start_date': 'Target Start',
            'target_end_date': 'Target Finish', 'act_start_date': 'Actual Start', 'act_end_date': 'Actual Finish',
            'phys_complete_pct': 'Physical %',
        }
        df = df.rename(columns={k: v for k, v in rename.items() if k in df.columns})
        df.to_excel(writer, sheet_name='Activities', index=False)

    def _write_critical_path_tab(self, writer, pd):
        columns = ['task_code', 'task_name', 'wbs_name', 'original_duration_days', 'total_float_days', 'early_start_date', 'early_end_date', 'status_text']
        rename = {'task_code': 'Activity ID', 'task_name': 'Activity Name', 'wbs_name': 'WBS', 'original_duration_days': 'Duration (d)', 'total_float_days': 'Total Float (d)', 'early_start_date': 'Early Start', 'early_end_date': 'Early Finish', 'status_text': 'Status'}

        crit = getattr(self.engine, 'critical_activities', []) or []
        rows = [{col: act.get(col, '') for col in columns} for act in crit]

        if not rows:
            df = pd.DataFrame([{'Info': 'No critical activities found'}])
        else:
            df = pd.DataFrame(rows).rename(columns=rename)

        df.to_excel(writer, sheet_name='Critical Path', index=False)

    def _write_relationships_tab(self, writer, pd):
        columns = ['pred_code', 'pred_name', 'succ_code', 'succ_name', 'type_text', 'lag_days']
        rename = {'pred_code': 'Predecessor ID', 'pred_name': 'Predecessor Name', 'succ_code': 'Successor ID', 'succ_name': 'Successor Name', 'type_text': 'Type', 'lag_days': 'Lag (d)'}

        rels = getattr(self.engine, 'relationships', []) or []
        rows = [{col: rel.get(col, '') for col in columns} for rel in rels]

        if not rows:
            df = pd.DataFrame([{'Info': 'No relationships found'}])
        else:
            df = pd.DataFrame(rows).rename(columns=rename)

        df.to_excel(writer, sheet_name='Relationships', index=False)

    def _write_issues_tab(self, writer, pd):
        issues = []

        for check_name, result in (self.engine.dcma_results or {}).items():
            if result.get('pass') is not False:
                continue

            items = result.get('activities') or result.get('items') or []
            clean_check = check_name.replace('_', ' ')

            if not items:
                issues.append({
                    'Issue Type': clean_check, 'Activity ID': '', 'Activity Name': '(No item list)',
                    'WBS': '', 'Status': '', 'Total Float': '', 'Duration': '', 'Detail': f"Count={result.get('count', '')}",
                })
                continue

            for item in items[:100]:
                code, name, extra = self._format_issue_item(item)
                issues.append({
                    'Issue Type': clean_check,
                    'Activity ID': code,
                    'Activity Name': name,
                    'WBS': (item.get('wbs_name') or '')[:60] if isinstance(item, dict) else '',
                    'Status': item.get('status_text', '') if isinstance(item, dict) else '',
                    'Total Float': item.get('total_float_days', '') if isinstance(item, dict) else '',
                    'Duration': item.get('original_duration_days', '') if isinstance(item, dict) else '',
                    'Detail': extra,
                })

        if not issues:
            df = pd.DataFrame([{'Info': 'No failed DCMA checks'}])
        else:
            df = pd.DataFrame(issues)

        df.to_excel(writer, sheet_name='Issues', index=False)

    def _format_issue_item(self, item):
        if not isinstance(item, dict):
            return str(item), '', ''

        if ('pred_code' in item or 'succ_code' in item or 'pred_task_id' in item or item.get('pred_type') or item.get('type_text')):
            p = item.get('pred_code') or item.get('pred_task_id', '')
            s = item.get('succ_code') or item.get('task_code') or item.get('task_id', '')
            pn = item.get('pred_name', '')
            sn = item.get('succ_name', '')
            rt = item.get('type_text') or item.get('pred_type', '')
            lag = item.get('lag_days', '')
            return f"{p} → {s}", f"{pn} → {sn}", f"{rt}, lag={lag}"

        return item.get('task_code', '') or item.get('wbs_short_name', ''), item.get('task_name', '') or item.get('wbs_name', ''), ''

    def _style_worksheet(self, ws, pass_fail_col=None):
        from openpyxl.styles import Font, PatternFill, Alignment

        header_font = Font(bold=True, color='FFFFFF', size=11)
        header_fill = PatternFill(start_color='1E40AF', end_color='1E40AF', fill_type='solid')
        green = PatternFill(start_color='D1FAE5', end_color='D1FAE5', fill_type='solid')
        red = PatternFill(start_color='FEE2E2', end_color='FEE2E2', fill_type='solid')
        gray = PatternFill(start_color='F1F5F9', end_color='F1F5F9', fill_type='solid')

        if ws.max_row < 1:
            return

        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='left', vertical='center')

        ws.freeze_panes = 'A2'

        for col in ws.columns:
            max_len = 8
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    val = str(cell.value) if cell.value is not None else ''
                    if len(val) > max_len:
                        max_len = len(val)
                except Exception:
                    pass
            ws.column_dimensions[col_letter].width = min(max_len + 2, 55)

        result_col_idx = None
        for idx, cell in enumerate(ws[1], start=1):
            if str(cell.value).strip().lower() == 'result':
                result_col_idx = idx
                break

        if result_col_idx:
            for row in ws.iter_rows(min_row=2, min_col=result_col_idx, max_col=result_col_idx):
                cell = row[0]
                val = str(cell.value or '').upper()
                if val == 'PASS':
                    cell.fill = green
                    cell.font = Font(bold=True, color='065F46')
                elif val == 'FAIL':
                    cell.fill = red
                    cell.font = Font(bold=True, color='991B1B')
                elif val == 'N/A':
                    cell.fill = gray
'''

with open("reports.py", "w", encoding="utf-8") as f:
    f.write(REPORTS_CODE)
print("  ✅ Updated reports.py")


# ------------------------------------------------------------------------------
# 7. pdf_report_generator.py
# ------------------------------------------------------------------------------
PDF_REPORT_GENERATOR_CODE = '''"""
PDF EXECUTIVE REPORT GENERATOR
================================
"""

from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
    PageBreak, HRFlowable, KeepTogether
)
from datetime import datetime
from xml.sax.saxutils import escape
import io
import logging

logger = logging.getLogger(__name__)

try:
    from config import (
        COMPANY_NAME,
        APP_TITLE,
        MAX_ITEMS_PER_CHECK_PDF,
        MAX_TOP_ACTIONS_PDF,
        get_theme,
    )
except ImportError:
    COMPANY_NAME = "P6 Schedule Analyzer"
    APP_TITLE = "P6 Schedule Analyzer"
    MAX_ITEMS_PER_CHECK_PDF = 50
    MAX_TOP_ACTIONS_PDF = 15

    def get_theme():
        return {
            'primary': '#1e40af',
            'success': '#10b981',
            'warning': '#f59e0b',
            'danger': '#dc2626',
            'muted': '#64748b',
        }


SEVERITY_LEVELS = {
    'critical': ['critical'],
    'high': ['critical', 'high'],
    'medium': ['critical', 'high', 'medium'],
    'all': ['critical', 'high', 'medium', 'low', 'info'],
}

SEVERITY_WEIGHT = {
    'critical': 100,
    'high': 50,
    'medium': 20,
    'low': 5,
    'info': 0,
}


class PDFReportGenerator:
    def __init__(self, health_data, file_name='', severity_filter='all'):
        self.data = health_data or {}
        self.file_name = file_name or ''
        self.severity_filter = (severity_filter or 'all').lower()
        self.allowed_severities = SEVERITY_LEVELS.get(
            self.severity_filter, SEVERITY_LEVELS['all']
        )
        self.theme = get_theme()
        self.primary = colors.HexColor(self.theme.get('primary', '#1e40af'))
        self.danger = colors.HexColor(self.theme.get('danger', '#dc2626'))
        self.muted = colors.HexColor(self.theme.get('muted', '#64748b'))
        self.styles = self._create_styles()
        self.max_items = int(MAX_ITEMS_PER_CHECK_PDF or 50)
        self.max_top = int(MAX_TOP_ACTIONS_PDF or 15)

    def _safe(self, text):
        if text is None:
            return ''
        return escape(str(text))

    def _p(self, text, style_name='CheckBody'):
        style = self.styles[style_name]
        return Paragraph(self._safe(text), style)

    def _p_markup(self, markup, style_name='CheckBody'):
        return Paragraph(str(markup), self.styles[style_name])

    def _matches_severity(self, check_or_action):
        sev = (check_or_action.get('severity') or 'low').lower()
        return sev in self.allowed_severities

    def _activity_line(self, item):
        code = self._safe(item.get('code', ''))
        name = self._safe(item.get('name', ''))
        wbs = self._safe(item.get('wbs', ''))
        line = f"• {code}"
        if name:
            line += f" - {name}"
        if wbs:
            line += f" ({wbs})"
        return line

    def _iter_failed_checks(self):
        for std_name, std_data in (self.data.get('standards') or {}).items():
            for cat in std_data.get('categories') or []:
                cat_name = cat.get('name', '')
                for check in cat.get('checks') or []:
                    if check.get('status') != 'fail':
                        continue
                    if not self._matches_severity(check):
                        continue
                    yield std_name, cat_name, check

    def _build_priority_actions(self, limit=None):
        actions = []
        for std_name, cat_name, check in self._iter_failed_checks():
            count = check.get('count', 0) or 0
            sev = (check.get('severity') or 'low').lower()
            priority = SEVERITY_WEIGHT.get(sev, 5) + min(count, 100)
            actions.append({
                'standard': std_name,
                'category': cat_name,
                'id': check.get('id'),
                'name': check.get('name'),
                'severity': sev,
                'count': count,
                'total': check.get('total', 0),
                'percentage': check.get('percentage', 0),
                'value': check.get('value'),
                'threshold': check.get('threshold', ''),
                'description': check.get('description', ''),
                'recommendation': check.get('recommendation', ''),
                'failed_items': check.get('failed_items') or [],
                'priority': priority,
            })

        if not actions:
            for a in (self.data.get('top_actions') or []):
                if self._matches_severity(a):
                    actions.append(a)

        actions.sort(key=lambda x: x.get('priority', 0), reverse=True)
        if limit is not None:
            return actions[:limit]
        return actions

    def _score_color(self, score):
        try:
            s = float(score)
        except (TypeError, ValueError):
            s = 0
        if s >= 90:
            return colors.HexColor('#059669')
        if s >= 80:
            return colors.HexColor('#2563eb')
        if s >= 70:
            return colors.HexColor('#d97706')
        return colors.HexColor('#dc2626')

    def _severity_color_hex(self, severity):
        return {
            'CRITICAL': '#7f1d1d',
            'HIGH': '#dc2626',
            'MEDIUM': '#f59e0b',
            'LOW': '#64748b',
            'INFO': '#64748b',
        }.get((severity or 'LOW').upper(), '#64748b')

    def _create_styles(self):
        styles = getSampleStyleSheet()
        primary_hex = self.theme.get('primary', '#1e40af')

        styles.add(ParagraphStyle(
            name='CustomTitle', parent=styles['Heading1'], fontSize=22, leading=26,
            textColor=colors.HexColor(primary_hex), spaceAfter=8, alignment=TA_CENTER, fontName='Helvetica-Bold'
        ))
        styles.add(ParagraphStyle(
            name='CustomSubtitle', parent=styles['Normal'], fontSize=12, leading=15,
            textColor=colors.HexColor('#64748b'), alignment=TA_CENTER, spaceAfter=12
        ))
        styles.add(ParagraphStyle(
            name='SectionHeader', parent=styles['Heading2'], fontSize=14, leading=18,
            textColor=colors.HexColor(primary_hex), spaceBefore=12, spaceAfter=8, fontName='Helvetica-Bold'
        ))
        styles.add(ParagraphStyle(
            name='ScoreBig', parent=styles['Normal'], fontSize=48, leading=54,
            textColor=colors.HexColor(primary_hex), alignment=TA_CENTER, fontName='Helvetica-Bold', spaceBefore=6, spaceAfter=2
        ))
        styles.add(ParagraphStyle(
            name='ScoreSub', parent=styles['Normal'], fontSize=11, leading=14,
            textColor=colors.HexColor('#64748b'), alignment=TA_CENTER, spaceBefore=0, spaceAfter=12
        ))
        styles.add(ParagraphStyle(
            name='RecommendationText', parent=styles['Normal'], fontSize=9,
            textColor=colors.HexColor('#92400e'), leading=12, spaceBefore=0, spaceAfter=0
        ))
        styles.add(ParagraphStyle(
            name='CheckBody', parent=styles['Normal'], fontSize=9,
            textColor=colors.HexColor('#334155'), leading=13, spaceBefore=2, spaceAfter=2, alignment=TA_LEFT
        ))
        styles.add(ParagraphStyle(
            name='CheckTitle', parent=styles['Normal'], fontSize=10,
            textColor=colors.HexColor('#0f172a'), leading=14, spaceBefore=4, spaceAfter=2, fontName='Helvetica-Bold'
        ))
        return styles

    def _recommendation_box(self, text):
        if not text:
            return Spacer(1, 0.05 * cm)

        clean = self._safe(str(text).replace('\\n', ' ').strip())
        if len(clean) > 800:
            clean = clean[:800] + '…'

        para = Paragraph(f"💡 {clean}", self.styles['RecommendationText'])
        table = Table([[para]], colWidths=[16 * cm])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#FEF3C7')),
            ('BOX', (0, 0), (-1, -1), 0.5, colors.HexColor('#F59E0B')),
            ('LEFTPADDING', (0, 0), (-1, -1), 8),
            ('RIGHTPADDING', (0, 0), (-1, -1), 8),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ]))
        return table

    def _add_page_decor(self, canvas, doc):
        canvas.saveState()
        canvas.setFont('Helvetica', 8)
        canvas.setFillColor(colors.HexColor('#64748b'))
        page_w, _ = A4
        fname = (self.file_name or APP_TITLE)[:50]
        canvas.drawString(1.5 * cm, 0.6 * cm, fname)
        canvas.drawRightString(page_w - 1.5 * cm, 0.6 * cm, f"Page {doc.page}")
        canvas.drawCentredString(page_w / 2, 0.6 * cm, f"{COMPANY_NAME} | Confidential")
        canvas.restoreState()

    def _build_doc(self, buffer):
        return SimpleDocTemplate(
            buffer, pagesize=A4, topMargin=1.2 * cm, bottomMargin=1.4 * cm,
            leftMargin=1.5 * cm, rightMargin=1.5 * cm,
            title=f"{APP_TITLE} Health Report", author=COMPANY_NAME
        )

    def generate_executive_report(self):
        buffer = io.BytesIO()
        doc = self._build_doc(buffer)
        story = []

        selected_std = self.data.get('selected_standard', 'all')
        score = self.data.get('overall_score', 0)

        story.append(Spacer(1, 0.3 * cm))
        story.append(Paragraph("SCHEDULE HEALTH", self.styles['CustomTitle']))
        story.append(Paragraph("Executive Assessment Report", self.styles['CustomSubtitle']))
        story.append(Spacer(1, 0.2 * cm))

        meta_lines = [
            f"<b>Project File:</b> {self._safe(self.file_name)}",
            f"<b>Analysis Date:</b> {self._safe(self.data.get('analysis_date', datetime.now().strftime('%Y-%m-%d %H:%M')))}",
            f"<b>Standard Scope:</b> {self._safe(str(selected_std).upper())}",
            f"<b>Severity Filter (action lists):</b> {self._safe(self.severity_filter.upper())}",
        ]
        proj = self.data.get('project_info') or {}
        if proj.get('name'):
            meta_lines.append(f"<b>Project Name:</b> {self._safe(proj.get('name'))}")
        if proj.get('data_date'):
            meta_lines.append(f"<b>Data Date:</b> {self._safe(proj.get('data_date'))}")

        for line in meta_lines:
            story.append(self._p_markup(line, 'CheckBody'))

        story.append(Spacer(1, 0.35 * cm))
        story.append(Paragraph("OVERALL HEALTH SCORE", self.styles['SectionHeader']))
        score_style = ParagraphStyle(
            'ScoreBigDynamic', parent=self.styles['ScoreBig'], textColor=self._score_color(score)
        )
        story.append(Paragraph(self._safe(str(score)), score_style))
        story.append(Paragraph("out of 100 (weighted)", self.styles['ScoreSub']))

        stats_data = [
            [Paragraph('<b>Metric</b>', self.styles['CheckBody']), Paragraph('<b>Value</b>', self.styles['CheckBody'])],
            ['Total Checks Performed (full run)', str(self.data.get('total_checks', 0))],
            ['Checks Passed', f"{self.data.get('passed_checks', 0)} ({self.data.get('pass_rate', 0)}%)"],
            ['Checks Failed', str(self.data.get('failed_checks', 0))],
            ['Critical Failures (full run)', str(self.data.get('critical_failures', 0))],
            ['High-Severity Failures (full run)', str(self.data.get('high_failures', 0))],
            ['Action list severity filter', self.severity_filter.upper()],
        ]
        stats_table_data = []
        for i, row in enumerate(stats_data):
            if i == 0:
                stats_table_data.append(row)
            else:
                stats_table_data.append([
                    Paragraph(self._safe(row[0]), self.styles['CheckBody']),
                    Paragraph(self._safe(str(row[1])), self.styles['CheckBody']),
                ])

        stats_table = Table(stats_table_data, colWidths=[10 * cm, 6 * cm])
        stats_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), self.primary),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#cbd5e1')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.HexColor('#f8fafc'), colors.white]),
        ]))
        story.append(stats_table)

        story.append(PageBreak())

        story.append(Paragraph("STANDARDS COMPLIANCE SUMMARY", self.styles['SectionHeader']))
        story.append(HRFlowable(width="100%", thickness=2, color=self.primary))
        story.append(Spacer(1, 0.35 * cm))

        std_header = ['Standard', 'Score', 'Grade', 'Passed', 'Failed', 'Critical']
        std_rows = [std_header]
        for std_name, std_score in (self.data.get('standard_scores') or {}).items():
            std_rows.append([
                str(std_name), str(std_score.get('score', 0)), str(std_score.get('grade', '-')),
                str(std_score.get('passed', 0)), str(std_score.get('failed', 0)), str(std_score.get('critical_failures', 0)),
            ])

        if len(std_rows) == 1:
            std_rows.append(['—', '—', '—', '—', '—', '—'])

        std_table_data = []
        for i, row in enumerate(std_rows):
            std_table_data.append([
                Paragraph(f"<b>{self._safe(c)}</b>" if i == 0 else self._safe(c), self.styles['CheckBody'])
                for c in row
            ])

        std_table = Table(std_table_data, colWidths=[3.5 * cm, 2.2 * cm, 2.2 * cm, 2.5 * cm, 2.5 * cm, 2.5 * cm])
        std_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), self.primary),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (1, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
            ('TOPPADDING', (0, 0), (-1, -1), 5),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#cbd5e1')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.HexColor('#f8fafc'), colors.white]),
        ]))
        story.append(std_table)
        story.append(Spacer(1, 0.6 * cm))

        story.append(Paragraph(f"TOP PRIORITY ACTIONS (Severity: {self.severity_filter.upper()})", self.styles['SectionHeader']))
        story.append(HRFlowable(width="100%", thickness=2, color=self.danger))
        story.append(Spacer(1, 0.25 * cm))

        filtered_actions = self._build_priority_actions(limit=min(10, self.max_top))

        if not filtered_actions:
            story.append(self._p_markup(f"<i>No actions found matching severity filter: {self._safe(self.severity_filter.upper())}</i>", 'CheckBody'))
        else:
            for idx, action in enumerate(filtered_actions, 1):
                block = []
                severity = (action.get('severity') or 'low').upper()
                sev_hex = self._severity_color_hex(severity)

                title = (
                    f"<b>{idx}. [{self._safe(action.get('id', ''))}] {self._safe(action.get('name', ''))}</b><br/>"
                    f"<font size='9' color='#64748b'>Standard: {self._safe(action.get('standard', ''))} | Severity: <font color='{sev_hex}'><b>{self._safe(severity)}</b></font> | Affected: {self._safe(action.get('count', 0))} ({self._safe(action.get('percentage', 0))}%)</font>"
                )
                block.append(self._p_markup(title, 'CheckBody'))

                if action.get('recommendation'):
                    block.append(Spacer(1, 0.08 * cm))
                    block.append(self._recommendation_box(action.get('recommendation')))

                items = action.get('failed_items') or []
                if items:
                    block.append(self._p_markup("<b>Affected Activities:</b>", 'CheckBody'))
                    for item in items[:10]:
                        block.append(self._p_markup(self._activity_line(item), 'CheckBody'))
                    if len(items) > 10:
                        block.append(self._p_markup(f"<i>… and {len(items) - 10} more</i>", 'CheckBody'))

                block.append(Spacer(1, 0.25 * cm))
                story.append(KeepTogether(block))

        doc.build(story, onFirstPage=self._add_page_decor, onLaterPages=self._add_page_decor)
        buffer.seek(0)
        return buffer

    def generate_actions_report(self):
        buffer = io.BytesIO()
        doc = self._build_doc(buffer)
        story = []

        selected_std = self.data.get('selected_standard', 'all')

        story.append(Paragraph("SCHEDULE HEALTH ACTIONS", self.styles['CustomTitle']))
        story.append(Paragraph("Failed Checks & Corrective Action List", self.styles['CustomSubtitle']))
        story.append(self._p_markup(f"<i>Generated: {self._safe(datetime.now().strftime('%Y-%m-%d %H:%M'))} | File: {self._safe(self.file_name)}</i>", 'CheckBody'))
        story.append(self._p_markup(f"<b>Standard Scope:</b> {self._safe(str(selected_std).upper())} | <b>Severity Filter:</b> {self._safe(self.severity_filter.upper())}", 'CheckBody'))
        story.append(HRFlowable(width="100%", thickness=2, color=self.danger))
        story.append(Spacer(1, 0.4 * cm))

        summary_data = [
            ['Total Failed Checks (full run)', str(self.data.get('failed_checks', 0))],
            ['Critical Failures (full run)', str(self.data.get('critical_failures', 0))],
            ['High Severity Failures (full run)', str(self.data.get('high_failures', 0))],
            ['Overall Health Score', f"{self.data.get('overall_score', 0)} / 100"],
            ['This report severity filter', self.severity_filter.upper()],
        ]
        summary_table_data = [
            [Paragraph(self._safe(a), self.styles['CheckBody']), Paragraph(self._safe(b), self.styles['CheckBody'])]
            for a, b in summary_data
        ]
        summary_table = Table(summary_table_data, colWidths=[10 * cm, 6 * cm])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#FEF3C7')),
            ('BACKGROUND', (1, 0), (1, -1), colors.white),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#cbd5e1')),
        ]))
        story.append(summary_table)
        story.append(Spacer(1, 0.55 * cm))

        story.append(Paragraph(f"TOP PRIORITY ACTIONS (Severity: {self.severity_filter.upper()})", self.styles['SectionHeader']))
        story.append(HRFlowable(width="100%", thickness=2, color=self.danger))
        story.append(Spacer(1, 0.25 * cm))

        filtered_top = self._build_priority_actions(limit=self.max_top)

        if not filtered_top:
            story.append(self._p_markup(f"<i>No priority actions matched severity filter: {self._safe(self.severity_filter.upper())}</i>", 'CheckBody'))
        else:
            for idx, action in enumerate(filtered_top, 1):
                block = self._render_action_block(idx, action)
                story.append(KeepTogether(block))

        story.append(PageBreak())

        story.append(Paragraph(f"DETAILED FAILED CHECKS (Severity: {self.severity_filter.upper()})", self.styles['SectionHeader']))
        story.append(Spacer(1, 0.2 * cm))

        any_failures = False
        by_std = {}
        for std_name, cat_name, check in self._iter_failed_checks():
            by_std.setdefault(std_name, []).append((cat_name, check))

        for std_name, items in by_std.items():
            any_failures = True
            story.append(Spacer(1, 0.3 * cm))
            story.append(self._p_markup(f"<b>{self._safe(std_name)} Standard</b> ({len(items)} filtered failures)", 'SectionHeader'))

            for cat_name, check in items:
                block = []
                severity = (check.get('severity') or 'low').upper()
                sev_hex = self._severity_color_hex(severity)

                title = f'<font color="{sev_hex}">[{self._safe(severity)}]</font> {self._safe(check.get("id", ""))}: {self._safe(check.get("name", ""))}'
                block.append(self._p_markup(title, 'CheckTitle'))
                block.append(self._p_markup(f"Category: {self._safe(cat_name)}", 'CheckBody'))

                if check.get('description'):
                    block.append(self._p_markup(self._safe(check.get('description')), 'CheckBody'))

                if check.get('count') is not None:
                    metric_line = f"Affected: <b>{self._safe(check.get('count'))}</b> of {self._safe(check.get('total'))} ({self._safe(check.get('percentage', 0))}%) | Threshold: {self._safe(check.get('threshold', ''))}"
                    block.append(self._p_markup(metric_line, 'CheckBody'))
                elif check.get('value') is not None:
                    metric_line = f"Value: <b>{self._safe(check.get('value'))}</b> | Threshold: {self._safe(check.get('threshold', ''))}"
                    block.append(self._p_markup(metric_line, 'CheckBody'))

                if check.get('recommendation'):
                    block.append(Spacer(1, 0.1 * cm))
                    block.append(self._recommendation_box(check.get('recommendation')))

                failed_items = check.get('failed_items') or []
                if failed_items:
                    block.append(Spacer(1, 0.08 * cm))
                    block.append(self._p_markup("<b>Affected Activities:</b>", 'CheckBody'))
                    shown = failed_items[: self.max_items]
                    for item in shown:
                        block.append(self._p_markup(self._activity_line(item), 'CheckBody'))
                    leftover = len(failed_items) - len(shown)
                    if leftover > 0:
                        block.append(self._p_markup(f"<i>… and {leftover} more (see Excel export)</i>", 'CheckBody'))
                else:
                    block.append(self._p_markup("<i>No activity list available for this metric.</i>", 'CheckBody'))

                block.append(Spacer(1, 0.3 * cm))
                story.append(KeepTogether(block))

        if not any_failures:
            story.append(self._p_markup(f"<i>No failures matched severity filter: {self._safe(self.severity_filter.upper())}</i>", 'CheckBody'))

        doc.build(story, onFirstPage=self._add_page_decor, onLaterPages=self._add_page_decor)
        buffer.seek(0)
        return buffer

    def _render_action_block(self, idx, action):
        block = []
        severity = (action.get('severity') or 'low').upper()
        sev_hex = self._severity_color_hex(severity)

        title = f"<b>{idx}. <font color='{sev_hex}'>[{self._safe(severity)}]</font> {self._safe(action.get('id', ''))}: {self._safe(action.get('name', ''))}</b>"
        block.append(self._p_markup(title, 'CheckTitle'))

        meta_parts = []
        if action.get('standard'):
            meta_parts.append(f"Standard: {self._safe(action.get('standard'))}")
        if action.get('category'):
            meta_parts.append(f"Category: {self._safe(action.get('category'))}")
        if action.get('count') is not None:
            meta_parts.append(f"Affected: {self._safe(action.get('count', 0))} ({self._safe(action.get('percentage', 0))}%)")
        elif action.get('value') is not None:
            meta_parts.append(f"Value: {self._safe(action.get('value'))}")
        if action.get('threshold'):
            meta_parts.append(f"Threshold: {self._safe(action.get('threshold'))}")

        if meta_parts:
            block.append(self._p_markup(" | ".join(meta_parts), 'CheckBody'))

        if action.get('description'):
            block.append(self._p_markup(self._safe(action.get('description')), 'CheckBody'))

        if action.get('recommendation'):
            block.append(Spacer(1, 0.08 * cm))
            block.append(self._recommendation_box(action.get('recommendation')))

        failed_items = action.get('failed_items') or []
        if failed_items:
            block.append(Spacer(1, 0.1 * cm))
            block.append(self._p_markup("<b>Affected Activities:</b>", 'CheckBody'))
            shown = failed_items[: self.max_items]
            for item in shown:
                block.append(self._p_markup(self._activity_line(item), 'CheckBody'))
            leftover = len(failed_items) - len(shown)
            if leftover > 0:
                block.append(self._p_markup(f"<i>… and {leftover} more (see Excel export)</i>", 'CheckBody'))
        else:
            block.append(self._p_markup("<i>No activity list available for this metric.</i>", 'CheckBody'))

        block.append(Spacer(1, 0.3 * cm))
        return block
'''

with open("pdf_report_generator.py", "w", encoding="utf-8") as f:
    f.write(PDF_REPORT_GENERATOR_CODE)
print("  ✅ Updated pdf_report_generator.py")


# ------------------------------------------------------------------------------
# 8. app.py
# ------------------------------------------------------------------------------
APP_CODE = '''"""
P6 SCHEDULE ANALYZER - MAIN WEB APPLICATION (app.py)
=====================================================
"""

from flask import Flask, render_template, request, jsonify, send_file, session
from flask_cors import CORS
from werkzeug.utils import secure_filename
import os
import uuid
import time
import logging
from datetime import datetime

try:
    from config import (
        get_config,
        MAX_UPLOAD_SIZE_MB,
        SECRET_KEY,
        SESSION_LIFETIME_HOURS,
    )
except ImportError:
    MAX_UPLOAD_SIZE_MB = 100
    SECRET_KEY = 'dev-only-CHANGE-ME'
    SESSION_LIFETIME_HOURS = 24
    def get_config():
        return {
            'company_name': 'MK Constructions',
            'app_title': 'P6 Schedule Analyzer',
            'app_subtitle': 'DCMA 14-Point Check & Analytics',
            'use_logo_image': False,
            'theme': {'primary': '#1e40af', 'accent': '#3b82f6'},
            'features': {'gantt': True, 'comparison': True, 'evm': True, 'export': True, 'health': True}
        }

from parser import XERParser
from data_engine import ScheduleEngine
from reports import ReportGenerator

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

try:
    from comparison_engine import ScheduleComparator
    logger.info("✅ ScheduleComparator imported")
except Exception as e:
    ScheduleComparator = None
    logger.warning("❌ ScheduleComparator import failed: %s", e)

try:
    from evm_engine import EVMEngine
    logger.info("✅ EVMEngine imported")
except Exception as e:
    EVMEngine = None
    logger.warning("❌ EVMEngine import failed: %s", e)

try:
    from advanced_health_engine import AdvancedHealthEngine
    logger.info("✅ AdvancedHealthEngine imported")
except Exception as e:
    AdvancedHealthEngine = None
    logger.warning("❌ AdvancedHealthEngine import failed: %s", e)

try:
    from pdf_report_generator import PDFReportGenerator
    logger.info("✅ PDFReportGenerator imported")
except Exception as e:
    PDFReportGenerator = None
    logger.warning("❌ PDFReportGenerator import failed: %s", e)


app = Flask(__name__)
app.secret_key = SECRET_KEY
CORS(app)

UPLOAD_FOLDER = 'uploads'
OUTPUT_FOLDER = 'output'
ALLOWED_EXTENSIONS = {'xer'}

app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['OUTPUT_FOLDER'] = OUTPUT_FOLDER
app.config['MAX_CONTENT_LENGTH'] = MAX_UPLOAD_SIZE_MB * 1024 * 1024

os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(OUTPUT_FOLDER, exist_ok=True)


def cleanup_old_files(folder, max_age_hours=SESSION_LIFETIME_HOURS):
    if not os.path.exists(folder):
        return
    cutoff = time.time() - (max_age_hours * 3600)
    for fname in os.listdir(folder):
        fpath = os.path.join(folder, fname)
        try:
            if os.path.isfile(fpath) and os.path.getmtime(fpath) < cutoff:
                os.remove(fpath)
                logger.info("🧹 Cleaned up old file: %s", fname)
        except Exception as e:
            logger.warning("Could not delete %s: %s", fname, e)

cleanup_old_files(UPLOAD_FOLDER)
cleanup_old_files(OUTPUT_FOLDER)


SESSION_STORAGE = {}

def get_session_data():
    sid = session.get('sid')
    if not sid or sid not in SESSION_STORAGE:
        sid = uuid.uuid4().hex
        session['sid'] = sid
        SESSION_STORAGE[sid] = {
            'analysis': {'engine': None, 'dashboard_data': None, 'file_name': None, 'analyzed_at': None},
            'comparison': {'comparator': None, 'results': None, 'baseline_file': None, 'current_file': None},
            'health_cache': {},
        }
    return SESSION_STORAGE[sid]


def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


def analyze_xer_file(file_path_or_stream, original_filename, session_data):
    logger.info("🔍 Analyzing XER: %s", original_filename)
    parser = XERParser()
    tables = parser.parse(file_path_or_stream)

    if tables is None or not tables:
        return {'error': 'Failed to parse XER file or file is empty.'}

    engine = ScheduleEngine()
    engine.load_data(tables)
    engine.analyze()

    dashboard_data = engine.get_dashboard_data()

    analysis = session_data['analysis']
    analysis['engine'] = engine
    analysis['dashboard_data'] = dashboard_data
    analysis['file_name'] = original_filename
    analysis['analyzed_at'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    session_data['health_cache'] = {}

    return dashboard_data


@app.context_processor
def inject_config():
    return {'config': get_config()}


@app.errorhandler(413)
def request_entity_too_large(error):
    return jsonify({'error': f'File size exceeds maximum limit of {MAX_UPLOAD_SIZE_MB} MB.'}), 413


@app.route('/')
def home():
    return render_template('index.html')


@app.route('/api/upload', methods=['POST'])
def upload_file():
    if 'file' not in request.files:
        return jsonify({'error': 'No file provided'}), 400

    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'No file selected'}), 400

    if not allowed_file(file.filename):
        return jsonify({'error': 'File must be a .xer file'}), 400

    original_name = secure_filename(file.filename)
    unique_name = f"{uuid.uuid4().hex}_{original_name}"
    file_path = os.path.join(app.config['UPLOAD_FOLDER'], unique_name)
    file.save(file_path)

    sess_data = get_session_data()

    try:
        dashboard_data = analyze_xer_file(file_path, original_name, sess_data)
        if 'error' in dashboard_data:
            return jsonify(dashboard_data), 400

        analysis = sess_data['analysis']
        return jsonify({
            'success': True,
            'file_name': analysis['file_name'],
            'analyzed_at': analysis['analyzed_at'],
            'data': dashboard_data
        })
    except Exception as e:
        logger.exception("Upload analysis error")
        return jsonify({'error': str(e)}), 500


@app.route('/api/dashboard')
def get_dashboard():
    sess_data = get_session_data()
    analysis = sess_data['analysis']

    if analysis['dashboard_data'] is None:
        return jsonify({'has_data': False})

    return jsonify({
        'has_data': True,
        'file_name': analysis['file_name'],
        'analyzed_at': analysis['analyzed_at'],
        'data': analysis['dashboard_data']
    })


@app.route('/api/load-sample')
def load_sample():
    sample_path = os.path.join('input', 'sample.xer')
    if not os.path.exists(sample_path):
        return jsonify({'error': 'sample.xer not found in input/ folder'}), 404

    sess_data = get_session_data()
    try:
        dashboard_data = analyze_xer_file(sample_path, 'sample.xer', sess_data)
        analysis = sess_data['analysis']
        return jsonify({
            'success': True,
            'file_name': 'sample.xer',
            'analyzed_at': analysis['analyzed_at'],
            'data': dashboard_data
        })
    except Exception as e:
        logger.exception("Sample load error")
        return jsonify({'error': str(e)}), 500


@app.route('/api/export-excel')
def export_excel():
    sess_data = get_session_data()
    engine = sess_data['analysis']['engine']

    if engine is None:
        return jsonify({'error': 'No schedule loaded. Please upload an XER file first.'}), 400

    out_name = f"schedule_report_{uuid.uuid4().hex[:8]}.xlsx"
    output_path = os.path.join(app.config['OUTPUT_FOLDER'], out_name)
    
    reporter = ReportGenerator(engine)
    res_path = reporter.generate_full_report(output_path)

    if not res_path or not os.path.exists(res_path):
        return jsonify({'error': 'Failed to generate Excel report.'}), 500

    return send_file(
        res_path,
        as_attachment=True,
        download_name=f"schedule_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


@app.route('/gantt')
def gantt_view():
    return render_template('gantt.html')


@app.route('/api/gantt-data')
def get_gantt_data():
    sess_data = get_session_data()
    analysis = sess_data['analysis']

    if analysis['engine'] is None:
        return jsonify({'error': 'No data loaded. Please upload an XER file first.'}), 400

    max_acts = request.args.get('max', 2000, type=int)
    gantt_data = analysis['engine'].get_gantt_data(max_activities=max_acts)

    return jsonify({
        'success': True,
        'file_name': analysis['file_name'],
        'data': gantt_data
    })


@app.route('/comparison')
def comparison_view():
    return render_template('comparison.html')


@app.route('/api/compare', methods=['POST'])
def compare_schedules():
    if ScheduleComparator is None:
        return jsonify({'error': 'comparison_engine.py is missing!'}), 500

    if 'baseline' not in request.files or 'current' not in request.files:
        return jsonify({'error': 'Both baseline and current files required'}), 400

    baseline_file = request.files['baseline']
    current_file = request.files['current']

    if not (allowed_file(baseline_file.filename) and allowed_file(current_file.filename)):
        return jsonify({'error': 'Both files must be .xer files'}), 400

    baseline_name = secure_filename(baseline_file.filename)
    current_name = secure_filename(current_file.filename)

    baseline_path = os.path.join(app.config['UPLOAD_FOLDER'], f"bl_{uuid.uuid4().hex[:8]}_{baseline_name}")
    current_path = os.path.join(app.config['UPLOAD_FOLDER'], f"cur_{uuid.uuid4().hex[:8]}_{current_name}")

    baseline_file.save(baseline_path)
    current_file.save(current_path)

    sess_data = get_session_data()

    try:
        comparator = ScheduleComparator()
        comparator.load_baseline(baseline_path)
        comparator.load_current(current_path)
        results = comparator.compare()

        comp = sess_data['comparison']
        comp['comparator'] = comparator
        comp['results'] = results
        comp['baseline_file'] = baseline_name
        comp['current_file'] = current_name

        return jsonify({
            'success': True,
            'baseline_file': baseline_name,
            'current_file': current_name,
            'results': results
        })
    except Exception as e:
        logger.exception("Comparison error")
        return jsonify({'error': str(e)}), 500


@app.route('/api/comparison-data')
def get_comparison_data():
    sess_data = get_session_data()
    comp = sess_data['comparison']

    if comp['results'] is None:
        return jsonify({'has_data': False})

    return jsonify({
        'has_data': True,
        'baseline_file': comp['baseline_file'],
        'current_file': comp['current_file'],
        'results': comp['results']
    })


@app.route('/evm')
def evm_view():
    return render_template('evm.html')


@app.route('/api/evm-data')
def get_evm_data():
    sess_data = get_session_data()
    analysis = sess_data['analysis']

    if analysis['engine'] is None:
        return jsonify({'error': 'No data loaded. Upload an XER file first.'}), 400

    if EVMEngine is None:
        return jsonify({'error': 'evm_engine.py is missing!'}), 500

    try:
        evm = EVMEngine(analysis['engine'])
        results = evm.calculate()
        return jsonify({
            'success': True,
            'file_name': analysis['file_name'],
            'data': results
        })
    except Exception as e:
        logger.exception("EVM calculation error")
        return jsonify({'error': str(e)}), 500


@app.route('/health')
def health_view():
    return render_template('health.html')


@app.route('/api/health-data')
def get_health_data():
    sess_data = get_session_data()
    analysis = sess_data['analysis']

    if analysis['engine'] is None:
        return jsonify({'error': 'No data loaded. Upload a file first.'}), 400

    if AdvancedHealthEngine is None:
        return jsonify({'error': 'advanced_health_engine.py is missing!'}), 500

    selected_standard = request.args.get('standard', 'all')

    try:
        health = AdvancedHealthEngine(analysis['engine'])
        results = health.run_all_checks(selected_standard=selected_standard)

        return jsonify({
            'success': True,
            'file_name': analysis['file_name'],
            'data': results
        })
    except Exception as e:
        logger.exception("Health analysis error")
        return jsonify({'error': str(e)}), 500


@app.route('/api/executive-pdf')
def download_executive_pdf():
    sess_data = get_session_data()
    analysis = sess_data['analysis']

    if analysis['engine'] is None:
        return jsonify({'error': 'No data loaded'}), 400

    if PDFReportGenerator is None or AdvancedHealthEngine is None:
        return jsonify({'error': 'PDF or Health engine module missing!'}), 500

    selected_standard = request.args.get('standard', 'all')
    severity_filter = request.args.get('severity', 'all')

    try:
        health = AdvancedHealthEngine(analysis['engine'])
        results = health.run_all_checks(selected_standard=selected_standard)

        generator = PDFReportGenerator(
            results,
            analysis['file_name'],
            severity_filter=severity_filter
        )
        pdf_buffer = generator.generate_executive_report()

        return send_file(
            pdf_buffer,
            as_attachment=True,
            download_name=f"executive_report_{selected_standard}_{severity_filter}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf",
            mimetype='application/pdf'
        )
    except Exception as e:
        logger.exception("PDF generation error")
        return jsonify({'error': str(e)}), 500


@app.route('/api/actions-pdf')
def download_actions_pdf():
    sess_data = get_session_data()
    analysis = sess_data['analysis']

    if analysis['engine'] is None:
        return jsonify({'error': 'No data loaded'}), 400

    if PDFReportGenerator is None or AdvancedHealthEngine is None:
        return jsonify({'error': 'PDF or Health engine module missing!'}), 500

    selected_standard = request.args.get('standard', 'all')
    severity_filter = request.args.get('severity', 'all')

    try:
        health = AdvancedHealthEngine(analysis['engine'])
        results = health.run_all_checks(selected_standard=selected_standard)

        generator = PDFReportGenerator(
            results,
            analysis['file_name'],
            severity_filter=severity_filter
        )
        pdf_buffer = generator.generate_actions_report()

        return send_file(
            pdf_buffer,
            as_attachment=True,
            download_name=f"action_list_{selected_standard}_{severity_filter}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf",
            mimetype='application/pdf'
        )
    except Exception as e:
        logger.exception("PDF actions error")
        return jsonify({'error': str(e)}), 500


@app.route('/api/actions-excel')
def download_actions_excel():
    sess_data = get_session_data()
    analysis = sess_data['analysis']

    if analysis['engine'] is None:
        return jsonify({'error': 'No data loaded. Upload a file first.'}), 400

    if AdvancedHealthEngine is None:
        return jsonify({'error': 'advanced_health_engine.py is missing!'}), 500

    selected_standard = request.args.get('standard', 'all')
    severity_filter = request.args.get('severity', 'all').lower()

    severity_levels = {
        'critical': ['critical'],
        'high': ['critical', 'high'],
        'medium': ['critical', 'high', 'medium'],
        'all': ['critical', 'high', 'medium', 'low', 'info']
    }
    allowed_severities = severity_levels.get(severity_filter, severity_levels['all'])

    try:
        import pandas as pd
        from io import BytesIO
        from openpyxl.styles import Font, PatternFill, Alignment

        health = AdvancedHealthEngine(analysis['engine'])
        results = health.run_all_checks(selected_standard=selected_standard)

        top_actions = results.get('top_actions', []) or []
        standards_data = results.get('standards', {}) or {}

        filtered_top_actions = [
            a for a in top_actions
            if (a.get('severity') or 'low').lower() in allowed_severities
        ]

        meta_rows = [
            ['Report Type', 'Schedule Health - Top Actions Export'],
            ['Selected Standard', selected_standard],
            ['Severity Filter', severity_filter.upper()],
            ['File Name', analysis.get('file_name', '')],
            ['Generated At', results.get('analysis_date', '')],
            ['', ''],
            ['Overall Score', results.get('overall_score', '')],
            ['Total Checks', results.get('total_checks', '')],
            ['Passed Checks', results.get('passed_checks', '')],
            ['Failed Checks', results.get('failed_checks', '')],
            ['Critical Failures', results.get('critical_failures', '')],
            ['High Failures', results.get('high_failures', '')],
            ['Pass Rate (%)', results.get('pass_rate', '')],
            ['', ''],
            ['Filtered Actions Count', len(filtered_top_actions)],
        ]

        top_summary_rows = []
        for idx, action in enumerate(filtered_top_actions, 1):
            top_summary_rows.append({
                'Rank': idx,
                'Standard': action.get('standard', ''),
                'Check ID': action.get('id', ''),
                'Check Name': action.get('name', ''),
                'Category': action.get('category', ''),
                'Severity': (action.get('severity') or '').upper(),
                'Affected Count': action.get('count', 0),
                'Total': action.get('total', 0),
                'Percentage': action.get('percentage', 0),
                'Threshold': action.get('threshold', ''),
                'Value': action.get('value', ''),
                'Recommendation': action.get('recommendation', ''),
                'Description': action.get('description', ''),
            })

        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            pd.DataFrame(meta_rows, columns=['Field', 'Value']).to_excel(
                writer, sheet_name='Report Info', index=False
            )

            if top_summary_rows:
                pd.DataFrame(top_summary_rows).to_excel(
                    writer, sheet_name='Top Actions Summary', index=False
                )
            else:
                pd.DataFrame([{'Info': f'No actions matched severity filter: {severity_filter.upper()}'}]).to_excel(
                    writer, sheet_name='Top Actions Summary', index=False
                )

            for std_name, std_data in standards_data.items():
                sheet_name = (std_name or 'Standard')[:31]
                rows = []

                failed_in_std = []
                for category in std_data.get('categories', []):
                    for check in category.get('checks', []):
                        if check.get('status') != 'fail':
                            continue
                        if (check.get('severity') or 'low').lower() not in allowed_severities:
                            continue
                        failed_in_std.append((category.get('name', ''), check))

                if not failed_in_std:
                    rows.append({
                        'Section': 'No Matching Failures',
                        'Field': '',
                        'Value': f'No {severity_filter.upper()} failures found for {std_name}',
                        'Activity ID': '',
                        'Activity Name': '',
                        'WBS': '',
                    })
                else:
                    for cat_name, check in failed_in_std:
                        rows.append({
                            'Section': 'METRIC',
                            'Field': 'Check ID',
                            'Value': check.get('id', ''),
                            'Activity ID': '', 'Activity Name': '', 'WBS': '',
                        })
                        rows.append({
                            'Section': '', 'Field': 'Check Name', 'Value': check.get('name', ''),
                            'Activity ID': '', 'Activity Name': '', 'WBS': '',
                        })
                        rows.append({
                            'Section': '', 'Field': 'Category', 'Value': cat_name,
                            'Activity ID': '', 'Activity Name': '', 'WBS': '',
                        })
                        rows.append({
                            'Section': '', 'Field': 'Severity', 'Value': (check.get('severity') or '').upper(),
                            'Activity ID': '', 'Activity Name': '', 'WBS': '',
                        })
                        rows.append({
                            'Section': '', 'Field': 'Description', 'Value': check.get('description', ''),
                            'Activity ID': '', 'Activity Name': '', 'WBS': '',
                        })
                        rows.append({
                            'Section': '', 'Field': 'Threshold', 'Value': check.get('threshold', ''),
                            'Activity ID': '', 'Activity Name': '', 'WBS': '',
                        })

                        if check.get('count') is not None:
                            rows.append({
                                'Section': '', 'Field': 'Affected',
                                'Value': f"{check.get('count', 0)} of {check.get('total', 0)} ({check.get('percentage', 0)}%)",
                                'Activity ID': '', 'Activity Name': '', 'WBS': '',
                            })
                        if check.get('value') is not None and check.get('value') != '':
                            rows.append({
                                'Section': '', 'Field': 'Value', 'Value': check.get('value', ''),
                                'Activity ID': '', 'Activity Name': '', 'WBS': '',
                            })

                        rows.append({
                            'Section': '', 'Field': 'Recommendation', 'Value': check.get('recommendation', ''),
                            'Activity ID': '', 'Activity Name': '', 'WBS': '',
                        })

                        items = check.get('failed_items', []) or []
                        rows.append({
                            'Section': 'AFFECTED ACTIVITIES', 'Field': f'Total: {len(items)}', 'Value': '',
                            'Activity ID': 'Activity ID', 'Activity Name': 'Activity Name', 'WBS': 'WBS',
                        })

                        if items:
                            for item in items:
                                rows.append({
                                    'Section': '', 'Field': '', 'Value': '',
                                    'Activity ID': item.get('code', ''),
                                    'Activity Name': item.get('name', ''),
                                    'WBS': item.get('wbs', ''),
                                })
                        else:
                            rows.append({
                                'Section': '', 'Field': '', 'Value': '(No activity list available)',
                                'Activity ID': '', 'Activity Name': '', 'WBS': '',
                            })

                        rows.append({'Section': '', 'Field': '', 'Value': '', 'Activity ID': '', 'Activity Name': '', 'WBS': ''})
                        rows.append({'Section': '', 'Field': '', 'Value': '', 'Activity ID': '', 'Activity Name': '', 'WBS': ''})

                pd.DataFrame(rows).to_excel(writer, sheet_name=sheet_name, index=False)

            workbook = writer.book
            header_font = Font(bold=True, color='FFFFFF', size=11)
            header_fill = PatternFill(start_color='1E40AF', end_color='1E40AF', fill_type='solid')
            metric_fill = PatternFill(start_color='FEF3C7', end_color='FEF3C7', fill_type='solid')
            activity_hdr_fill = PatternFill(start_color='DBEAFE', end_color='DBEAFE', fill_type='solid')
            bold_font = Font(bold=True)
            wrap_align = Alignment(wrap_text=True, vertical='top')

            for sheet_name in workbook.sheetnames:
                ws = workbook[sheet_name]
                for cell in ws[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal='left', vertical='center')

                for col in ws.columns:
                    max_len = 10
                    col_letter = col[0].column_letter
                    for cell in col:
                        try:
                            val = str(cell.value) if cell.value is not None else ''
                            if len(val) > max_len:
                                max_len = len(val)
                        except Exception:
                            pass
                    ws.column_dimensions[col_letter].width = min(max_len + 2, 60)

                ws.freeze_panes = 'A2'

                if sheet_name not in ['Report Info', 'Top Actions Summary']:
                    for row in ws.iter_rows(min_row=2):
                        section_cell = row[0]
                        if section_cell.value == 'METRIC':
                            for c in row:
                                c.fill = metric_fill
                                c.font = bold_font
                        elif section_cell.value == 'AFFECTED ACTIVITIES':
                            for c in row:
                                c.fill = activity_hdr_fill
                                c.font = bold_font
                        for c in row:
                            c.alignment = wrap_align

        output.seek(0)
        filename = f"health_top_actions_{selected_standard}_{severity_filter}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    except Exception as e:
        logger.exception("Actions Excel error")
        return jsonify({'error': str(e)}), 500


if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    debug_mode = os.environ.get('FLASK_ENV') != 'production'

    logger.info("============================================================")
    logger.info("🚀 P6 SCHEDULE ANALYZER - ALL FEATURES READY")
    logger.info("============================================================")
    logger.info("📌 Running on port %s", port)
    logger.info("🔧 Debug mode: %s", debug_mode)
    logger.info("👉 Open in browser: http://localhost:%s", port)

    app.run(debug=debug_mode, host='0.0.0.0', port=port)
'''

with open("app.py", "w", encoding="utf-8") as f:
    f.write(APP_CODE)
print("  ✅ Updated app.py")

print("\n🎉 Part 3 Patches Applied Successfully via Python!")
print("✨ ALL BACKEND ENGINES AND HEALTH STANDARDS ARE FULLY PATCHED!")