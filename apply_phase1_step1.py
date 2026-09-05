import os
import shutil
from datetime import datetime

print("🚀 Applying Phase 1 - Step 1: AI Executive Narrative Feature...")

# 1. Create Backup Folder
timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
backup_dir = f"_backup_phase1_step1_{timestamp}"
os.makedirs(backup_dir, exist_ok=True)
print(f"📦 Created backup folder: {backup_dir}")

files_to_backup = [
    "app.py",
    "templates/health.html",
    "static/health.js",
]

for file_path in files_to_backup:
    if os.path.exists(file_path):
        dest = os.path.join(backup_dir, os.path.basename(file_path.replace("/", os.sep)))
        shutil.copy2(file_path, dest)
        print(f"   Backed up {file_path}")


# ==============================================================================
# FILE 2: app.py (Full Rewrite with AI Narrative Route)
# ==============================================================================

APP_CODE = '''"""
P6 SCHEDULE ANALYZER - MAIN WEB APPLICATION (app.py)
=====================================================
Now with AI Executive Narrative Endpoint (Phase 1, Step 1)
"""

from flask import Flask, render_template, request, jsonify, send_file, session
from flask_cors import CORS
from werkzeug.utils import secure_filename
import os
import uuid
import time
import logging
from datetime import datetime

try:
    from config import (
        get_config,
        MAX_UPLOAD_SIZE_MB,
        SECRET_KEY,
        SESSION_LIFETIME_HOURS,
    )
except ImportError:
    MAX_UPLOAD_SIZE_MB = 100
    SECRET_KEY = 'dev-only-CHANGE-ME'
    SESSION_LIFETIME_HOURS = 24
    def get_config():
        return {
            'company_name': 'MK Constructions',
            'app_title': 'P6 Schedule Analyzer',
            'app_subtitle': 'DCMA 14-Point Check & Analytics',
            'use_logo_image': False,
            'theme': {'primary': '#1e40af', 'accent': '#3b82f6'},
            'features': {'gantt': True, 'comparison': True, 'evm': True, 'export': True, 'health': True}
        }

from parser import XERParser
from data_engine import ScheduleEngine
from reports import ReportGenerator

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

try:
    from comparison_engine import ScheduleComparator
    logger.info("✅ ScheduleComparator imported")
except Exception as e:
    ScheduleComparator = None
    logger.warning("❌ ScheduleComparator import failed: %s", e)

try:
    from evm_engine import EVMEngine
    logger.info("✅ EVMEngine imported")
except Exception as e:
    EVMEngine = None
    logger.warning("❌ EVMEngine import failed: %s", e)

try:
    from advanced_health_engine import AdvancedHealthEngine
    logger.info("✅ AdvancedHealthEngine imported")
except Exception as e:
    AdvancedHealthEngine = None
    logger.warning("❌ AdvancedHealthEngine import failed: %s", e)

try:
    from pdf_report_generator import PDFReportGenerator
    logger.info("✅ PDFReportGenerator imported")
except Exception as e:
    PDFReportGenerator = None
    logger.warning("❌ PDFReportGenerator import failed: %s", e)

try:
    from ai_narrative_engine import AINarrativeEngine
    logger.info("✅ AINarrativeEngine imported")
except Exception as e:
    AINarrativeEngine = None
    logger.warning("❌ AINarrativeEngine import failed: %s", e)


app = Flask(__name__)
app.secret_key = SECRET_KEY
CORS(app)

UPLOAD_FOLDER = 'uploads'
OUTPUT_FOLDER = 'output'
ALLOWED_EXTENSIONS = {'xer'}

app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['OUTPUT_FOLDER'] = OUTPUT_FOLDER
app.config['MAX_CONTENT_LENGTH'] = MAX_UPLOAD_SIZE_MB * 1024 * 1024

os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(OUTPUT_FOLDER, exist_ok=True)


def cleanup_old_files(folder, max_age_hours=SESSION_LIFETIME_HOURS):
    if not os.path.exists(folder):
        return
    cutoff = time.time() - (max_age_hours * 3600)
    for fname in os.listdir(folder):
        fpath = os.path.join(folder, fname)
        try:
            if os.path.isfile(fpath) and os.path.getmtime(fpath) < cutoff:
                os.remove(fpath)
                logger.info("🧹 Cleaned up old file: %s", fname)
        except Exception as e:
            logger.warning("Could not delete %s: %s", fname, e)

cleanup_old_files(UPLOAD_FOLDER)
cleanup_old_files(OUTPUT_FOLDER)


SESSION_STORAGE = {}

def get_session_data():
    sid = session.get('sid')
    if not sid or sid not in SESSION_STORAGE:
        sid = uuid.uuid4().hex
        session['sid'] = sid
        SESSION_STORAGE[sid] = {
            'analysis': {'engine': None, 'dashboard_data': None, 'file_name': None, 'analyzed_at': None},
            'comparison': {'comparator': None, 'results': None, 'baseline_file': None, 'current_file': None},
            'health_cache': {},
        }
    return SESSION_STORAGE[sid]


def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


def analyze_xer_file(file_path_or_stream, original_filename, session_data):
    logger.info("🔍 Analyzing XER: %s", original_filename)
    parser = XERParser()
    tables = parser.parse(file_path_or_stream)

    if tables is None or not tables:
        return {'error': 'Failed to parse XER file or file is empty.'}

    engine = ScheduleEngine()
    engine.load_data(tables)
    engine.analyze()

    dashboard_data = engine.get_dashboard_data()

    analysis = session_data['analysis']
    analysis['engine'] = engine
    analysis['dashboard_data'] = dashboard_data
    analysis['file_name'] = original_filename
    analysis['analyzed_at'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    session_data['health_cache'] = {}

    return dashboard_data


@app.context_processor
def inject_config():
    return {'config': get_config()}


@app.errorhandler(413)
def request_entity_too_large(error):
    return jsonify({'error': f'File size exceeds maximum limit of {MAX_UPLOAD_SIZE_MB} MB.'}), 413


@app.route('/')
def home():
    return render_template('index.html')


@app.route('/api/upload', methods=['POST'])
def upload_file():
    if 'file' not in request.files:
        return jsonify({'error': 'No file provided'}), 400

    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'No file selected'}), 400

    if not allowed_file(file.filename):
        return jsonify({'error': 'File must be a .xer file'}), 400

    original_name = secure_filename(file.filename)
    unique_name = f"{uuid.uuid4().hex}_{original_name}"
    file_path = os.path.join(app.config['UPLOAD_FOLDER'], unique_name)
    file.save(file_path)

    sess_data = get_session_data()

    try:
        dashboard_data = analyze_xer_file(file_path, original_name, sess_data)
        if 'error' in dashboard_data:
            return jsonify(dashboard_data), 400

        analysis = sess_data['analysis']
        return jsonify({
            'success': True,
            'file_name': analysis['file_name'],
            'analyzed_at': analysis['analyzed_at'],
            'data': dashboard_data
        })
    except Exception as e:
        logger.exception("Upload analysis error")
        return jsonify({'error': str(e)}), 500


@app.route('/api/dashboard')
def get_dashboard():
    sess_data = get_session_data()
    analysis = sess_data['analysis']

    if analysis['dashboard_data'] is None:
        return jsonify({'has_data': False})

    return jsonify({
        'has_data': True,
        'file_name': analysis['file_name'],
        'analyzed_at': analysis['analyzed_at'],
        'data': analysis['dashboard_data']
    })


@app.route('/api/load-sample')
def load_sample():
    sample_path = os.path.join('input', 'sample.xer')
    if not os.path.exists(sample_path):
        return jsonify({'error': 'sample.xer not found in input/ folder'}), 404

    sess_data = get_session_data()
    try:
        dashboard_data = analyze_xer_file(sample_path, 'sample.xer', sess_data)
        analysis = sess_data['analysis']
        return jsonify({
            'success': True,
            'file_name': 'sample.xer',
            'analyzed_at': analysis['analyzed_at'],
            'data': dashboard_data
        })
    except Exception as e:
        logger.exception("Sample load error")
        return jsonify({'error': str(e)}), 500


@app.route('/api/export-excel')
def export_excel():
    sess_data = get_session_data()
    engine = sess_data['analysis']['engine']

    if engine is None:
        return jsonify({'error': 'No schedule loaded. Please upload an XER file first.'}), 400

    out_name = f"schedule_report_{uuid.uuid4().hex[:8]}.xlsx"
    output_path = os.path.join(app.config['OUTPUT_FOLDER'], out_name)
    
    reporter = ReportGenerator(engine)
    res_path = reporter.generate_full_report(output_path)

    if not res_path or not os.path.exists(res_path):
        return jsonify({'error': 'Failed to generate Excel report.'}), 500

    return send_file(
        res_path,
        as_attachment=True,
        download_name=f"schedule_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


@app.route('/gantt')
def gantt_view():
    return render_template('gantt.html')


@app.route('/api/gantt-data')
def get_gantt_data():
    sess_data = get_session_data()
    analysis = sess_data['analysis']

    if analysis['engine'] is None:
        return jsonify({'error': 'No data loaded. Please upload an XER file first.'}), 400

    max_acts = request.args.get('max', 2000, type=int)
    gantt_data = analysis['engine'].get_gantt_data(max_activities=max_acts)

    return jsonify({
        'success': True,
        'file_name': analysis['file_name'],
        'data': gantt_data
    })


@app.route('/comparison')
def comparison_view():
    return render_template('comparison.html')


@app.route('/api/compare', methods=['POST'])
def compare_schedules():
    if ScheduleComparator is None:
        return jsonify({'error': 'comparison_engine.py is missing!'}), 500

    if 'baseline' not in request.files or 'current' not in request.files:
        return jsonify({'error': 'Both baseline and current files required'}), 400

    baseline_file = request.files['baseline']
    current_file = request.files['current']

    if not (allowed_file(baseline_file.filename) and allowed_file(current_file.filename)):
        return jsonify({'error': 'Both files must be .xer files'}), 400

    baseline_name = secure_filename(baseline_file.filename)
    current_name = secure_filename(current_file.filename)

    baseline_path = os.path.join(app.config['UPLOAD_FOLDER'], f"bl_{uuid.uuid4().hex[:8]}_{baseline_name}")
    current_path = os.path.join(app.config['UPLOAD_FOLDER'], f"cur_{uuid.uuid4().hex[:8]}_{current_name}")

    baseline_file.save(baseline_path)
    current_file.save(current_path)

    sess_data = get_session_data()

    try:
        comparator = ScheduleComparator()
        comparator.load_baseline(baseline_path)
        comparator.load_current(current_path)
        results = comparator.compare()

        comp = sess_data['comparison']
        comp['comparator'] = comparator
        comp['results'] = results
        comp['baseline_file'] = baseline_name
        comp['current_file'] = current_name

        return jsonify({
            'success': True,
            'baseline_file': baseline_name,
            'current_file': current_name,
            'results': results
        })
    except Exception as e:
        logger.exception("Comparison error")
        return jsonify({'error': str(e)}), 500


@app.route('/api/comparison-data')
def get_comparison_data():
    sess_data = get_session_data()
    comp = sess_data['comparison']

    if comp['results'] is None:
        return jsonify({'has_data': False})

    return jsonify({
        'has_data': True,
        'baseline_file': comp['baseline_file'],
        'current_file': comp['current_file'],
        'results': comp['results']
    })


@app.route('/evm')
def evm_view():
    return render_template('evm.html')


@app.route('/api/evm-data')
def get_evm_data():
    sess_data = get_session_data()
    analysis = sess_data['analysis']

    if analysis['engine'] is None:
        return jsonify({'error': 'No data loaded. Upload an XER file first.'}), 400

    if EVMEngine is None:
        return jsonify({'error': 'evm_engine.py is missing!'}), 500

    try:
        evm = EVMEngine(analysis['engine'])
        results = evm.calculate()
        return jsonify({
            'success': True,
            'file_name': analysis['file_name'],
            'data': results
        })
    except Exception as e:
        logger.exception("EVM calculation error")
        return jsonify({'error': str(e)}), 500


@app.route('/health')
def health_view():
    return render_template('health.html')


@app.route('/api/health-data')
def get_health_data():
    sess_data = get_session_data()
    analysis = sess_data['analysis']

    if analysis['engine'] is None:
        return jsonify({'error': 'No data loaded. Upload a file first.'}), 400

    if AdvancedHealthEngine is None:
        return jsonify({'error': 'advanced_health_engine.py is missing!'}), 500

    selected_standard = request.args.get('standard', 'all')

    try:
        health = AdvancedHealthEngine(analysis['engine'])
        results = health.run_all_checks(selected_standard=selected_standard)

        return jsonify({
            'success': True,
            'file_name': analysis['file_name'],
            'data': results
        })
    except Exception as e:
        logger.exception("Health analysis error")
        return jsonify({'error': str(e)}), 500


@app.route('/api/executive-pdf')
def download_executive_pdf():
    sess_data = get_session_data()
    analysis = sess_data['analysis']

    if analysis['engine'] is None:
        return jsonify({'error': 'No data loaded'}), 400

    if PDFReportGenerator is None or AdvancedHealthEngine is None:
        return jsonify({'error': 'PDF or Health engine module missing!'}), 500

    selected_standard = request.args.get('standard', 'all')
    severity_filter = request.args.get('severity', 'all')

    try:
        health = AdvancedHealthEngine(analysis['engine'])
        results = health.run_all_checks(selected_standard=selected_standard)

        generator = PDFReportGenerator(
            results,
            analysis['file_name'],
            severity_filter=severity_filter
        )
        pdf_buffer = generator.generate_executive_report()

        return send_file(
            pdf_buffer,
            as_attachment=True,
            download_name=f"executive_report_{selected_standard}_{severity_filter}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf",
            mimetype='application/pdf'
        )
    except Exception as e:
        logger.exception("PDF generation error")
        return jsonify({'error': str(e)}), 500


@app.route('/api/actions-pdf')
def download_actions_pdf():
    sess_data = get_session_data()
    analysis = sess_data['analysis']

    if analysis['engine'] is None:
        return jsonify({'error': 'No data loaded'}), 400

    if PDFReportGenerator is None or AdvancedHealthEngine is None:
        return jsonify({'error': 'PDF or Health engine module missing!'}), 500

    selected_standard = request.args.get('standard', 'all')
    severity_filter = request.args.get('severity', 'all')

    try:
        health = AdvancedHealthEngine(analysis['engine'])
        results = health.run_all_checks(selected_standard=selected_standard)

        generator = PDFReportGenerator(
            results,
            analysis['file_name'],
            severity_filter=severity_filter
        )
        pdf_buffer = generator.generate_actions_report()

        return send_file(
            pdf_buffer,
            as_attachment=True,
            download_name=f"action_list_{selected_standard}_{severity_filter}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf",
            mimetype='application/pdf'
        )
    except Exception as e:
        logger.exception("PDF actions error")
        return jsonify({'error': str(e)}), 500


@app.route('/api/actions-excel')
def download_actions_excel():
    sess_data = get_session_data()
    analysis = sess_data['analysis']

    if analysis['engine'] is None:
        return jsonify({'error': 'No data loaded. Upload a file first.'}), 400

    if AdvancedHealthEngine is None:
        return jsonify({'error': 'advanced_health_engine.py is missing!'}), 500

    selected_standard = request.args.get('standard', 'all')
    severity_filter = request.args.get('severity', 'all').lower()

    severity_levels = {
        'critical': ['critical'],
        'high': ['critical', 'high'],
        'medium': ['critical', 'high', 'medium'],
        'all': ['critical', 'high', 'medium', 'low', 'info']
    }
    allowed_severities = severity_levels.get(severity_filter, severity_levels['all'])

    try:
        import pandas as pd
        from io import BytesIO
        from openpyxl.styles import Font, PatternFill, Alignment

        health = AdvancedHealthEngine(analysis['engine'])
        results = health.run_all_checks(selected_standard=selected_standard)

        top_actions = results.get('top_actions', []) or []
        standards_data = results.get('standards', {}) or {}

        filtered_top_actions = [
            a for a in top_actions
            if (a.get('severity') or 'low').lower() in allowed_severities
        ]

        meta_rows = [
            ['Report Type', 'Schedule Health - Top Actions Export'],
            ['Selected Standard', selected_standard],
            ['Severity Filter', severity_filter.upper()],
            ['File Name', analysis.get('file_name', '')],
            ['Generated At', results.get('analysis_date', '')],
            ['', ''],
            ['Overall Score', results.get('overall_score', '')],
            ['Total Checks', results.get('total_checks', '')],
            ['Passed Checks', results.get('passed_checks', '')],
            ['Failed Checks', results.get('failed_checks', '')],
            ['Critical Failures', results.get('critical_failures', '')],
            ['High Failures', results.get('high_failures', '')],
            ['Pass Rate (%)', results.get('pass_rate', '')],
            ['', ''],
            ['Filtered Actions Count', len(filtered_top_actions)],
        ]

        top_summary_rows = []
        for idx, action in enumerate(filtered_top_actions, 1):
            top_summary_rows.append({
                'Rank': idx,
                'Standard': action.get('standard', ''),
                'Check ID': action.get('id', ''),
                'Check Name': action.get('name', ''),
                'Category': action.get('category', ''),
                'Severity': (action.get('severity') or '').upper(),
                'Affected Count': action.get('count', 0),
                'Total': action.get('total', 0),
                'Percentage': action.get('percentage', 0),
                'Threshold': action.get('threshold', ''),
                'Value': action.get('value', ''),
                'Recommendation': action.get('recommendation', ''),
                'Description': action.get('description', ''),
            })

        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            pd.DataFrame(meta_rows, columns=['Field', 'Value']).to_excel(
                writer, sheet_name='Report Info', index=False
            )

            if top_summary_rows:
                pd.DataFrame(top_summary_rows).to_excel(
                    writer, sheet_name='Top Actions Summary', index=False
                )
            else:
                pd.DataFrame([{'Info': f'No actions matched severity filter: {severity_filter.upper()}'}]).to_excel(
                    writer, sheet_name='Top Actions Summary', index=False
                )

            for std_name, std_data in standards_data.items():
                sheet_name = (std_name or 'Standard')[:31]
                rows = []
                failed_in_std = []
                for category in std_data.get('categories', []):
                    for check in category.get('checks', []):
                        if check.get('status') != 'fail':
                            continue
                        if (check.get('severity') or 'low').lower() not in allowed_severities:
                            continue
                        failed_in_std.append((category.get('name', ''), check))

                if not failed_in_std:
                    rows.append({
                        'Section': 'No Matching Failures', 'Field': '',
                        'Value': f'No {severity_filter.upper()} failures found for {std_name}',
                        'Activity ID': '', 'Activity Name': '', 'WBS': '',
                    })
                else:
                    for cat_name, check in failed_in_std:
                        rows.append({
                            'Section': 'METRIC', 'Field': 'Check ID', 'Value': check.get('id', ''),
                            'Activity ID': '', 'Activity Name': '', 'WBS': '',
                        })
                        rows.append({
                            'Section': '', 'Field': 'Check Name', 'Value': check.get('name', ''),
                            'Activity ID': '', 'Activity Name': '', 'WBS': '',
                        })
                        rows.append({
                            'Section': '', 'Field': 'Category', 'Value': cat_name,
                            'Activity ID': '', 'Activity Name': '', 'WBS': '',
                        })
                        rows.append({
                            'Section': '', 'Field': 'Severity', 'Value': (check.get('severity') or '').upper(),
                            'Activity ID': '', 'Activity Name': '', 'WBS': '',
                        })
                        rows.append({
                            'Section': '', 'Field': 'Description', 'Value': check.get('description', ''),
                            'Activity ID': '', 'Activity Name': '', 'WBS': '',
                        })
                        rows.append({
                            'Section': '', 'Field': 'Threshold', 'Value': check.get('threshold', ''),
                            'Activity ID': '', 'Activity Name': '', 'WBS': '',
                        })

                        if check.get('count') is not None:
                            rows.append({
                                'Section': '', 'Field': 'Affected',
                                'Value': f"{check.get('count', 0)} of {check.get('total', 0)} ({check.get('percentage', 0)}%)",
                                'Activity ID': '', 'Activity Name': '', 'WBS': '',
                            })
                        if check.get('value') is not None and check.get('value') != '':
                            rows.append({
                                'Section': '', 'Field': 'Value', 'Value': check.get('value', ''),
                                'Activity ID': '', 'Activity Name': '', 'WBS': '',
                            })

                        rows.append({
                            'Section': '', 'Field': 'Recommendation', 'Value': check.get('recommendation', ''),
                            'Activity ID': '', 'Activity Name': '', 'WBS': '',
                        })

                        items = check.get('failed_items', []) or []
                        rows.append({
                            'Section': 'AFFECTED ACTIVITIES', 'Field': f'Total: {len(items)}', 'Value': '',
                            'Activity ID': 'Activity ID', 'Activity Name': 'Activity Name', 'WBS': 'WBS',
                        })

                        if items:
                            for item in items:
                                rows.append({
                                    'Section': '', 'Field': '', 'Value': '',
                                    'Activity ID': item.get('code', ''),
                                    'Activity Name': item.get('name', ''),
                                    'WBS': item.get('wbs', ''),
                                })
                        else:
                            rows.append({
                                'Section': '', 'Field': '', 'Value': '(No activity list available)',
                                'Activity ID': '', 'Activity Name': '', 'WBS': '',
                            })

                        rows.append({'Section': '', 'Field': '', 'Value': '', 'Activity ID': '', 'Activity Name': '', 'WBS': ''})

                pd.DataFrame(rows).to_excel(writer, sheet_name=sheet_name, index=False)

            workbook = writer.book
            header_font = Font(bold=True, color='FFFFFF', size=11)
            header_fill = PatternFill(start_color='1E40AF', end_color='1E40AF', fill_type='solid')
            metric_fill = PatternFill(start_color='FEF3C7', end_color='FEF3C7', fill_type='solid')
            activity_hdr_fill = PatternFill(start_color='DBEAFE', end_color='DBEAFE', fill_type='solid')
            bold_font = Font(bold=True)
            wrap_align = Alignment(wrap_text=True, vertical='top')

            for sheet_name in workbook.sheetnames:
                ws = workbook[sheet_name]
                for cell in ws[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal='left', vertical='center')

                for col in ws.columns:
                    max_len = 10
                    col_letter = col[0].column_letter
                    for cell in col:
                        try:
                            val = str(cell.value) if cell.value is not None else ''
                            if len(val) > max_len:
                                max_len = len(val)
                        except Exception:
                            pass
                    ws.column_dimensions[col_letter].width = min(max_len + 2, 60)

                ws.freeze_panes = 'A2'

                if sheet_name not in ['Report Info', 'Top Actions Summary']:
                    for row in ws.iter_rows(min_row=2):
                        section_cell = row[0]
                        if section_cell.value == 'METRIC':
                            for c in row:
                                c.fill = metric_fill
                                c.font = bold_font
                        elif section_cell.value == 'AFFECTED ACTIVITIES':
                            for c in row:
                                c.fill = activity_hdr_fill
                                c.font = bold_font
                        for c in row:
                            c.alignment = wrap_align

        output.seek(0)
        filename = f"health_top_actions_{selected_standard}_{severity_filter}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    except Exception as e:
        logger.exception("Actions Excel error")
        return jsonify({'error': str(e)}), 500


# ════════════════════════════════════════════
# ROUTE 8: AI EXECUTIVE NARRATIVE BRIEFING
# ════════════════════════════════════════════

@app.route('/api/ai-narrative')
def get_ai_narrative():
    sess_data = get_session_data()
    analysis = sess_data['analysis']

    if analysis['engine'] is None:
        return jsonify({'error': 'No schedule loaded. Upload an XER file first.'}), 400

    if AINarrativeEngine is None:
        return jsonify({'error': 'ai_narrative_engine.py is missing!'}), 500

    try:
        health_data = {}
        if AdvancedHealthEngine is not None:
            try:
                health_gen = AdvancedHealthEngine(analysis['engine'])
                health_data = health_gen.run_all_checks('all')
            except Exception as he:
                logger.warning("Health data unavailable for narrative: %s", he)

        evm_data = {}
        if EVMEngine is not None:
            try:
                evm_data = EVMEngine(analysis['engine']).calculate()
            except Exception:
                pass

        comp_data = sess_data.get('comparison', {}).get('results', {}) or {}

        narrative_gen = AINarrativeEngine(
            health_data=health_data,
            comparison_data=comp_data,
            evm_data=evm_data
        )
        results = narrative_gen.generate_narrative()

        return jsonify({
            'success': True,
            'file_name': analysis['file_name'],
            'data': results
        })
    except Exception as e:
        logger.exception("AI Narrative generation error")
        return jsonify({'error': str(e)}), 500


if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    debug_mode = os.environ.get('FLASK_ENV') != 'production'

    logger.info("============================================================")
    logger.info("🚀 P6 SCHEDULE ANALYZER - AI NARRATIVE + FULL SUITE READY")
    logger.info("============================================================")
    logger.info("📌 Running on port %s", port)
    logger.info("🔧 Debug mode: %s", debug_mode)
    logger.info("👉 Open in browser: http://localhost:%s", port)

    app.run(debug=debug_mode, host='0.0.0.0', port=port)
'''

with open("app.py", "w", encoding="utf-8") as f:
    f.write(APP_CODE)
print("  ✅ Updated app.py")


# ==============================================================================
# FILE 3a: templates/health.html (Full Rewrite with AI Narrative UI Card)
# ==============================================================================

HEALTH_HTML_CODE = '''<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Schedule Health | {{ config.app_title }}</title>

    <style>
        :root {
            --color-primary: {{ config.theme.primary }};
            --color-primary-dark: {{ config.theme.primary_dark }};
            --color-accent: {{ config.theme.accent }};
            --color-success: {{ config.theme.success }};
            --color-warning: {{ config.theme.warning }};
            --color-danger: {{ config.theme.danger }};
            --color-bg: {{ config.theme.bg }};
            --color-surface: {{ config.theme.surface }};
            --color-text: {{ config.theme.text }};
            --color-muted: {{ config.theme.muted }};
            --color-border: {{ config.theme.border }};
        }
    </style>

    <link rel="stylesheet" href="{{ url_for('static', filename='style.css') }}">

    <style>
        a.btn { text-decoration: none; }
        .app-header .btn-secondary[aria-current="page"] {
            background: rgba(255,255,255,0.4);
            border-color: #fff;
            font-weight: 700;
        }
        .std-selector { display: flex; flex-wrap: wrap; gap: 0.5rem; margin-bottom: 1.25rem; }
        .std-select-btn {
            padding: 0.45rem 0.9rem; border-radius: 999px; border: 1px solid var(--color-border);
            background: var(--color-surface); color: var(--color-text);
            font-size: 0.85rem; font-weight: 600; cursor: pointer; transition: all 0.15s;
        }
        .std-select-btn:hover { border-color: var(--color-accent); color: var(--color-accent); }
        .std-select-btn.active { background: var(--color-primary); border-color: var(--color-primary); color: #fff; }
        
        .health-hero {
            background: var(--color-surface); border: 1px solid var(--color-border);
            border-radius: 12px; padding: 1.5rem; margin-bottom: 1.25rem;
            box-shadow: 0 2px 4px rgba(0,0,0,0.05);
            display: grid; grid-template-columns: minmax(160px, 220px) 1fr;
            gap: 1.5rem; align-items: center;
        }
        @media (max-width: 768px) { .health-hero { grid-template-columns: 1fr; text-align: center; } }
        .score-ring {
            text-align: center; padding: 1rem; border-radius: 12px;
            background: linear-gradient(135deg, #eff6ff, #f8fafc); border: 1px solid var(--color-border);
        }
        .score-ring .big { font-size: 3rem; font-weight: 800; color: var(--color-primary); line-height: 1.1; }
        .score-ring .sub { font-size: 0.85rem; color: var(--color-muted); }
        .hero-stats { display: grid; grid-template-columns: repeat(auto-fit, minmax(120px, 1fr)); gap: 0.75rem; }
        .hero-stat { background: #f8fafc; border: 1px solid var(--color-border); border-radius: 8px; padding: 0.75rem; }
        .hero-stat .val { font-size: 1.35rem; font-weight: 700; color: var(--color-text); }
        .hero-stat .lbl { font-size: 0.75rem; color: var(--color-muted); text-transform: uppercase; letter-spacing: 0.02em; }
        .hero-stat.danger .val { color: var(--color-danger); }

        .score-grid { display: grid; grid-template-columns: repeat(auto-fill, minmax(160px, 1fr)); gap: 0.85rem; margin-bottom: 1.5rem; }
        .std-score-card {
            background: var(--color-surface); border: 1px solid var(--color-border); border-radius: 10px;
            padding: 1rem; box-shadow: 0 1px 3px rgba(0,0,0,0.04);
            transition: transform 0.15s, box-shadow 0.15s; border-top: 4px solid #94a3b8;
        }
        .std-score-card:hover { transform: translateY(-2px); box-shadow: 0 6px 16px rgba(0,0,0,0.08); }
        .std-score-card.green { border-top-color: #10b981; }
        .std-score-card.blue { border-top-color: #3b82f6; }
        .std-score-card.orange { border-top-color: #f59e0b; }
        .std-score-card.red { border-top-color: #dc2626; }
        .std-score-value { font-size: 1.75rem; font-weight: 800; margin: 0.25rem 0; color: var(--color-text); }
        .std-score-grade { display: inline-block; padding: 0.15rem 0.55rem; border-radius: 999px; font-size: 0.75rem; font-weight: 700; margin-bottom: 0.35rem; }
        .grade-A { background: #d1fae5; color: #065f46; }
        .grade-B { background: #dbeafe; color: #1e40af; }
        .grade-C { background: #fef3c7; color: #92400e; }
        .grade-D { background: #ffedd5; color: #9a3412; }
        .grade-F { background: #fee2e2; color: #991b1b; }
        .std-score-details { font-size: 0.8rem; color: var(--color-muted); line-height: 1.4; }
        
        .health-toolbar {
            display: flex; flex-wrap: wrap; gap: 0.75rem; align-items: center; justify-content: space-between;
            margin-bottom: 1.25rem; padding: 0.85rem 1rem;
            background: var(--color-surface); border: 1px solid var(--color-border); border-radius: 10px;
        }
        .health-toolbar .left, .health-toolbar .right { display: flex; flex-wrap: wrap; gap: 0.5rem; align-items: center; }
        .health-toolbar label { font-size: 0.8rem; color: var(--color-muted); font-weight: 600; }
        .health-toolbar select, .filter-bar select, .filter-bar input {
            padding: 0.4rem 0.6rem; border: 1px solid var(--color-border); border-radius: 6px; font-size: 0.85rem; background: #fff;
        }
        
        .action-item { display: flex; gap: 0.85rem; padding: 1rem; border: 1px solid var(--color-border); border-radius: 10px; background: #fff; margin-bottom: 0.65rem; }
        .action-priority { width: 2rem; height: 2rem; border-radius: 50%; color: #fff; display: flex; align-items: center; justify-content: center; font-weight: 800; font-size: 0.85rem; flex-shrink: 0; }

        .badge { display: inline-block; padding: 0.12rem 0.45rem; border-radius: 999px; font-size: 0.7rem; font-weight: 700; text-transform: uppercase; margin-left: 0.25rem; vertical-align: middle; }
        .badge-critical { background: #7f1d1d; color: #fff; }
        .badge-high { background: #fee2e2; color: #991b1b; }
        .badge-medium { background: #fef3c7; color: #92400e; }
        .badge-low { background: #e2e8f0; color: #475569; }
        .badge-info { background: #e2e8f0; color: #334155; }
        .badge-std { background: #dbeafe; color: #1e40af; }
        
        .recommendation-box { margin-top: 0.5rem; padding: 0.55rem 0.75rem; background: #fffbeb; border: 1px solid #f59e0b; border-radius: 6px; font-size: 0.85rem; color: #92400e; }
        
        .filter-bar { display: flex; flex-wrap: wrap; gap: 0.65rem; align-items: center; margin-bottom: 1rem; padding: 0.75rem 1rem; background: #f8fafc; border: 1px solid var(--color-border); border-radius: 8px; }
        .filter-bar input[type="search"], .filter-bar input[type="text"] { min-width: 200px; flex: 1; }

        .category-section { background: var(--color-surface); border: 1px solid var(--color-border); border-radius: 10px; padding: 1rem 1.15rem; margin-bottom: 1rem; }
        .category-header { display: flex; justify-content: space-between; align-items: flex-start; gap: 1rem; margin-bottom: 0.85rem; padding-bottom: 0.65rem; border-bottom: 1px solid var(--color-border); }
        .category-header h3 { font-size: 1.05rem; margin: 0; }
        .category-stats { font-size: 0.85rem; color: var(--color-muted); white-space: nowrap; }

        .check-item { display: flex; gap: 0.75rem; padding: 0.85rem; border-radius: 8px; margin-bottom: 0.5rem; border: 1px solid var(--color-border); background: #f8fafc; }
        .check-item.pass { border-left: 4px solid #10b981; background: #f0fdf4; }
        .check-item.fail { border-left: 4px solid #dc2626; background: #fef2f2; }
        .check-item.info, .check-item.na { border-left: 4px solid #94a3b8; background: #f8fafc; }
        .check-icon { font-size: 1.1rem; line-height: 1.4; }
        .check-content { flex: 1; min-width: 0; }
        .check-title { font-weight: 600; margin-bottom: 0.25rem; display: flex; flex-wrap: wrap; align-items: center; gap: 0.15rem; }

        .section-title { font-size: 1.15rem; margin: 1.25rem 0 0.75rem; color: var(--color-text); }

        /* AI Narrative Card Styles */
        #aiNarrativeSection {
            background: linear-gradient(135deg, #ffffff, #f5f9ff);
            border: 1px solid var(--color-accent);
            border-left: 6px solid var(--color-accent);
        }
        #aiNarrativeBody h3 { margin: 1rem 0 0.4rem; color: var(--color-primary); font-size: 1.1rem; }
        #aiNarrativeBody strong { color: var(--color-primary-dark); }
    </style>
</head>
<body>
    <header class="app-header">
        <div class="header-content">
            <div class="logo-section">
                {% if config.use_logo_image %}
                    <img src="{{ url_for('static', filename=config.logo_path) }}" alt="Logo" style="height: 45px;">
                {% else %}
                    <span class="logo-icon">🏥</span>
                {% endif %}
                <div>
                    <h1>{{ config.app_title }} — Health</h1>
                    <p class="subtitle">622+ metrics across DCMA, DOE, NASA, GAO, AACE &amp; Industry</p>
                </div>
            </div>
            <div class="header-actions">
                <a href="/" class="btn btn-secondary">📊 Dashboard</a>
                <a href="/gantt" class="btn btn-secondary">📅 Gantt</a>
                <a href="/comparison" class="btn btn-secondary">🔄 Compare</a>
                <a href="/evm" class="btn btn-secondary">📈 EVM</a>
                <a href="/health" class="btn btn-secondary" aria-current="page">🏥 Health</a>
            </div>
        </div>
    </header>

    <main class="app-main">
        <div id="loadingMessage" class="loading-screen" style="display:flex;">
            <div class="spinner"></div>
            <p>Loading health analytics…</p>
        </div>

        <div id="healthContent" style="display:none;">

            <div class="std-selector" role="tablist" aria-label="Health standard">
                <button type="button" class="std-select-btn active" data-std="all">All Standards</button>
                <button type="button" class="std-select-btn" data-std="DCMA">DCMA</button>
                <button type="button" class="std-select-btn" data-std="DOE">DOE</button>
                <button type="button" class="std-select-btn" data-std="NASA">NASA</button>
                <button type="button" class="std-select-btn" data-std="GAO">GAO</button>
                <button type="button" class="std-select-btn" data-std="AACE">AACE</button>
                <button type="button" class="std-select-btn" data-std="Industry">Industry</button>
            </div>

            <div class="health-hero">
                <div class="score-ring">
                    <div class="big" id="overallScore">--</div>
                    <div class="sub">Overall score / 100</div>
                    <div style="margin-top:0.75rem;">
                        <div id="reportTitle" style="font-weight:700;font-size:1rem;">Assessment</div>
                        <div id="reportSubtitle" style="font-size:0.85rem;color:#64748b;">—</div>
                    </div>
                </div>
                <div class="hero-stats">
                    <div class="hero-stat"><div class="val" id="totalChecks">--</div><div class="lbl">Total checks</div></div>
                    <div class="hero-stat"><div class="val" id="passedChecks">--</div><div class="lbl">Passed</div></div>
                    <div class="hero-stat"><div class="val" id="failedChecks">--</div><div class="lbl">Failed</div></div>
                    <div class="hero-stat danger"><div class="val" id="criticalFailures">--</div><div class="lbl">Critical fails</div></div>
                </div>
            </div>

            <!-- ═══════════════════════════════════════ -->
            <!-- 🤖 AI EXECUTIVE NARRATIVE BRIEFING -->
            <!-- ═══════════════════════════════════════ -->
            <div class="dcma-section" id="aiNarrativeSection" style="margin-bottom: 1.75rem;">
                <div style="display: flex; justify-content: space-between; align-items: center; margin-bottom: 1rem; flex-wrap: wrap; gap: 0.5rem;">
                    <h2 style="margin: 0;">🤖 AI Executive Narrative Briefing</h2>
                    <button type="button" class="btn btn-primary" onclick="fetchAINarrative(true)">
                        🔄 Generate Briefing
                    </button>
                </div>
                <div id="aiNarrativeMethod" style="font-size: 0.8rem; color: var(--color-muted); margin-bottom: 0.75rem;"></div>
                <div id="aiNarrativeBody" style="font-size: 0.95rem; line-height: 1.6; color: var(--color-text);">
                    <p style="color: var(--color-muted);">💡 Click <strong>"Generate Briefing"</strong> to synthesize health, EVM, and delay metrics into a professional executive narrative report.</p>
                </div>
            </div>

            <div class="health-toolbar">
                <div class="left">
                    <label for="excelSeverityFilter">Export severity</label>
                    <select id="excelSeverityFilter" title="Applies to PDF and Excel exports and Top Actions list">
                        <option value="all">All</option>
                        <option value="critical">Critical only</option>
                        <option value="high">Critical + High</option>
                        <option value="medium">Critical + High + Medium</option>
                    </select>
                </div>
                <div class="right">
                    <button type="button" class="btn btn-primary" onclick="downloadPDF(event)">📄 Executive PDF</button>
                    <button type="button" class="btn btn-secondary" onclick="downloadActionsPDF(event)" style="background:#334155;color:#fff;border:none;">📋 Actions PDF</button>
                    <button type="button" class="btn btn-success" onclick="downloadActionsExcel(event)">⬇️ Actions Excel</button>
                </div>
            </div>

            <h2 class="section-title">Standards scores</h2>
            <p style="font-size:0.85rem;color:#64748b;margin:-0.35rem 0 0.75rem;">Click a card to drill into that standard only.</p>
            <div id="scoreGrid" class="score-grid"></div>

            <div id="topActionsSection" style="display:none;">
                <h2 class="section-title">Top priority actions</h2>
                <p style="font-size:0.85rem;color:#64748b;margin:-0.35rem 0 0.75rem;">Filtered by export severity. Expand rows to see affected activities.</p>
                <div id="topActionsList"></div>
            </div>

            <h2 class="section-title">Detailed check results</h2>
            <div class="filter-bar">
                <label for="filterStatus">Status</label>
                <select id="filterStatus">
                    <option value="all">All</option>
                    <option value="fail">Failed</option>
                    <option value="pass">Passed</option>
                    <option value="info">Info</option>
                    <option value="na">N/A</option>
                </select>
                <label for="filterSeverity">Severity</label>
                <select id="filterSeverity">
                    <option value="all">All</option>
                    <option value="critical">Critical</option>
                    <option value="high">Critical + High</option>
                    <option value="medium">Crit + High + Med</option>
                </select>
                <label for="filterSearch">Search</label>
                <input type="search" id="filterSearch" placeholder="Check ID or name…" autocomplete="off">
                <button type="button" class="btn btn-outline" onclick="applyFilter()">Apply</button>
            </div>
            <div id="detailedResults"></div>
        </div>
    </main>

    <footer class="app-footer">
        <p>&copy; {{ config.footer_year }} {{ config.company_name }} | {{ config.footer_text }}</p>
    </footer>

    <script src="{{ url_for('static', filename='health.js') }}"></script>
</body>
</html>
'''

os.makedirs("templates", exist_ok=True)
with open("templates/health.html", "w", encoding="utf-8") as f:
    f.write(HEALTH_HTML_CODE)
print("  ✅ Updated templates/health.html")


# ==============================================================================
# FILE 3b: static/health.js (Full Rewrite with AI Narrative Fetcher)
# ==============================================================================

HEALTH_JS_CODE = '''/*
    ADVANCED HEALTH DASHBOARD (Patched + AI Narrative Support)
    ==========================================================
*/

let healthData = null;
let currentStandard = 'all';
let searchTimer = null;

const MAX_ITEMS_UI = 200;
const MAX_TOP_ACTIONS_UI = 15;

const SEVERITY_LEVELS = {
    critical: ['critical'],
    high: ['critical', 'high'],
    medium: ['critical', 'high', 'medium'],
    all: ['critical', 'high', 'medium', 'low', 'info']
};

// ═══════════════════════════════════════════
// UTILITIES
// ═══════════════════════════════════════════

function esc(s) {
    if (s == null) return '';
    return String(s)
        .replace(/&/g, '&amp;')
        .replace(/</g, '&lt;')
        .replace(/>/g, '&gt;')
        .replace(/"/g, '&quot;')
        .replace(/'/g, '&#39;');
}

function severityAllowed(sev, filter) {
    const f = (filter || 'all').toLowerCase();
    const list = SEVERITY_LEVELS[f] || SEVERITY_LEVELS.all;
    return list.indexOf((sev || 'low').toLowerCase()) >= 0;
}

function getSelectedSeverity() {
    const el = document.getElementById('excelSeverityFilter');
    return el && el.value ? el.value.toLowerCase() : 'all';
}

function getFilterValue(id, fallback) {
    const el = document.getElementById(id);
    if (!el) return fallback;
    return el.value != null ? el.value : fallback;
}

async function safeFetchJSON(url, options) {
    const res = await fetch(url, options || {});
    let data = {};
    try {
        data = await res.json();
    } catch (e) {
        if (!res.ok) throw new Error('Request failed (' + res.status + ')');
        throw new Error('Invalid JSON response from server');
    }
    if (!res.ok || data.error) {
        throw new Error(data.error || ('Request failed (' + res.status + ')'));
    }
    return data;
}

async function downloadFile(url, defaultName, btn) {
    const label = btn ? btn.textContent : '';
    try {
        if (btn) {
            btn.disabled = true;
            btn.textContent = '⏳ Working…';
        }
        const res = await fetch(url);
        if (!res.ok) {
            let msg = 'Download failed (' + res.status + ')';
            try {
                const j = await res.json();
                if (j.error) msg = j.error;
            } catch (e) { /* ignore */ }
            throw new Error(msg);
        }
        const blob = await res.blob();
        const cd = res.headers.get('content-disposition') || '';
        const m = cd.match(/filename\\*?=(?:UTF-8''|")?([^\\";]+)/i);
        const name = m ? decodeURIComponent(m[1].replace(/"/g, '')) : defaultName;
        const a = document.createElement('a');
        a.href = URL.createObjectURL(blob);
        a.download = name;
        document.body.appendChild(a);
        a.click();
        a.remove();
        URL.revokeObjectURL(a.href);
    } catch (err) {
        alert('❌ ' + (err.message || 'Download failed'));
    } finally {
        if (btn) {
            btn.disabled = false;
            btn.textContent = label;
        }
    }
}

function showHealthError(msg) {
    const loading = document.getElementById('loadingMessage');
    const content = document.getElementById('healthContent');
    if (content) content.style.display = 'none';
    if (loading) {
        loading.style.display = 'block';
        loading.innerHTML =
            '<p style="color:#dc2626;">❌ ' + esc(msg) + '</p>' +
            '<p style="color:#64748b;margin-top:0.5rem;">Upload an XER on the Dashboard first.</p>' +
            '<a href="/" class="btn btn-primary" style="margin-top:1rem;display:inline-flex;">← Dashboard</a>';
    }
}

function itemRowHtml(item) {
    const code = esc(item && item.code);
    const name = item && item.name ? ' - ' + esc(item.name) : '';
    const wbs = item && item.wbs
        ? ' <span style="color:#64748b;">(' + esc(item.wbs) + ')</span>'
        : '';
    return (
        '<div style="font-size:0.82rem;padding:0.15rem 0;border-bottom:1px solid #f1f5f9;">' +
        '<strong>' + code + '</strong>' + name + wbs +
        '</div>'
    );
}

function renderItemsBlock(items, summaryLabel) {
    const list = Array.isArray(items) ? items : [];
    if (!list.length) {
        return (
            '<div style="margin-top:0.4rem;font-size:0.82rem;color:#64748b;">' +
            'No activity list available for this metric.</div>'
        );
    }
    const shown = list.slice(0, MAX_ITEMS_UI);
    const more = list.length - shown.length;
    let html =
        '<details style="margin-top:0.5rem;">' +
        '<summary style="cursor:pointer;color:#1d4ed8;font-size:0.85rem;">' +
        esc(summaryLabel || 'Show affected items') + ' (' + list.length + ')</summary>' +
        '<div style="margin-top:0.4rem;background:#fff;border:1px solid #e2e8f0;' +
        'border-radius:6px;padding:0.6rem;max-height:280px;overflow:auto;">';

    shown.forEach(function (item) { html += itemRowHtml(item); });
    if (more > 0) {
        html +=
            '<div style="font-size:0.82rem;color:#64748b;padding-top:0.35rem;">' +
            '… and ' + more + ' more (see Excel export for full list)</div>';
    }
    html += '</div></details>';
    return html;
}

function statusIcon(status) {
    const map = { pass: '✅', fail: '❌', info: 'ℹ️', na: '⚪' };
    return map[status] || 'ℹ️';
}

// ═══════════════════════════════════════════
// BOOT
// ═══════════════════════════════════════════

document.addEventListener('DOMContentLoaded', function () {
    try {
        const params = new URLSearchParams(window.location.search || '');
        const std = params.get('standard');
        if (std) currentStandard = std;
    } catch (e) { /* ignore */ }

    wireFilterListeners();
    loadHealthData(currentStandard);
});

function wireFilterListeners() {
    const search = document.getElementById('filterSearch');
    if (search) {
        search.addEventListener('input', function () {
            clearTimeout(searchTimer);
            searchTimer = setTimeout(function () { applyFilter(); }, 250);
        });
    }

    ['filterStatus', 'filterSeverity'].forEach(function (id) {
        const el = document.getElementById(id);
        if (el) el.addEventListener('change', applyFilter);
    });

    const sevExport = document.getElementById('excelSeverityFilter');
    if (sevExport) {
        sevExport.addEventListener('change', function () { renderTopActions(); });
    }

    document.querySelectorAll('.std-select-btn').forEach(function (btn) {
        btn.addEventListener('click', function () {
            const std = btn.getAttribute('data-std');
            if (std) selectStandard(std);
        });
    });
}

// ═══════════════════════════════════════════
// DATA LOAD
// ═══════════════════════════════════════════

function selectStandard(standard) {
    currentStandard = standard || 'all';
    document.querySelectorAll('.std-select-btn').forEach(function (btn) {
        btn.classList.toggle('active', btn.getAttribute('data-std') === currentStandard);
    });
    loadHealthData(currentStandard);
}

async function loadHealthData(standard) {
    const loading = document.getElementById('loadingMessage');
    const content = document.getElementById('healthContent');
    if (loading) {
        loading.style.display = 'block';
        loading.innerHTML = '<p>Loading health analytics…</p>';
    }
    if (content) content.style.display = 'none';

    const std = standard || 'all';

    try {
        const response = await safeFetchJSON('/api/health-data?standard=' + encodeURIComponent(std));
        healthData = response.data || {};
        currentStandard = std;
        renderDashboard();
    } catch (err) {
        console.error(err);
        showHealthError(err.message || 'Failed to load health data');
    }
}

// ═══════════════════════════════════════════
// DASHBOARD RENDER
// ═══════════════════════════════════════════

function setText(id, value) {
    const el = document.getElementById(id);
    if (el) el.textContent = value == null ? '' : String(value);
}

function renderDashboard() {
    const loading = document.getElementById('loadingMessage');
    const content = document.getElementById('healthContent');
    if (loading) loading.style.display = 'none';
    if (content) content.style.display = 'block';

    if (!healthData) { showHealthError('No health data returned'); return; }

    setText('overallScore', healthData.overall_score);
    setText('totalChecks', healthData.total_checks);
    setText('passedChecks', healthData.passed_checks);
    setText('failedChecks', healthData.failed_checks);
    setText('criticalFailures', healthData.critical_failures);

    const stdName = currentStandard === 'all' ? 'All Standards' : currentStandard;
    setText('reportTitle', currentStandard === 'all' ? 'Comprehensive Assessment' : stdName + ' Assessment');
    setText('reportSubtitle',
        currentStandard === 'all'
            ? 'Analysis based on all applicable standards'
            : 'Detailed analysis of ' + stdName + ' compliance'
    );

    renderStandardsScores();
    renderTopActions();
    renderDetailedResults();
}

function renderStandardsScores() {
    const container = document.getElementById('scoreGrid');
    if (!container) return;
    container.innerHTML = '';

    const scores = healthData.standard_scores || {};
    const keys = Object.keys(scores);
    if (!keys.length) {
        container.innerHTML = '<p style="color:#64748b;">No standards evaluated.</p>';
        return;
    }

    keys.forEach(function (std) {
        const data = scores[std] || {};
        const div = document.createElement('div');
        div.className = 'std-score-card ' + esc(data.color || '');
        div.style.cursor = 'pointer';

        const failedLine = data.failed > 0
            ? '<span style="color:#dc2626;">' + esc(data.failed) + ' failed</span>'
            : 'All passed ✅';

        div.innerHTML =
            '<div style="font-size:0.85rem;color:#64748b;font-weight:600;">' + esc(std) + '</div>' +
            '<div class="std-score-value">' + esc(data.score) + '</div>' +
            '<div class="std-score-grade grade-' + esc(data.grade) + '">Grade ' + esc(data.grade) + '</div>' +
            '<div class="std-score-details">' +
            esc(data.passed) + '/' + esc(data.total_checks) + ' passed<br>' + failedLine +
            '</div>';

        div.addEventListener('click', function () { selectStandard(std); });
        container.appendChild(div);
    });
}

function renderTopActions() {
    const container = document.getElementById('topActionsList');
    const section = document.getElementById('topActionsSection');
    if (!container || !section) return;

    const raw = (healthData && healthData.top_actions) || [];
    const sev = getSelectedSeverity();
    const list = raw
        .filter(function (a) { return severityAllowed(a.severity, sev); })
        .slice(0, MAX_TOP_ACTIONS_UI);

    if (!list.length) {
        if (!raw.length) { section.style.display = 'none'; return; }
        section.style.display = 'block';
        container.innerHTML =
            '<p style="color:#64748b;padding:0.5rem 0;">No top actions match severity filter: ' +
            esc(sev.toUpperCase()) + '.</p>';
        return;
    }

    section.style.display = 'block';
    container.innerHTML = '';

    list.forEach(function (action, idx) {
        const severity = (action.severity || 'low').toLowerCase();
        const severityColor = {
            critical: '#7f1d1d', high: '#dc2626', medium: '#f59e0b', low: '#64748b', info: '#64748b'
        }[severity] || '#64748b';

        const failedItems = action.failed_items || [];
        const itemsHtml = renderItemsBlock(failedItems, 'Show affected activities');

        let metricText;
        if (action.count !== undefined && action.count !== null) {
            metricText = esc(action.count) + ' activities affected (' + esc(action.percentage || 0) + '%)';
        } else if (action.value !== undefined && action.value !== null) {
            metricText = 'Value: ' + esc(action.value);
        } else {
            metricText = 'Review required';
        }

        const div = document.createElement('div');
        div.className = 'action-item';
        div.innerHTML =
            '<div class="action-priority" style="background:' + severityColor + ';">' + (idx + 1) + '</div>' +
            '<div style="flex:1;">' +
            '<div style="font-weight:600;">' +
            esc(action.id || '') + ': ' + esc(action.name || '') + ' ' +
            '<span class="badge badge-' + esc(severity) + '">' + esc(severity.toUpperCase()) + '</span> ' +
            '<span class="badge badge-std">' + esc(action.standard || '') + '</span>' +
            '</div>' +
            '<div style="font-size:0.85rem;color:#64748b;margin-top:0.25rem;">' +
            (action.category ? 'Category: ' + esc(action.category) + ' | ' : '') + metricText +
            '</div>' +
            (action.recommendation
                ? '<div class="recommendation-box">💡 ' + esc(action.recommendation) + '</div>'
                : '') +
            itemsHtml + '</div>';

        container.appendChild(div);
    });
}

function renderDetailedResults() {
    const container = document.getElementById('detailedResults');
    if (!container) return;
    container.innerHTML = '';

    if (!healthData || !healthData.standards) {
        container.innerHTML = '<p style="text-align:center;padding:2rem;color:#64748b;">No detailed results.</p>';
        return;
    }

    const filterStatus = getFilterValue('filterStatus', 'all');
    const filterSeverity = getFilterValue('filterSeverity', 'all').toLowerCase();
    const filterSearch = String(getFilterValue('filterSearch', '')).toLowerCase();

    const frag = document.createDocumentFragment();
    let sections = 0;

    Object.keys(healthData.standards).forEach(function (stdName) {
        const stdData = healthData.standards[stdName] || {};
        const categories = stdData.categories || [];

        categories.forEach(function (category) {
            const checks = category.checks || [];
            const filteredChecks = checks.filter(function (check) {
                if (filterStatus !== 'all' && check.status !== filterStatus) return false;
                if (filterSeverity !== 'all' && !severityAllowed(check.severity, filterSeverity)) return false;
                if (filterSearch) {
                    const name = String(check.name || '').toLowerCase();
                    const id = String(check.id || '').toLowerCase();
                    if (name.indexOf(filterSearch) < 0 && id.indexOf(filterSearch) < 0) return false;
                }
                return true;
            });

            if (!filteredChecks.length) return;

            sections += 1;
            const section = document.createElement('div');
            section.className = 'category-section';

            const passed = filteredChecks.filter(function (c) { return c.passed; }).length;
            const failed = filteredChecks.filter(function (c) { return c.status === 'fail'; }).length;

            section.innerHTML =
                '<div class="category-header">' +
                '<div><h3>' + esc(category.name) + '</h3>' +
                '<div style="font-size:0.85rem;color:#64748b;">' + esc(stdName) + '</div></div>' +
                '<div class="category-stats">' +
                passed + '/' + filteredChecks.length + ' passed' +
                (failed > 0 ? ' | <span style="color:#dc2626;">' + failed + ' failed</span>' : '') +
                '</div></div>' +
                '<div class="checks-list"></div>';

            const checksList = section.querySelector('.checks-list');
            filteredChecks.forEach(function (check) { checksList.appendChild(createCheckItem(check)); });
            frag.appendChild(section);
        });
    });

    if (!sections) {
        container.innerHTML = '<p style="text-align:center;padding:2rem;color:#64748b;">No checks match your filter criteria.</p>';
        return;
    }
    container.appendChild(frag);
}

function createCheckItem(check) {
    const div = document.createElement('div');
    const status = check.status || (check.passed ? 'pass' : 'fail');
    div.className = 'check-item ' + esc(status);

    const icon = statusIcon(status);
    const severity = (check.severity || 'low').toLowerCase();

    let details = '';
    if (check.value !== undefined && check.value !== null && check.value !== '') {
        details = '<strong>Value:</strong> ' + esc(check.value) + esc(check.unit || '');
    } else if (check.count !== undefined && check.count !== null) {
        details = '<strong>Count:</strong> ' + esc(check.count) + ' / ' + esc(check.total) + ' (' + esc(check.percentage) + '%)';
    }

    const itemsHtml = (check.failed_items && check.failed_items.length)
        ? renderItemsBlock(check.failed_items, 'Show affected items')
        : '';

    div.innerHTML =
        '<div class="check-icon">' + icon + '</div>' +
        '<div class="check-content">' +
        '<div class="check-title">' +
        '<span>' + esc(check.id) + ': ' + esc(check.name) + '</span> ' +
        '<span class="badge badge-' + esc(severity) + '">' + esc(severity) + '</span> ' +
        '<span class="badge badge-std">' + esc(check.standard) + '</span>' +
        '</div>' +
        '<div style="font-size:0.85rem;color:#64748b;margin-bottom:0.5rem;">' + esc(check.description) + '</div>' +
        '<div style="font-size:0.85rem;">' + details +
        (check.threshold ? ' | <strong>Threshold:</strong> ' + esc(check.threshold) : '') +
        '</div>' +
        (check.recommendation ? '<div class="recommendation-box">💡 ' + esc(check.recommendation) + '</div>' : '') +
        itemsHtml + '</div>';

    return div;
}

function applyFilter() { renderDetailedResults(); }

// ═══════════════════════════════════════════
// EXPORTS
// ═══════════════════════════════════════════

function downloadPDF(ev) {
    const severity = getSelectedSeverity();
    const btn = ev && ev.currentTarget ? ev.currentTarget : null;
    const url = '/api/executive-pdf?standard=' + encodeURIComponent(currentStandard) + '&severity=' + encodeURIComponent(severity);
    downloadFile(url, 'executive_report.pdf', btn);
}

function downloadActionsPDF(ev) {
    const severity = getSelectedSeverity();
    const btn = ev && ev.currentTarget ? ev.currentTarget : null;
    const url = '/api/actions-pdf?standard=' + encodeURIComponent(currentStandard) + '&severity=' + encodeURIComponent(severity);
    downloadFile(url, 'action_list.pdf', btn);
}

function downloadActionsExcel(ev) {
    const severity = getSelectedSeverity();
    const btn = ev && ev.currentTarget ? ev.currentTarget : null;
    const url = '/api/actions-excel?standard=' + encodeURIComponent(currentStandard) + '&severity=' + encodeURIComponent(severity);
    downloadFile(url, 'health_top_actions.xlsx', btn);
}

// ═══════════════════════════════════════════
// AI EXECUTIVE NARRATIVE
// ═══════════════════════════════════════════

async function fetchAINarrative(forceRefresh) {
    const body = document.getElementById('aiNarrativeBody');
    const methodEl = document.getElementById('aiNarrativeMethod');
    if (!body) return;

    body.innerHTML = '<p style="color:var(--color-muted);">⏳ Synthesizing executive briefing from health, EVM, and variance data...</p>';
    if (methodEl) methodEl.textContent = '';

    try {
        const res = await fetch('/api/ai-narrative');
        const data = await res.json().catch(function () { return {}; });

        if (!res.ok || data.error) {
            throw new Error(data.error || 'Failed to generate narrative');
        }

        const narrativeText = (data.data && data.data.narrative) || '';
        const method = (data.data && data.data.method) || 'Engine';

        if (methodEl) {
            methodEl.innerHTML = '<span>✨ Generated via: <strong>' + esc(method) + '</strong></span>';
        }

        // Convert basic markdown to HTML (with escape safety)
        let formattedHtml = esc(narrativeText)
            .replace(/^### (.*)$/gm, '<h3>$1</h3>')
            .replace(/\\*\\*(.*?)\\*\\*/g, '<strong>$1</strong>')
            .replace(/\\n\\n/g, '<br><br>')
            .replace(/\\n- /g, '<br>• ')
            .replace(/^- /gm, '• ');

        body.innerHTML = formattedHtml;
    } catch (err) {
        body.innerHTML = '<p style="color:var(--color-danger);">❌ ' + esc(err.message) + '</p>';
    }
}

// Back-compat exports
window.downloadPDF = downloadPDF;
window.downloadActionsPDF = downloadActionsPDF;
window.downloadActionsExcel = downloadActionsExcel;
window.selectStandard = selectStandard;
window.applyFilter = applyFilter;
window.loadHealthData = loadHealthData;
window.fetchAINarrative = fetchAINarrative;
'''

os.makedirs("static", exist_ok=True)
with open("static/health.js", "w", encoding="utf-8") as f:
    f.write(HEALTH_JS_CODE)
print("  ✅ Updated static/health.js")

print("\n🎉 Phase 1 - Step 1 (AI Executive Narrative) Applied Successfully!")
print("✨ Restart your Flask server (python app.py) and navigate to /health to see the new AI Narrative card!")