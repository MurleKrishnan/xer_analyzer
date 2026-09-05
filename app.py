"""
P6 SCHEDULE ANALYZER - MAIN WEB APPLICATION (app.py)
=====================================================
Unified application entry point including:
- Dashboard, Gantt, Comparison, EVM, Health, Trends
- AI Narrative, Longest Path, Resource Analytics, Activity Inspector
- Multi-Tenant Authentication, User Roles, Admin Panel, Database Persistence
"""

import os
import sys
import uuid
import time
import logging
from datetime import datetime

from flask import Flask, render_template, request, jsonify, send_file, session, redirect, url_for
from flask_cors import CORS
from werkzeug.utils import secure_filename

# ─── CONFIGURATION ───
try:
    from config import (
        get_config,
        MAX_UPLOAD_SIZE_MB,
        SECRET_KEY,
        SESSION_LIFETIME_HOURS,
    )
except ImportError:
    MAX_UPLOAD_SIZE_MB = 1000
    SECRET_KEY = 'dev-only-CHANGE-ME'
    SESSION_LIFETIME_HOURS = 24
    def get_config():
        return {
            'company_name': 'MK Constructions',
            'app_title': 'P6 Schedule Analyzer',
            'app_subtitle': 'DCMA 14-Point Check & Analytics',
            'use_logo_image': False,
            'theme': {'primary': '#1e40af', 'accent': '#3b82f6'},
            'features': {'gantt': True, 'comparison': True, 'evm': True, 'export': True, 'health': True}
        }

# ─── DATABASE & STORAGE ───
from database import init_db, cleanup_old_projects, get_db, Project
from project_service import ProjectService
from storage import get_storage

# ─── AUTHENTICATION ───
from auth_service import AuthService
from auth_decorators import login_required, role_required, get_current_user, get_current_org_id
import auth_models  # Registers SQLAlchemy models

# ─── PARSERS & CORE ENGINES ───
try:
    from universal_parser import UniversalParser
except ImportError:
    from parser import XERParser as UniversalParser

from data_engine import ScheduleEngine
from reports import ReportGenerator

# ─── LOGGING & OPTIONAL ENGINES ───
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

try:
    from comparison_engine import ScheduleComparator
    logger.info("✅ ScheduleComparator imported")
except Exception as e:
    ScheduleComparator = None
    logger.warning("❌ ScheduleComparator import failed: %s", e)

try:
    from evm_engine import EVMEngine
    logger.info("✅ EVMEngine imported")
except Exception as e:
    EVMEngine = None
    logger.warning("❌ EVMEngine import failed: %s", e)

try:
    from advanced_health_engine import AdvancedHealthEngine
    logger.info("✅ AdvancedHealthEngine imported")
except Exception as e:
    AdvancedHealthEngine = None
    logger.warning("❌ AdvancedHealthEngine import failed: %s", e)

try:
    from pdf_report_generator import PDFReportGenerator
    logger.info("✅ PDFReportGenerator imported")
except Exception as e:
    PDFReportGenerator = None
    logger.warning("❌ PDFReportGenerator import failed: %s", e)

try:
    from ai_narrative_engine import AINarrativeEngine
    logger.info("✅ AINarrativeEngine imported")
except Exception as e:
    AINarrativeEngine = None
    logger.warning("❌ AINarrativeEngine import failed: %s", e)

try:
    from resource_engine import ResourceEngine
    logger.info("✅ ResourceEngine imported")
except Exception as e:
    ResourceEngine = None
    logger.warning("❌ ResourceEngine import failed: %s", e)

try:
    from longest_path_engine import LongestPathEngine
    logger.info("✅ LongestPathEngine imported")
except Exception as e:
    LongestPathEngine = None
    logger.warning("❌ LongestPathEngine import failed: %s", e)

try:
    from trend_engine import TrendAnalysisEngine
    logger.info("✅ TrendAnalysisEngine imported")
except Exception as e:
    TrendAnalysisEngine = None
    logger.warning("❌ TrendAnalysisEngine import failed: %s", e)

try:
    from activity_detail_engine import ActivityDetailEngine
    logger.info("✅ ActivityDetailEngine imported")
except Exception as e:
    ActivityDetailEngine = None
    logger.warning("❌ ActivityDetailEngine import failed: %s", e)


# ─── INITIALIZE FLASK & DATABASE ───
app = Flask(__name__)
app.secret_key = SECRET_KEY
CORS(app)

UPLOAD_FOLDER = 'uploads'
OUTPUT_FOLDER = 'output'
ALLOWED_EXTENSIONS = {'xer', 'xml'}

app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['OUTPUT_FOLDER'] = OUTPUT_FOLDER
app.config['MAX_CONTENT_LENGTH'] = MAX_UPLOAD_SIZE_MB * 1024 * 1024

os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(OUTPUT_FOLDER, exist_ok=True)

# Initialize database schema on startup
init_db()
logger.info("✅ Database schema initialized")

# Cleanup old storage
cleanup_old_files = lambda folder, hours=24: None  # Fallback
try:
    def cleanup_old_files(folder, max_age_hours=SESSION_LIFETIME_HOURS):
        if not os.path.exists(folder):
            return
        cutoff = time.time() - (max_age_hours * 3600)
        for fname in os.listdir(folder):
            fpath = os.path.join(folder, fname)
            try:
                if os.path.isfile(fpath) and os.path.getmtime(fpath) < cutoff:
                    os.remove(fpath)
                    logger.info("🧹 Cleaned up old file: %s", fname)
            except Exception as e:
                logger.warning("Could not delete %s: %s", fname, e)
    cleanup_old_files(UPLOAD_FOLDER)
    cleanup_old_files(OUTPUT_FOLDER)
except Exception as e:
    logger.warning("Cleanup error: %s", e)


# ─── SESSION STORAGE ───
SESSION_STORAGE = {}

def get_session_data():
    sid = session.get('sid')
    if not sid or sid not in SESSION_STORAGE:
        sid = uuid.uuid4().hex
        session['sid'] = sid
        SESSION_STORAGE[sid] = {
            'analysis': {'engine': None, 'dashboard_data': None, 'file_name': None, 'analyzed_at': None},
            'comparison': {'comparator': None, 'results': None, 'baseline_file': None, 'current_file': None},
            'health_cache': {},
            'longest_path_cache': None,
            'trends': {'engine': None, 'results': None, 'periods': []},
        }
    return SESSION_STORAGE[sid]


def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


def analyze_xer_file(file_path_or_stream, original_filename, session_data):
    logger.info("🔍 Analyzing File: %s", original_filename)
    parser = UniversalParser()
    
    if hasattr(parser, 'parse') and parser.__class__.__name__ == 'UniversalParser':
        with open(file_path_or_stream, 'rb') as f:
            tables = parser.parse(f, original_filename)
    else:
        tables = parser.parse(file_path_or_stream)

    if tables is None or not tables:
        return {'error': 'Failed to parse schedule file or file is empty.'}

    engine = ScheduleEngine()
    engine.load_data(tables)
    engine.analyze()

    # Auto-compute Longest Path
    if LongestPathEngine is not None:
        try:
            lp_engine = LongestPathEngine(engine)
            lp_results = lp_engine.calculate()
            engine.longest_path_ids = lp_engine.longest_path_ids
            engine.longest_path_results = lp_results
        except Exception as e:
            logger.warning("Longest Path auto-calc failed: %s", e)
            engine.longest_path_ids = set()
            engine.longest_path_results = {}
    else:
        engine.longest_path_ids = set()

    dashboard_data = engine.get_dashboard_data()

    analysis = session_data['analysis']
    analysis['engine'] = engine
    analysis['dashboard_data'] = dashboard_data
    analysis['file_name'] = original_filename
    analysis['analyzed_at'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    session_data['health_cache'] = {}
    session_data['longest_path_cache'] = None

    return dashboard_data


@app.context_processor
def inject_config():
    return {'config': get_config()}


@app.errorhandler(413)
def request_entity_too_large(error):
    return jsonify({'error': f'File size exceeds maximum limit of {MAX_UPLOAD_SIZE_MB} MB.'}), 413


# ════════════════════════════════════════════
# ROUTE 1: DASHBOARD
# ════════════════════════════════════════════

@app.route('/')
def home():
    return render_template('index.html')


@app.route('/api/upload', methods=['POST'])
def upload_file():
    if 'file' not in request.files:
        return jsonify({'error': 'No file provided'}), 400

    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'No file selected'}), 400

    if not allowed_file(file.filename):
        return jsonify({'error': 'File must be a .xer or .xml file'}), 400

    original_name = secure_filename(file.filename)
    unique_name = f"{uuid.uuid4().hex}_{original_name}"
    file_path = os.path.join(app.config['UPLOAD_FOLDER'], unique_name)
    file.save(file_path)

    sess_data = get_session_data()

    try:
        dashboard_data = analyze_xer_file(file_path, original_name, sess_data)
        if 'error' in dashboard_data:
            return jsonify(dashboard_data), 400

        analysis = sess_data['analysis']
        return jsonify({
            'success': True,
            'file_name': analysis['file_name'],
            'analyzed_at': analysis['analyzed_at'],
            'data': dashboard_data
        })
    except Exception as e:
        logger.exception("Upload analysis error")
        return jsonify({'error': str(e)}), 500


@app.route('/api/dashboard')
def get_dashboard():
    sess_data = get_session_data()
    analysis = sess_data['analysis']

    if analysis['dashboard_data'] is None:
        return jsonify({'has_data': False})

    return jsonify({
        'has_data': True,
        'file_name': analysis['file_name'],
        'analyzed_at': analysis['analyzed_at'],
        'data': analysis['dashboard_data']
    })


@app.route('/api/load-sample')
def load_sample():
    sample_path = os.path.join('input', 'sample.xer')
    if not os.path.exists(sample_path):
        return jsonify({'error': 'sample.xer not found in input/ folder'}), 404

    sess_data = get_session_data()
    try:
        dashboard_data = analyze_xer_file(sample_path, 'sample.xer', sess_data)
        analysis = sess_data['analysis']
        return jsonify({
            'success': True,
            'file_name': 'sample.xer',
            'analyzed_at': analysis['analyzed_at'],
            'data': dashboard_data
        })
    except Exception as e:
        logger.exception("Sample load error")
        return jsonify({'error': str(e)}), 500


@app.route('/api/export-excel')
def export_excel():
    sess_data = get_session_data()
    engine = sess_data['analysis']['engine']

    if engine is None:
        return jsonify({'error': 'No schedule loaded. Please upload a schedule file first.'}), 400

    out_name = f"schedule_report_{uuid.uuid4().hex[:8]}.xlsx"
    output_path = os.path.join(app.config['OUTPUT_FOLDER'], out_name)
    
    reporter = ReportGenerator(engine)
    res_path = reporter.generate_full_report(output_path)

    if not res_path or not os.path.exists(res_path):
        return jsonify({'error': 'Failed to generate Excel report.'}), 500

    return send_file(
        res_path,
        as_attachment=True,
        download_name=f"schedule_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


# ════════════════════════════════════════════
# ROUTE 2: GANTT CHART & LONGEST PATH
# ════════════════════════════════════════════

@app.route('/gantt')
def gantt_view():
    return render_template('gantt.html')


@app.route('/api/gantt-data')
def get_gantt_data():
    sess_data = get_session_data()
    analysis = sess_data['analysis']

    if analysis['engine'] is None:
        return jsonify({'error': 'No data loaded. Please upload a schedule file first.'}), 400

    max_acts = request.args.get('max', 2000, type=int)
    gantt_data = analysis['engine'].get_gantt_data(max_activities=max_acts)

    return jsonify({
        'success': True,
        'file_name': analysis['file_name'],
        'data': gantt_data
    })


@app.route('/api/longest-path')
def get_longest_path():
    sess_data = get_session_data()
    analysis = sess_data['analysis']

    if analysis['engine'] is None:
        return jsonify({'error': 'No data loaded. Upload a schedule file first.'}), 400

    if LongestPathEngine is None:
        return jsonify({'error': 'longest_path_engine.py is missing!'}), 500

    if sess_data.get('longest_path_cache'):
        return jsonify({
            'success': True,
            'file_name': analysis['file_name'],
            'data': sess_data['longest_path_cache'],
            'cached': True
        })

    try:
        lp_engine = LongestPathEngine(analysis['engine'])
        results = lp_engine.calculate()
        sess_data['longest_path_cache'] = results
        analysis['engine'].longest_path_ids = lp_engine.longest_path_ids
        return jsonify({
            'success': True,
            'file_name': analysis['file_name'],
            'data': results
        })
    except Exception as e:
        logger.exception("Longest Path calculation error")
        return jsonify({'error': str(e)}), 500


# ════════════════════════════════════════════
# ROUTE 3: COMPARISON (Baseline vs Current)
# ════════════════════════════════════════════

@app.route('/comparison')
def comparison_view():
    return render_template('comparison.html')


@app.route('/api/compare', methods=['POST'])
def compare_schedules():
    if ScheduleComparator is None:
        return jsonify({'error': 'comparison_engine.py is missing!'}), 500

    if 'baseline' not in request.files or 'current' not in request.files:
        return jsonify({'error': 'Both baseline and current files required'}), 400

    baseline_file = request.files['baseline']
    current_file = request.files['current']

    if not (allowed_file(baseline_file.filename) and allowed_file(current_file.filename)):
        return jsonify({'error': 'Both files must be .xer or .xml files'}), 400

    baseline_name = secure_filename(baseline_file.filename)
    current_name = secure_filename(current_file.filename)

    baseline_path = os.path.join(app.config['UPLOAD_FOLDER'], f"bl_{uuid.uuid4().hex[:8]}_{baseline_name}")
    current_path = os.path.join(app.config['UPLOAD_FOLDER'], f"cur_{uuid.uuid4().hex[:8]}_{current_name}")

    baseline_file.save(baseline_path)
    current_file.save(current_path)

    sess_data = get_session_data()

    try:
        comparator = ScheduleComparator()
        comparator.load_baseline(baseline_path)
        comparator.load_current(current_path)
        results = comparator.compare()

        comp = sess_data['comparison']
        comp['comparator'] = comparator
        comp['results'] = results
        comp['baseline_file'] = baseline_name
        comp['current_file'] = current_name

        return jsonify({
            'success': True,
            'baseline_file': baseline_name,
            'current_file': current_name,
            'results': results
        })
    except Exception as e:
        logger.exception("Comparison error")
        return jsonify({'error': str(e)}), 500


@app.route('/api/comparison-data')
def get_comparison_data():
    sess_data = get_session_data()
    comp = sess_data['comparison']

    if comp['results'] is None:
        return jsonify({'has_data': False})

    return jsonify({
        'has_data': True,
        'baseline_file': comp['baseline_file'],
        'current_file': comp['current_file'],
        'results': comp['results']
    })


# ════════════════════════════════════════════
# ROUTE 4: EVM & RESOURCE ANALYTICS
# ════════════════════════════════════════════

@app.route('/evm')
def evm_view():
    return render_template('evm.html')


@app.route('/api/evm-data')
def get_evm_data():
    sess_data = get_session_data()
    analysis = sess_data['analysis']

    if analysis['engine'] is None:
        return jsonify({'error': 'No data loaded. Upload a schedule file first.'}), 400

    if EVMEngine is None:
        return jsonify({'error': 'evm_engine.py is missing!'}), 500

    try:
        evm = EVMEngine(analysis['engine'])
        results = evm.calculate()
        return jsonify({
            'success': True,
            'file_name': analysis['file_name'],
            'data': results
        })
    except Exception as e:
        logger.exception("EVM calculation error")
        return jsonify({'error': str(e)}), 500


@app.route('/api/resource-data')
def get_resource_data():
    sess_data = get_session_data()
    analysis = sess_data['analysis']

    if analysis['engine'] is None:
        return jsonify({'error': 'No data loaded. Upload a schedule file first.'}), 400

    if ResourceEngine is None:
        return jsonify({'error': 'resource_engine.py is missing!'}), 500

    try:
        engine = ResourceEngine(analysis['engine'])
        results = engine.calculate()
        return jsonify({
            'success': True,
            'file_name': analysis['file_name'],
            'data': results
        })
    except Exception as e:
        logger.exception("Resource analytics error")
        return jsonify({'error': str(e)}), 500


# ════════════════════════════════════════════
# ROUTE 5: MULTI-PERIOD TRENDS
# ════════════════════════════════════════════

@app.route('/trends')
def trends_view():
    return render_template('trends.html')


@app.route('/api/trend-upload', methods=['POST'])
def upload_trend_files():
    if TrendAnalysisEngine is None:
        return jsonify({'error': 'trend_engine.py is missing!'}), 500

    files = request.files.getlist('files')
    if not files or len(files) < 2:
        return jsonify({'error': 'At least 2 schedule files required for trend analysis'}), 400

    for f in files:
        if not allowed_file(f.filename):
            return jsonify({'error': f'File {f.filename} is not a .xer or .xml file'}), 400

    sess_data = get_session_data()
    trend_engine = TrendAnalysisEngine()

    saved_files = []
    try:
        for file in files:
            original_name = secure_filename(file.filename)
            unique_name = f"trend_{uuid.uuid4().hex[:8]}_{original_name}"
            fpath = os.path.join(app.config['UPLOAD_FOLDER'], unique_name)
            file.save(fpath)
            saved_files.append({'path': fpath, 'name': original_name})

            label = original_name.replace('.xer', '').replace('.XER', '').replace('.xml', '').replace('.XML', '')
            trend_engine.add_period(fpath, period_label=label)

        results = trend_engine.analyze()

        sess_data['trends']['engine'] = trend_engine
        sess_data['trends']['results'] = results
        sess_data['trends']['periods'] = [f['name'] for f in saved_files]

        return jsonify({
            'success': True,
            'period_count': len(saved_files),
            'periods': [f['name'] for f in saved_files],
            'data': results,
        })
    except Exception as e:
        logger.exception("Trend upload error")
        return jsonify({'error': str(e)}), 500


@app.route('/api/trend-data')
def get_trend_data():
    sess_data = get_session_data()
    trends = sess_data.get('trends', {})
    if not trends.get('results'):
        return jsonify({'has_data': False})
    return jsonify({
        'has_data': True,
        'periods': trends['periods'],
        'data': trends['results'],
    })


@app.route('/api/trend-reset', methods=['POST'])
def reset_trend_analysis():
    sess_data = get_session_data()
    sess_data['trends'] = {'engine': None, 'results': None, 'periods': []}
    return jsonify({'success': True})


# ════════════════════════════════════════════
# ROUTE 6: HEALTH ANALYTICS & REPORTS
# ════════════════════════════════════════════

@app.route('/health')
def health_view():
    return render_template('health.html')


@app.route('/api/health-data')
def get_health_data():
    sess_data = get_session_data()
    analysis = sess_data['analysis']

    if analysis['engine'] is None:
        return jsonify({'error': 'No data loaded. Upload a file first.'}), 400

    if AdvancedHealthEngine is None:
        return jsonify({'error': 'advanced_health_engine.py is missing!'}), 500

    selected_standard = request.args.get('standard', 'all')

    try:
        health = AdvancedHealthEngine(analysis['engine'])
        results = health.run_all_checks(selected_standard=selected_standard)

        return jsonify({
            'success': True,
            'file_name': analysis['file_name'],
            'data': results
        })
    except Exception as e:
        logger.exception("Health analysis error")
        return jsonify({'error': str(e)}), 500


@app.route('/api/executive-pdf')
def download_executive_pdf():
    sess_data = get_session_data()
    analysis = sess_data['analysis']

    if analysis['engine'] is None:
        return jsonify({'error': 'No data loaded'}), 400

    if PDFReportGenerator is None or AdvancedHealthEngine is None:
        return jsonify({'error': 'PDF or Health engine module missing!'}), 500

    selected_standard = request.args.get('standard', 'all')
    severity_filter = request.args.get('severity', 'all')

    try:
        health = AdvancedHealthEngine(analysis['engine'])
        results = health.run_all_checks(selected_standard=selected_standard)

        generator = PDFReportGenerator(
            results,
            analysis['file_name'],
            severity_filter=severity_filter
        )
        pdf_buffer = generator.generate_executive_report()

        return send_file(
            pdf_buffer,
            as_attachment=True,
            download_name=f"executive_report_{selected_standard}_{severity_filter}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf",
            mimetype='application/pdf'
        )
    except Exception as e:
        logger.exception("PDF generation error")
        return jsonify({'error': str(e)}), 500


@app.route('/api/actions-pdf')
def download_actions_pdf():
    sess_data = get_session_data()
    analysis = sess_data['analysis']

    if analysis['engine'] is None:
        return jsonify({'error': 'No data loaded'}), 400

    if PDFReportGenerator is None or AdvancedHealthEngine is None:
        return jsonify({'error': 'PDF or Health engine module missing!'}), 500

    selected_standard = request.args.get('standard', 'all')
    severity_filter = request.args.get('severity', 'all')

    try:
        health = AdvancedHealthEngine(analysis['engine'])
        results = health.run_all_checks(selected_standard=selected_standard)

        generator = PDFReportGenerator(
            results,
            analysis['file_name'],
            severity_filter=severity_filter
        )
        pdf_buffer = generator.generate_actions_report()

        return send_file(
            pdf_buffer,
            as_attachment=True,
            download_name=f"action_list_{selected_standard}_{severity_filter}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf",
            mimetype='application/pdf'
        )
    except Exception as e:
        logger.exception("PDF actions error")
        return jsonify({'error': str(e)}), 500


@app.route('/api/actions-excel')
def download_actions_excel():
    sess_data = get_session_data()
    analysis = sess_data['analysis']

    if analysis['engine'] is None:
        return jsonify({'error': 'No data loaded. Upload a file first.'}), 400

    if AdvancedHealthEngine is None:
        return jsonify({'error': 'advanced_health_engine.py is missing!'}), 500

    selected_standard = request.args.get('standard', 'all')
    severity_filter = request.args.get('severity', 'all').lower()

    severity_levels = {
        'critical': ['critical'],
        'high': ['critical', 'high'],
        'medium': ['critical', 'high', 'medium'],
        'all': ['critical', 'high', 'medium', 'low', 'info']
    }
    allowed_severities = severity_levels.get(severity_filter, severity_levels['all'])

    try:
        import pandas as pd
        from io import BytesIO
        from openpyxl.styles import Font, PatternFill, Alignment

        health = AdvancedHealthEngine(analysis['engine'])
        results = health.run_all_checks(selected_standard=selected_standard)

        top_actions = results.get('top_actions', []) or []
        standards_data = results.get('standards', {}) or {}

        filtered_top_actions = [
            a for a in top_actions
            if (a.get('severity') or 'low').lower() in allowed_severities
        ]

        meta_rows = [
            ['Report Type', 'Schedule Health - Top Actions Export'],
            ['Selected Standard', selected_standard],
            ['Severity Filter', severity_filter.upper()],
            ['File Name', analysis.get('file_name', '')],
            ['Generated At', results.get('analysis_date', '')],
            ['', ''],
            ['Overall Score', results.get('overall_score', '')],
            ['Total Checks', results.get('total_checks', '')],
            ['Passed Checks', results.get('passed_checks', '')],
            ['Failed Checks', results.get('failed_checks', '')],
            ['Critical Failures', results.get('critical_failures', '')],
            ['High Failures', results.get('high_failures', '')],
            ['Pass Rate (%)', results.get('pass_rate', '')],
            ['', ''],
            ['Filtered Actions Count', len(filtered_top_actions)],
        ]

        top_summary_rows = []
        for idx, action in enumerate(filtered_top_actions, 1):
            top_summary_rows.append({
                'Rank': idx,
                'Standard': action.get('standard', ''),
                'Check ID': action.get('id', ''),
                'Check Name': action.get('name', ''),
                'Category': action.get('category', ''),
                'Severity': (action.get('severity') or '').upper(),
                'Affected Count': action.get('count', 0),
                'Total': action.get('total', 0),
                'Percentage': action.get('percentage', 0),
                'Threshold': action.get('threshold', ''),
                'Value': action.get('value', ''),
                'Recommendation': action.get('recommendation', ''),
                'Description': action.get('description', ''),
            })

        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            pd.DataFrame(meta_rows, columns=['Field', 'Value']).to_excel(
                writer, sheet_name='Report Info', index=False
            )

            if top_summary_rows:
                pd.DataFrame(top_summary_rows).to_excel(
                    writer, sheet_name='Top Actions Summary', index=False
                )
            else:
                pd.DataFrame([{'Info': f'No actions matched severity filter: {severity_filter.upper()}'}]).to_excel(
                    writer, sheet_name='Top Actions Summary', index=False
                )

            for std_name, std_data in standards_data.items():
                sheet_name = (std_name or 'Standard')[:31]
                rows = []

                failed_in_std = []
                for category in std_data.get('categories', []):
                    for check in category.get('checks', []):
                        if check.get('status') != 'fail':
                            continue
                        if (check.get('severity') or 'low').lower() not in allowed_severities:
                            continue
                        failed_in_std.append((category.get('name', ''), check))

                if not failed_in_std:
                    rows.append({
                        'Section': 'No Matching Failures',
                        'Field': '',
                        'Value': f'No {severity_filter.upper()} failures found for {std_name}',
                        'Activity ID': '',
                        'Activity Name': '',
                        'WBS': '',
                    })
                else:
                    for cat_name, check in failed_in_std:
                        rows.append({'Section': 'METRIC', 'Field': 'Check ID', 'Value': check.get('id', ''), 'Activity ID': '', 'Activity Name': '', 'WBS': ''})
                        rows.append({'Section': '', 'Field': 'Check Name', 'Value': check.get('name', ''), 'Activity ID': '', 'Activity Name': '', 'WBS': ''})
                        rows.append({'Section': '', 'Field': 'Category', 'Value': cat_name, 'Activity ID': '', 'Activity Name': '', 'WBS': ''})
                        rows.append({'Section': '', 'Field': 'Severity', 'Value': (check.get('severity') or '').upper(), 'Activity ID': '', 'Activity Name': '', 'WBS': ''})
                        rows.append({'Section': '', 'Field': 'Description', 'Value': check.get('description', ''), 'Activity ID': '', 'Activity Name': '', 'WBS': ''})
                        rows.append({'Section': '', 'Field': 'Threshold', 'Value': check.get('threshold', ''), 'Activity ID': '', 'Activity Name': '', 'WBS': ''})

                        if check.get('count') is not None:
                            rows.append({
                                'Section': '', 'Field': 'Affected',
                                'Value': f"{check.get('count', 0)} of {check.get('total', 0)} ({check.get('percentage', 0)}%)",
                                'Activity ID': '', 'Activity Name': '', 'WBS': '',
                            })
                        if check.get('value') is not None and check.get('value') != '':
                            rows.append({
                                'Section': '', 'Field': 'Value', 'Value': check.get('value', ''),
                                'Activity ID': '', 'Activity Name': '', 'WBS': '',
                            })

                        rows.append({'Section': '', 'Field': 'Recommendation', 'Value': check.get('recommendation', ''), 'Activity ID': '', 'Activity Name': '', 'WBS': ''})

                        items = check.get('failed_items', []) or []
                        rows.append({'Section': 'AFFECTED ACTIVITIES', 'Field': f'Total: {len(items)}', 'Value': '', 'Activity ID': 'Activity ID', 'Activity Name': 'Activity Name', 'WBS': 'WBS'})

                        if items:
                            for item in items:
                                rows.append({
                                    'Section': '', 'Field': '', 'Value': '',
                                    'Activity ID': item.get('code', ''),
                                    'Activity Name': item.get('name', ''),
                                    'WBS': item.get('wbs', ''),
                                })
                        else:
                            rows.append({'Section': '', 'Field': '', 'Value': '(No activity list available)', 'Activity ID': '', 'Activity Name': '', 'WBS': ''})

                        rows.append({'Section': '', 'Field': '', 'Value': '', 'Activity ID': '', 'Activity Name': '', 'WBS': ''})

                pd.DataFrame(rows).to_excel(writer, sheet_name=sheet_name, index=False)

            workbook = writer.book
            header_font = Font(bold=True, color='FFFFFF', size=11)
            header_fill = PatternFill(start_color='1E40AF', end_color='1E40AF', fill_type='solid')
            metric_fill = PatternFill(start_color='FEF3C7', end_color='FEF3C7', fill_type='solid')
            activity_hdr_fill = PatternFill(start_color='DBEAFE', end_color='DBEAFE', fill_type='solid')
            bold_font = Font(bold=True)
            wrap_align = Alignment(wrap_text=True, vertical='top')

            for sheet_name in workbook.sheetnames:
                ws = workbook[sheet_name]
                for cell in ws[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal='left', vertical='center')

                for col in ws.columns:
                    max_len = 10
                    col_letter = col[0].column_letter
                    for cell in col:
                        try:
                            val = str(cell.value) if cell.value is not None else ''
                            if len(val) > max_len: max_len = len(val)
                        except Exception: pass
                    ws.column_dimensions[col_letter].width = min(max_len + 2, 60)

                ws.freeze_panes = 'A2'

                if sheet_name not in ['Report Info', 'Top Actions Summary']:
                    for row in ws.iter_rows(min_row=2):
                        section_cell = row[0]
                        if section_cell.value == 'METRIC':
                            for c in row: c.fill = metric_fill; c.font = bold_font
                        elif section_cell.value == 'AFFECTED ACTIVITIES':
                            for c in row: c.fill = activity_hdr_fill; c.font = bold_font
                        for c in row: c.alignment = wrap_align

        output.seek(0)
        filename = f"health_top_actions_{selected_standard}_{severity_filter}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    except Exception as e:
        logger.exception("Actions Excel error")
        return jsonify({'error': str(e)}), 500


@app.route('/api/ai-narrative')
def get_ai_narrative():
    sess_data = get_session_data()
    analysis = sess_data['analysis']

    if analysis['engine'] is None:
        return jsonify({'error': 'No schedule loaded. Upload an XER file first.'}), 400

    if AINarrativeEngine is None:
        return jsonify({'error': 'ai_narrative_engine.py is missing!'}), 500

    try:
        health_data = {}
        if AdvancedHealthEngine is not None:
            try:
                health_gen = AdvancedHealthEngine(analysis['engine'])
                health_data = health_gen.run_all_checks('all')
            except Exception as he:
                logger.warning("Health data unavailable for narrative: %s", he)

        evm_data = {}
        if EVMEngine is not None:
            try:
                evm_data = EVMEngine(analysis['engine']).calculate()
            except Exception: pass

        comp_data = sess_data.get('comparison', {}).get('results', {}) or {}

        narrative_gen = AINarrativeEngine(
            health_data=health_data,
            comparison_data=comp_data,
            evm_data=evm_data
        )
        results = narrative_gen.generate_narrative()

        return jsonify({
            'success': True,
            'file_name': analysis['file_name'],
            'data': results
        })
    except Exception as e:
        logger.exception("AI Narrative generation error")
        return jsonify({'error': str(e)}), 500


@app.route('/api/activity-detail/<path:activity_code>')
def get_activity_detail(activity_code):
    sess_data = get_session_data()
    analysis = sess_data['analysis']

    if analysis['engine'] is None:
        return jsonify({'error': 'No schedule loaded. Upload an XER file first.'}), 400

    if ActivityDetailEngine is None:
        return jsonify({'error': 'activity_detail_engine.py is missing!'}), 500

    try:
        detail_engine = ActivityDetailEngine(analysis['engine'])
        result = detail_engine.get_detail(activity_code)
        if 'error' in result:
            return jsonify(result), 404
        return jsonify({'success': True, 'activity_code': activity_code, 'data': result})
    except Exception as e:
        logger.exception("Activity detail error")
        return jsonify({'error': str(e)}), 500


# ═══════════════════════════════════════════
# ROUTE 7: PROJECT PERSISTENCE
# ═══════════════════════════════════════════

@app.route('/api/projects', methods=['GET'])
def list_user_projects():
    sess_id = session.get('sid')
    if not sess_id:
        return jsonify({'projects': []})
    service = ProjectService(sess_id)
    return jsonify({'projects': service.list_projects()})


@app.route('/api/projects/<int:project_id>', methods=['DELETE'])
def delete_user_project(project_id):
    sess_id = session.get('sid')
    if not sess_id:
        return jsonify({'error': 'Not authenticated'}), 401
    service = ProjectService(sess_id)
    ok = service.delete_project(project_id)
    if not ok:
        return jsonify({'error': 'Project not found'}), 404
    return jsonify({'success': True})


@app.route('/api/projects/<int:project_id>/activate', methods=['POST'])
def activate_project(project_id):
    sess_id = session.get('sid')
    if not sess_id:
        return jsonify({'error': 'Not authenticated'}), 401

    service = ProjectService(sess_id)
    project = service.get_project(project_id)
    if not project:
        return jsonify({'error': 'Project not found'}), 404

    engine = service.get_engine(project_id)
    if not engine:
        return jsonify({'error': 'Failed to load project engine'}), 500

    sess_data = get_session_data()
    sess_data['analysis']['engine'] = engine
    sess_data['analysis']['dashboard_data'] = engine.get_dashboard_data()
    sess_data['analysis']['file_name'] = project['file_name']
    sess_data['analysis']['analyzed_at'] = project.get('processed_at', '')
    sess_data['analysis']['project_id'] = project_id
    sess_data['health_cache'] = {}
    sess_data['longest_path_cache'] = None

    return jsonify({
        'success': True,
        'project': project,
        'data': sess_data['analysis']['dashboard_data']
    })


# ═══════════════════════════════════════════
# ROUTE 8: AUTHENTICATION & SAAS MANAGEMENT
# ═══════════════════════════════════════════

@app.route('/login')
def login_page():
    return render_template('auth/login.html')


@app.route('/register')
def register_page():
    return render_template('auth/register.html')


@app.route('/admin')
@role_required('owner', 'admin')
def admin_page():
    return render_template('auth/admin.html')


@app.route('/logout')
def logout():
    session.pop('user', None)
    session.pop('sid', None)
    return redirect('/login')


@app.route('/api/auth/register', methods=['POST'])
def api_register():
    data = request.get_json() or {}
    user, err = AuthService.register_new_org(
        org_name=data.get('org_name', '').strip(),
        email=data.get('email', '').strip(),
        password=data.get('password', ''),
        full_name=data.get('full_name', '').strip(),
    )
    if err:
        return jsonify({'error': err}), 400
    session['user'] = user
    session['sid'] = str(user['id'])
    return jsonify({'success': True, 'user': user})


@app.route('/api/auth/login', methods=['POST'])
def api_login():
    data = request.get_json() or {}
    user, err = AuthService.login(
        email=data.get('email', '').strip(),
        password=data.get('password', ''),
    )
    if err:
        return jsonify({'error': err}), 401
    session['user'] = user
    session['sid'] = str(user['id'])
    return jsonify({'success': True, 'user': user})


@app.route('/api/auth/me')
def api_current_user():
    user = get_current_user()
    if not user:
        return jsonify({'user': None, 'authenticated': False})
    return jsonify({'user': user, 'authenticated': True})


@app.route('/api/auth/users')
@role_required('owner', 'admin')
def api_list_users():
    org_id = get_current_org_id()
    users = AuthService.list_org_users(org_id)
    return jsonify({'users': users})


@app.route('/api/auth/users/<int:user_id>/role', methods=['POST'])
@role_required('owner', 'admin')
def api_change_role(user_id):
    data = request.get_json() or {}
    new_role = data.get('role')
    actor = get_current_user()
    ok, err = AuthService.update_user_role(user_id, new_role, actor['id'])
    if not ok:
        return jsonify({'error': err}), 400
    return jsonify({'success': True})


@app.route('/api/auth/users/<int:user_id>/deactivate', methods=['POST'])
@role_required('owner', 'admin')
def api_deactivate_user(user_id):
    actor = get_current_user()
    ok, err = AuthService.deactivate_user(user_id, actor['id'])
    if not ok:
        return jsonify({'error': err}), 400
    return jsonify({'success': True})


@app.route('/api/auth/invite', methods=['POST'])
@role_required('owner', 'admin')
def api_invite():
    data = request.get_json() or {}
    actor = get_current_user()
    inv, err = AuthService.create_invitation(
        org_id=actor['org_id'],
        email=data.get('email', '').strip(),
        role=data.get('role', 'viewer'),
        inviter_user_id=actor['id'],
    )
    if err:
        return jsonify({'error': err}), 400
    return jsonify({'success': True, 'invitation': inv})


@app.route('/accept-invite')
def accept_invite_page():
    token = request.args.get('token', '')
    return render_template('auth/register.html', invite_token=token)


@app.route('/api/auth/accept-invite', methods=['POST'])
def api_accept_invite():
    data = request.get_json() or {}
    user, err = AuthService.accept_invitation(
        token=data.get('token', ''),
        password=data.get('password', ''),
        full_name=data.get('full_name', ''),
    )
    if err:
        return jsonify({'error': err}), 400
    session['user'] = user
    session['sid'] = str(user['id'])
    return jsonify({'success': True, 'user': user})


if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    debug_mode = os.environ.get('FLASK_ENV') != 'production'

    logger.info("============================================================")
    logger.info("🚀 P6 SCHEDULE ANALYZER - UNIFIED ENTERPRISE EDITION")
    logger.info("============================================================")
    logger.info("📌 Running on port %s", port)
    logger.info("🔧 Debug mode: %s", debug_mode)
    logger.info("👉 Open in browser: http://localhost:%s", port)

    app.run(debug=debug_mode, host='0.0.0.0', port=port)
